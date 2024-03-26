"""
This module contains utility functions and classes for the MONSCE project.

Author: "Benkirane Ismail"
Email: "ibenkirane@mgb.org"
version: "1.0.0"
Date: 2023-10-19
"""

import re
import os
import cv2
import ast
import json
import math
import dlib
import torch
import random
import string
import logging
import librosa
import openpyxl
from itertools import combinations
from textstat.textstat import textstat
from collections import Counter

from git import Repo
from git.exc import GitCommandError

import numpy as np
import pandas as pd
import seaborn as sns
from tqdm import tqdm
import mediapipe as mp
import noisereduce as nr
import ipywidgets as widgets

import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from matplotlib.patches import Ellipse
from matplotlib_venn import venn2, venn3
from matplotlib.colors import LinearSegmentedColormap

from docx import Document
from textblob import TextBlob
from moviepy.editor import VideoFileClip
from datetime import datetime, timedelta
from transformers import BertTokenizer, BertModel
from matplotlib.backends.backend_pdf import PdfPages

from scipy.ndimage import median_filter
from scipy.stats import pearsonr, spearmanr, kendalltau, ttest_1samp, wilcoxon
from scipy.signal import find_peaks, savgol_filter, resample, butter, filtfilt, lfilter

import nltk
from nltk.tag import pos_tag
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from nltk.tokenize import word_tokenize
from nltk.sentiment import SentimentIntensityAnalyzer
from nltk.tokenize import word_tokenize, sent_tokenize

from sklearn.svm import SVC
from sklearn.cluster import KMeans
from sklearn.manifold import TSNE,MDS
from sklearn.tree import DecisionTreeClassifier
from sklearn.feature_selection import RFE, RFECV
from sklearn.metrics import silhouette_score, accuracy_score, confusion_matrix
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.ensemble import RandomForestClassifier, VotingClassifier, AdaBoostClassifier, StackingClassifier
from sklearn.model_selection import StratifiedKFold, train_test_split
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.decomposition import LatentDirichletAllocation, PCA

from xgboost import XGBClassifier
        
from imblearn.over_sampling import SMOTE

from mlxtend.feature_selection import SequentialFeatureSelector

from tensorflow.keras.utils import to_categorical

class EMPATICA:
    
    def __init__(self, sample_rate_to_use = 64):
        """
        Initialize the Empatica class and fetch the necessary .

        Parameters
        ----------
        sample_rate_to_use : int
            The sample rate to use.

        """

        if os.path.exists("metadata/empatica_metadata.json"):
            with open("metadata/empatica_metadata.json", 'r') as json_file:
                self.metadata = json.load(json_file)
        elif os.path.exists("../metadata/empatica_metadata.json"):
            with open("../metadata/empatica_metadata.json", 'r') as json_file:
                self.metadata = json.load(json_file)
        else:
            print("Empatica metadata file not found.")
            return

        self.sample_rate_to_use = sample_rate_to_use
        
        self.utilities = UTILITIES()
    
    def get_data(self, folder, verbose = False):
        """
        Get the empatica data from the csv files.
        
        Parameters
        ----------
        folder : str
            The path to the folder containing the data.
        verbose : bool
            Whether to show messages.
        """

        data = dict()

        acc_data_path = os.path.join(folder, 'ACC.csv')
        bvp_data_path = os.path.join(folder, 'BVP.csv')
        eda_data_path = os.path.join(folder, 'EDA.csv')
        hr_data_path = os.path.join(folder, 'HR.csv')
        temp_data_path = os.path.join(folder, 'TEMP.csv')
        tags_data_path = os.path.join(folder, 'tags.csv')

        if os.path.getsize(acc_data_path) != 0:
            data["acc"] = self.metadata["acc"].copy()
            data["acc"]['initial_time'], data["acc"]['sample_rate'], data["acc"]['data'] = self._load_empatica_csv(acc_data_path)
            data["acc"]['data'].columns = ['accX', 'accY', 'accZ']
            data['acc']['data']['Time (s)'] = self.utilities._construct_time(data["acc"]['sample_rate'], len(data["acc"]['data']))    
        else:
            if verbose:
                print(acc_data_path, "is empty.")
            
        if os.path.getsize(bvp_data_path) != 0:
            data['bvp'] = self.metadata['bvp'].copy()
            data["bvp"]['initial_time'], data["bvp"]['sample_rate'], data["bvp"]['data'] = self._load_empatica_csv(bvp_data_path)
            data["bvp"]['data'].columns = ['BVP Value']
            data['bvp']['data']['Time (s)'] = self.utilities._construct_time(data["bvp"]['sample_rate'], len(data["bvp"]['data'])) 
        else:   
            if verbose:
                print(bvp_data_path, "is empty.")   

        if os.path.getsize(eda_data_path) != 0:
            data["eda"] = self.metadata["eda"].copy()
            data["eda"]['initial_time'], data["eda"]['sample_rate'], data["eda"]['data'] = self._load_empatica_csv(eda_data_path)
            data["eda"]['data'].columns = ['EDA (μS)']
            data['eda']['data']['Time (s)'] = self.utilities._construct_time(data["eda"]['sample_rate'], len(data["eda"]['data']))
        else:
            if verbose:
                print(eda_data_path, "is empty.")
            
        if os.path.getsize(hr_data_path) != 0:
            data["hr"] = self.metadata["hr"].copy()
            data["hr"]['initial_time'], data["hr"]['sample_rate'], data["hr"]['data'] = self._load_empatica_csv(hr_data_path)
            data["hr"]['data'].columns = ['Heart Rate (bpm)']
            data['hr']['data']['Time (s)'] = self.utilities._construct_time(data["hr"]['sample_rate'], len(data["hr"]['data']))
        else:
            if verbose:
                print(hr_data_path, "is empty.")

        if os.path.getsize(temp_data_path) != 0:
            data["temp"] = self.metadata["temp"].copy()
            data["temp"]['initial_time'], data["temp"]['sample_rate'], data["temp"]['data'] = self._load_empatica_csv(temp_data_path)
            data["temp"]['data'].columns = ['Temperature (°C)']
            data['temp']['data']['Time (s)'] = self.utilities._construct_time(data["temp"]['sample_rate'], len(data["temp"]['data']))
        else:
            if verbose:
                print(temp_data_path, "is empty.")
            
        if os.path.getsize(tags_data_path) != 0:
            tags_df = pd.read_csv(tags_data_path, header=None)  
            tags = self.utilities._convert_unix_timestamp_to_utc(float(tags_df[0][0]))
            return data, tags
        else:
            if verbose:
                print(tags_data_path, "is empty.")
            return data, None     
            
    def label_data(self, data, emotion_timing, empatica_shift, include_neutral, verbose=False, increase_window = False):
        """
        Label the empatica data based on the timings.
        
        Parameters
        ----------
        data : dict
            The dictionary that contains all data.
        emotion_timing : dict
            The mapping of timing to emotions.
        empatica_shift : int
            The shift between the empatica and the GoPro videos.
        include_neutral : bool
            Whether to include the neutral label.
        verbose : bool
            Whether to show messages.
        increase_window : bool
            Whether to increase the window of the labels.
        """

        emotion_order = list(emotion_timing.keys())

        if include_neutral:
            emotion_order.insert(1, 'Neutral')

        for rec in data.keys():
            
            data[rec]['emotion_order'] = emotion_order

            sr = data[rec]['sample_rate']
            data[rec]['data']['label'] = 'None'

            neutral_start_idx = None
            neutral_end_idx = None

            if rec == 'hr':
                empatica_shift -= 10 #There's an additional shift for hr data across all subjects

            for idx, (label, (start, end)) in enumerate(emotion_timing.items()):
                
                if increase_window:
                    start_idx = int(((self.utilities._time_to_seconds(start)+empatica_shift)-20) * sr)
                    end_idx = int(((self.utilities._time_to_seconds(end)+empatica_shift)+40) * sr)
                else:
                    start_idx = int(((self.utilities._time_to_seconds(start)+empatica_shift)) * sr)
                    end_idx = int(((self.utilities._time_to_seconds(end)+empatica_shift)) * sr)

                if end_idx <= len(data[rec]['data']):
                    data[rec]['data'].loc[start_idx:end_idx, 'label'] = label
                else:
                    if verbose:
                        print(f"        {label} label is out of range in the {rec} data. Skipping..")

                if include_neutral:
                    if idx == 0:
                        neutral_start_idx = end_idx
                    elif idx == 1:
                        neutral_end_idx = start_idx
                        data[rec]['data'].loc[neutral_start_idx:neutral_end_idx, 'label'] = 'Neutral'

            data[rec]['data'] = data[rec]['data'][data[rec]['data']['label'] != 'None']

    def _load_empatica_csv(self, filepath):
        """ 
        Extract Empatica data from a CSV file.
        
        Parameters
        ----------
        filepath : str
            The path of the CSV file.
            
        Returns
        -------
        initial_time : datetime.datetime
            The initial time of the recording.
        sample_rate : float
            The sample rate of the recording.
        data : pandas.DataFrame 
            The extracted data.
        """
        
        with open(filepath, 'r') as file:
            initial_times = file.readline().strip().split(',')
            initial_time_unix = float(initial_times[0])
            initial_time = self.utilities._convert_unix_timestamp_to_utc(initial_time_unix)

            sample_rate_list = file.readline().strip().split(',')
            _sample_rate = float(sample_rate_list[0])

            _data = pd.read_csv(file, header=None)

            if _sample_rate != self.sample_rate_to_use:
                new_num_samples = int(len(_data) * self.sample_rate_to_use / _sample_rate)
                data = pd.DataFrame(resample(_data, new_num_samples))
                sample_rate = int(self.sample_rate_to_use)
            else:
                data = _data
                sample_rate = int(_sample_rate)

        return initial_time, sample_rate, data
    
    def plot_data(self, subject_id, data, data_type, filter):
        """
        Plot empatica data.
        
        Parameters
        ----------
        subject_id : int
            The subject ID.
        data : dict
            The dictionary that contains all data.
        data_type : str
            The data type ('acc', 'temp', ...).
        filter : bool
            Whether to apply a filter.
        """

        if subject_id not in data.keys():
            print("subject ID not found in data")
            return
        
        print("Visualization of Empatica data for subject " + str(subject_id))

        sample_rate = data[subject_id]['Empatica'][data_type]['sample_rate']
        if data_type == "acc":
            plt.figure(figsize=(14, 8))
            x_title = data[subject_id]['Empatica'][data_type]['x_title']
            plt.title(f"subject {subject_id} - {data[subject_id]['Empatica'][data_type]['name']} Over Time")
            for i in range(3):
                y = f"y_title_{i+1}"
                y_title = data[subject_id]['Empatica'][data_type][y]
                plt.subplot(3, 1, i+1)
                if filter:
                    y_data = self.utilities._filter_empatica(data[subject_id]['Empatica'][data_type]['data'][y_title], sample_rate, data_type)
                else:
                    y_data = data[subject_id]['Empatica'][data_type]['data'][y_title]
                x_data = data[subject_id]['Empatica'][data_type]['data'][x_title]
                plt.plot(x_data, y_data)
                plt.ylabel(y_title)
            plt.xlabel(x_title)
            plt.tight_layout()
            plt.show()
        else:
            plt.figure(figsize=(12, 6))
            x_title = data[subject_id]['Empatica'][data_type]['x_title']
            y_title = data[subject_id]['Empatica'][data_type]['y_title']
            if filter:
                y_data = self.utilities._filter_empatica(data[subject_id]['Empatica'][data_type]['data'][y_title], sample_rate, data_type)
            else:
                y_data = data[subject_id]['Empatica'][data_type]['data'][y_title]
            x_data = data[subject_id]['Empatica'][data_type]['data'][x_title]
            
            plt.plot(x_data, y_data)
            if filter:
                plt.title(f"subject {subject_id} - {data[subject_id]['Empatica'][data_type]['name']} Over Time - Filtered")
            else:
                plt.title(f"subject {subject_id} - {data[subject_id]['Empatica'][data_type]['name']} Over Time - Raw")
            plt.xlabel(x_title)
            plt.ylabel(y_title)
            plt.show()
    
    def plot_labeled_data(self, subject_id, d, data_type, filter = False, exclude_neutral=False, verbose = False, pdf=None):
        """
        Plot the labeled data.
        
        Parameters
        ----------
        subject_id : int
            The subject ID.
        d : dict
            The dictionary that contains all data.
        data_type : str
            The data_type for the data to plot ('acc', 'temp', ...).
        filter : bool
            Whether to apply a filter.
        column : str or list
            The column(s) to plot.
        exclude_neutral : bool
            Whether to exclude the neutral label.
        verbose : bool
            Whether to print messages.
        pdf : matplotlib.backends.backend_pdf.PdfPages
            The PDF file to save the plots to.
        """

        if verbose:
            print(f"Plotting labelled data for subject {subject_id}..")

        label_colors = {
            'Neutral': 'blue',
            'Joy': 'yellow',
            'Shame': 'orange',
            'Frustration': 'green',
            'Pride': 'red'
        }

        data = d[subject_id]['Empatica'][data_type]["data"]

        sample_rate = d[subject_id]['Empatica'][data_type]["sample_rate"]
        
        unique_labels = data["label"].unique()
        
        columns_to_plot = [col for col in data.columns if col not in ["Time (s)", "label"]]

        legend_handles = {} 

        for col in columns_to_plot:
            plt.figure(figsize=(15, 7))
            for label in unique_labels:
                if exclude_neutral and label == 'Neutral':
                    continue
                subset = data[data["label"] == label]
                if label == 'Neutral':
                    breaks = np.where(np.diff(subset["Time (s)"]) > 1)[0] + 1
                    segments = np.split(subset, breaks)
                    for segment in segments:
                        if filter:
                            data_to_plot = self.utilities._filter_empatica(segment[col], sample_rate, data_type)
                        else:
                            data_to_plot = segment[col]
                        line, = plt.plot(segment["Time (s)"], data_to_plot, color=label_colors[label])
                    if 'Neutral' not in legend_handles:
                        legend_handles['Neutral'] = line
                else:
                    if filter:
                        data_to_plot = self.utilities._filter_empatica(subset[col], sample_rate, data_type)
                    else:
                        data_to_plot = subset[col]
                    line, = plt.plot(subset["Time (s)"], data_to_plot, label=label, color=label_colors[label])
                    legend_handles[label] = line  

            if filter:
                plt.title(f'subject {subject_id} - Data Visualization for {col} - Filtered')
            else:
                plt.title(f'subject {subject_id} - Data Visualization for {col} - Raw')

            plt.xlabel('Time (s)')
            plt.ylabel(col)

            plt.legend(handles=list(legend_handles.values()), labels=list(legend_handles.keys()))
            plt.grid(True)
            plt.tight_layout()

            if pdf is not None:
                pdf.savefig()
                plt.close()
            else:
                plt.show()
    
    def save_empatica_data_to_pdf(self, data, filename, filter=False):
        """
        Save the raw empatica data to a PDF file.

        Parameters
        ----------
        data : dict
            The dictionary that contains all data.
        filename : str
            The name of the PDF file.
        filter : bool
            Whether to apply a filter.
        """

        with PdfPages(filename) as pdf:
            for subject_id in data.keys():
                recordings = list(data[subject_id]['Empatica'].keys())
                if 'ibi' in recordings:
                    recordings.remove('ibi')
                plt.figure(figsize=(11, 8.5))  
                plt.text(0.5, 0.5, f'Subject {subject_id}', ha='center', va='center', size=24)
                plt.axis('off')
                pdf.savefig() 
                plt.close() 
                for k in recordings:
                    self.plot_labeled_data(subject_id, data, k, filter= filter, verbose= False, pdf=pdf)
            
class TRANSCRIPT:
    
    def __init__(self):
        return
    
    def get_data(self, filepath, include_neutral):
        """
        Extracts the data from the Word document.
        
        Parameters
        ----------
        filepath : str
            The name of the Word document.
        include_neutral : bool
            Whether to add the neutral label.
        
        Returns
        -------
        data : dict
            The extracted data.
        """

        text = self._extract_text_from_word(filepath)
        emotions = re.findall(r'(\w+) Prompt', text)
        data = {}

        for i in range(len(emotions)):
            if i == 1 and include_neutral:
                data['Neutral'] = self._synthesize_neutral_text()

            if i < len(emotions) - 1:
                between_text = re.search(f'(?<={emotions[i]} Prompt)(.*?)(?={emotions[i+1]} Prompt)', text, re.DOTALL)
            else:
                between_text = re.search(f'(?<={emotions[-1]} Prompt)(.*)', text, re.DOTALL)

            if between_text:
                section = between_text.group(1).strip()
                data[emotions[i]] = self._extract_values(section) 
        
        return data
    
    def _synthesize_neutral_text(self):
        """
        Generates neutral text based on predefined templates.
        
        Returns
        -------
        dict 
            A dict of strings of synthesized neutral text.
        """

        templates = [
            "Today, the weather is {} and the temperature is around {} degrees.",
            "I spent some time reading a book which is about {}.",
            "This morning, I had a cup of {} for breakfast.",
            "Then, I went for a walk and noticed {}.",
            "The current time is {}and I have a meeting scheduled at {}.",
            "On my way to work, I usually pass by a {}.",
            "Yesterday, I organized my workspace. Now, it feels more {}.",
            "During my lunch break, I usually eat {}.",
            "In the evening, I plan to watch a documentary about {}.",
            "This weekend, I might visit the local {}.",
            "My favorite color is {}. It's a very calming color.",
            "For exercise, I often go {}. It's quite refreshing.",
            "The book I'm currently reading is about {}. It's very informative.",
            "My morning routine includes {} which helps me start my day.",
            "I recently learned about {}. It's quite interesting.",
            "The last movie I watched was about {}. It was quite engaging.",
            "One of my hobbies is {}. It's very relaxing.",
            "For dinner last night, I made {}. It was quite delicious.",
            "I'm planning a trip to {}. I've heard it's a beautiful place.",
            "On the weekends, I enjoy {}. It's a great way to relax."
        ]

        fillers = [
            ["weather_conditions", "temperatures"],
            ["book_topics"],
            ["breakfast_items"],
            ["observations"],
            ["times", "meeting_times"],
            ["pass_by_places"],
            ["workspace_conditions"],
            ["lunch_items"],
            ["documentary_topics"],
            ["local_places"],
            ["colors"],
            ["exercise_activities"],
            ["book_topics"],
            ["breakfast_items"],
            ["learning_topics"],
            ["movie_topics"],
            ["hobbies"],
            ["dinner_items"],
            ["trip_destinations"],
            ["weekend_activities"]
        ]

        filler_options = {
            "weather_conditions": ["sunny", "cloudy", "clear", "overcast"],
            "temperatures": ["21", "18", "25", "20"],
            "book_topics": ["history", "science", "technology", "art"],
            "breakfast_items": ["coffee", "tea", "milk", "juice"],
            "observations": ["a tree with green leaves", "a quiet street", "birds in the sky", "cars passing by"],
            "times": ["10 AM", "3 PM", "6 PM", "9 AM"],
            "meeting_times": ["11 AM", "4 PM", "7 PM", "10 AM"],
            "pass_by_places": ["a park", "a coffee shop", "an old bookstore", "a river"],
            "workspace_conditions": ["organized", "spacious", "neat", "clean"],
            "lunch_items": ["a sandwich", "salad", "pasta", "rice and vegetables"],
            "documentary_topics": ["nature", "history", "science", "architecture"],
            "local_places": ["museum", "library", "garden", "market"],
            "colors": ["blue", "green", "gray", "beige"],
            "exercise_activities": ["jogging", "walking", "cycling", "swimming"],
            "learning_topics": ["a new language", "guitar chords", "cooking recipes", "photography skills"],
            "movie_topics": ["an historical event", "a scientific discovery", "a famous artist", "a cultural practice"],
            "hobbies": ["reading", "painting", "gardening", "knitting"],
            "dinner_items": ["stir-fry vegetables", "grilled chicken", "homemade pizza", "seafood paella"],
            "trip_destinations": ["the mountains", "a coastal town", "a nearby city", "the countryside"],
            "weekend_activities": ["reading", "hiking", "visiting friends", "exploring new cafes"]
        }

        paragraph = []

        for template, filler_types in zip(templates, fillers):
            selected_fillers = [random.choice(filler_options[filler]) for filler in filler_types]
            text = template.format(*selected_fillers)
            paragraph.append(text)

        return {
                    'Start': None,
                    'End': None,
                    'P': ' '.join(paragraph)
                }
    
    def _extract_text_from_word(self, filepath):
        """
        Extracts the text prompts from a Word document.
        
        Parameters
        ----------  
        filepath : str
            The name of the Word document.
        
        Returns
        -------
        full_text : str
            The text from the Word document.
        """

        doc = Document(filepath)
        full_text = []

        for paragraph in doc.paragraphs:
            full_text.append(paragraph.text)

        return '\n'.join(full_text)

    def _extract_values(self, section):
        """
        Extracts Start, End, and all Prompt values from the extract word section.
        
        Parameters
        ----------
        section : str
            The section to extract the values from.
        
        Returns
        -------
        values : dict
            The extracted values.
        """

        start_time = re.search(r'Start: (.*?)\n', section)
        end_time = re.search(r'End: (.*?)\n', section)
        p_values = re.findall(r'P: “(.*?)(?=”|$|Start:|End:|I:)', section, re.DOTALL)
        
        concatenated_p_values = ' '.join([p.strip() for p in p_values])
        
        return {
            'Start': start_time.group(1) if start_time else None,
            'End': end_time.group(1) if end_time else None,
            'P': concatenated_p_values
        }
        
class AUDIO:
    
    def __init__(self):
        
        self.utilities = UTILITIES()

        return
    
    def get_data(self, filepath, emotion_timing, sample_rate=11025):
        """

        Gets the amplitude data and the necessary information for the analysis of the audio file.
        
        Parameters
        ----------
        filepath : str
            The path to the audio file.
        emotion_timing : dict
            The mapping of timing to emotions.
        sample_rate : int
            The sampling rate of the audio file.
            
        Returns
        -------
        dict
            A dictionary containing the audio data, sampling rate, and the time of the first clap.
        """
        
        audio, sr = librosa.load(filepath, sr=sample_rate)

        #Get the end time of the last prompt
        max_end_time = max(self.utilities._time_to_seconds(times[1]) for times in emotion_timing.values()) + 200

        # Remove the end of the audio file (from the end of the last prompt to the end of the recording). 
        audio_trimmed = audio[:int(max_end_time * sr)]
        audio_silence_trimmed, _ = librosa.effects.trim(audio_trimmed, top_db=30)
        time_index = np.arange(0, len(audio_silence_trimmed)) / sr
        
        return {'data': pd.DataFrame({
                        'Time': time_index,
                        'Amplitude': audio_silence_trimmed
                        }),
                'sample_rate': sr,
                }

    def label_data(self, data, emotion_timing, audio_clap_time, go_pro_clap_time, include_neutral, verbose=False):
        """
        Label the data based on the segments.
        
        Parameters
        ----------
        data : dict
            The dictionary that contains all data.
        emotion_timing : dict
            The mapping of segments to emotions.
        audio_clap_time : str
            The time of the first clap in the audio file.
        go_pro_clap_time : str
            The time of the first clap in the GoPro video.
        include_neutral : bool
            Whether to include the neutral label.
        verbose : bool
            Whether to show messages.
        """

        emotion_order = list(emotion_timing.keys())

        if include_neutral:
            emotion_order.insert(1, 'Neutral')

        data['emotion_order'] = emotion_order

        sr = data['sample_rate']

        #Align the prompt timings of the audio with the one from the GoPro videos
        shift = int(self.utilities._time_to_seconds(go_pro_clap_time)) - audio_clap_time
        
        data['data']['label'] = 'None'

        neutral_start_idx = None
        neutral_end_idx = None

        for idx, (label, (start, end)) in enumerate(emotion_timing.items()):
            start_idx = int((self.utilities._time_to_seconds(start)-shift) * sr)
            end_idx = int((self.utilities._time_to_seconds(end)-shift) * sr)

            if end_idx <= len(data['data']):
                data['data'].loc[start_idx:end_idx, 'label'] = label
            else:
                if verbose:
                    print(f"        {label} label is out of range in the audio data. Skipping..")

            if include_neutral:
                if idx == 0:
                    neutral_start_idx = end_idx
                elif idx == 1:
                    neutral_end_idx = start_idx
                    data['data'].loc[neutral_start_idx:neutral_end_idx, 'label'] = 'Neutral'

        data['data'] = data['data'][data['data']['label'] != 'None']
    
    def plot_labeled_data(self, subject_id, d, filter = False, ploting_reduction_factor = 100, verbose = False, pdf=None):
        """
        Plot the labeled data.
        
        Parameters
        ----------
        subject_id : int
            The subject ID.
        d : dict
            The dictionary that contains all data.
        key : str
            The key for the data to plot.
        filter : bool
            Whether to apply a filter.
        ploting_reduction_factor: int
            The factor to reduce the number of points to plot.
        verbose : bool
            Whether to print messages.
        pdf : matplotlib.backends.backend_pdf.PdfPages
            The PDF file to save the plots to.
        """

        if verbose:
            print(f"Plotting labelled data for subject {subject_id}..")

        label_colors = {
            'Joy': 'yellow',
            'Shame': 'orange',
            'Frustration': 'green',
            'Pride': 'red',
            'Neutral': 'blue'
        }

        original_data = d[subject_id]['Audio']["data"]
        data = original_data[::ploting_reduction_factor]
        
        unique_labels = data["label"].unique()
        
        legend_handles = {} 

        plt.figure(figsize=(15, 7))
        
        for label in unique_labels:
            subset = data[data["label"] == label]

            if filter:
                data_to_plot = self.utilities._filter_audio(subset["Amplitude"], d[subject_id]['Audio']["sample_rate"])
            else:
                data_to_plot = subset["Amplitude"]

            line, = plt.plot(subset["Time"], data_to_plot, label=label, color=label_colors[label])
            legend_handles[label] = line  

        if filter:
            plt.title(f'subject {subject_id} - Audio Data Visualization - Filtered')
        else:
            plt.title(f'subject {subject_id} - Audio Data Visualization - Raw')

        plt.xlabel('Time (s)')
        plt.ylabel("Amplitude")

        plt.legend(handles=list(legend_handles.values()), labels=list(legend_handles.keys()))
        plt.grid(True)
        plt.tight_layout()

        if pdf is not None:
            pdf.savefig()
            plt.close()
        else:
            plt.show()
    
    def save_audio_data_to_pdf(self, data, filename, filter=False):
        """
        Save the audio data to a PDF file.

        Parameters
        ----------
        data : dict
            The dictionary that contains all data.
        filename : str
            The name of the PDF file.
        filter : bool
            Whether to apply a filter.
        """

        with PdfPages(filename) as pdf:
            for subject_id in data.keys():
                plt.figure(figsize=(11, 8.5))  
                plt.text(0.5, 0.5, f'Subject {subject_id}', ha='center', va='center', size=24)
                plt.axis('off')
                pdf.savefig() 
                plt.close() 
                self.plot_labeled_data(subject_id, data, filter= filter, verbose= False, pdf=pdf)

class WEBCAM:

    def __init__(self):
        """
        Initialize the Webcam class.
        """

        self.utilities = UTILITIES()

    def get_data(self, filepath):
        """
        Get the video data from the files using the VideoFileClip library.

        Parameters
        ----------
        filepath : str
            The path to the video file.
        
        Returns
        -------
        dict
            A dictionary containing the video data path, sampling rate, the duration and the time of the first clap.
        """

        clip = VideoFileClip(filepath)

        duration = clip.duration
        fps = clip.fps

        return {'data_path': filepath, 
                'fps': fps,
                'duration': duration,
                }
    
    def label_data(self, data, emotion_timing, webcam_clap_time, go_pro_clap_time, include_neutral, verbose):
        """
        Label the webcam data by extracting start and end frames of each emotion prompt.

        Parameters
        ----------
        data : dict
            The dictionary that contains all data.
        emotion_timing : dict
            The mapping of timing to emotions.
        webcam_clap_time : str
            The time of the first clap in the webcam video.
        go_pro_clap_time : str
            The time of the first clap in the GoPro video.
        include_neutral : bool
            Whether to include the neutral label.
        verbose: bool
            Whether to show messages.
        """

        emotion_order = list(emotion_timing.keys())

        if include_neutral:
            emotion_order.insert(1, 'Neutral')

        data['emotion_order'] = emotion_order
        
        shift = int(self.utilities._time_to_seconds(go_pro_clap_time) - webcam_clap_time)

        neutral_start_frame = None 
        neutral_end_frame = None

        data['label'] = dict()

        for idx, (emotion, (start, end)) in enumerate(emotion_timing.items()):

            start_frame = self.utilities._timestamp_to_frame(start, data['fps'], shift)
            end_frame = self.utilities._timestamp_to_frame(end, data['fps'], shift)

            if include_neutral:
                if idx == 0:
                    neutral_start_frame = end_frame
                elif idx == 1:
                    neutral_end_frame = start_frame
                    data['label']['Neutral'] = (neutral_start_frame, neutral_end_frame)

            data['label'][emotion] = (start_frame, end_frame)
        
        total_number_frames = data['duration']*data['fps']

        keys_to_delete = []

        for key, (_, end_frame) in data['label'].items():

            if end_frame > total_number_frames:
                keys_to_delete.append(key)

        for key in keys_to_delete:

            if verbose:
                print(f"        {key} label is out of range in the webcam video. Skipping..")
            del data['label'][key]

class GOPRO:

    def __init__(self):
        """
        Initialize the GoPro class.
        """

        self.utilities = UTILITIES()

        return
    
    def get_data(self, filepath):
        """
        Get the data from the cut files.

        Parameters
        ----------
        filepath : str
            The path to the video file.
        
        Returns
        -------
        dict
            A dictionary containing the video data.
        """

        vid = cv2.VideoCapture(filepath)
        fps = int(vid.get(cv2.CAP_PROP_FPS))+1 #To round the fps to 30
        total_frames = int(vid.get(cv2.CAP_PROP_FRAME_COUNT))
        duration = total_frames / fps
        vid.release()

        return {
            'data_path': filepath,
            'fps': fps,
            'duration': duration
        }

    def label_data(self, data, emotion_timing, include_neutral):
        """
        Label the data with the cut videos.

        Parameters
        ----------
        data : dict
            The dictionary that contains all data.
        emotion_timing : dict
            The mapping of timing to emotions.
        include_neutral : bool
            Whether to include the neutral label.
        """

        emotion_order = list(emotion_timing.keys())

        if include_neutral:
            emotion_order.insert(1, 'Neutral')

        data['emotion_order'] = emotion_order

        neutral_start_frame = None
        neutral_end_frame = None

        data['label'] = dict()

        for idx, (emotion, (start, end)) in enumerate(emotion_timing.items()):

            start_frame = self.utilities._timestamp_to_frame(start, data['fps'])
            end_frame = self.utilities._timestamp_to_frame(end, data['fps'])
            
            if include_neutral:

                if idx == 0:
                    neutral_start_frame = end_frame
                elif idx == 1:
                    neutral_end_frame = start_frame
                    data['label']['Neutral'] = (neutral_start_frame, neutral_end_frame)

            data['label'][emotion] = (start_frame, end_frame)

class SRE:
    
    def __init__(self):
        """
        Initialize the SRE class.
        """

        self.positive_emotions = ['Proud', 'Lively', 'Competent', 'Confident', 'Satisfied', 'Joy']
        self.negative_emotions = ['Angry', 'Ashamed', 'Annoyed', 'Worthless', 'Embarrassed', 'Frustrated']
        self.scores = {
            'Not at all': 0,
            'A little bit': 1,
            'Moderately': 2,
            'Quite a bit': 3,
            'Extremely': 4
        }

        return

    def _extract_data(self, path):
        """
        Extract Self Reported Emotion Data from the file and optionally saves it in a json format

        Parameters
        ----------
        path : str
            The path to the general folder.

        Returns:
        --------
        dict
            A dictionary containing the self reported emotion data.
        """

        filepath = os.path.join(path, 'RedCap Data', 'NonVerbalMeasurement_DATA_LABELS_2023-11-01_1238.csv')
        df = pd.read_csv(filepath)

        subject_id_list = df['Record ID'].tolist()
        emotion_order_list = df['Which condition was this participant assigned?'].tolist()

        angry_shame_list = df['Angry_Shame'].tolist()
        proud_shame_list = df['Proud_Shame'].tolist()
        lively_shame_list = df['Lively_Shame'].tolist()
        ashamed_shame_list = df['Ashamed_Shame'].tolist()
        competent_shame_list = df['Competent_Shame'].tolist()
        annoyed_shame_list = df['Annoyed_Shame'].tolist()
        confident_shame_list = df['Confident_Shame'].tolist()
        satisfied_shame_list = df['Satisfied_Shame'].tolist()
        worthless_shame_list = df['Worthless_Shame'].tolist()
        embarrassed_shame_list = df['Embarrassed_Shame'].tolist()
        frustrated_shame_list = df['Frustrated_Shame'].tolist()
        joy_shame_list = df['Joy_Shame'].tolist()

        angry_frustration_list = df['Angry_Frustration'].tolist()
        proud_frustration_list = df['Proud_Frustration'].tolist()
        lively_frustration_list = df['Lively_Frustration'].tolist()
        ashamed_frustration_list = df['Ashamed_Frustration'].tolist()
        competent_frustration_list = df['Competent_Frustration1'].tolist()
        annoyed_frustration_list = df['Annoyed_Frustration'].tolist()
        confident_frustration_list = df['Confident_Frustration'].tolist()
        satisfied_frustration_list = df['Satisfied_Frustration'].tolist()
        worthless_frustration_list = df['Worthless_Frustration'].tolist()
        embarrassed_frustration_list = df['Embarrassed_Frustration'].tolist()
        frustrated_frustration_list = df['Frustrated_Frustration'].tolist()
        joy_frustration_list = df['Joy_Frustration'].tolist()

        angry_pride_list = df['Angry_Pride'].tolist()
        proud_pride_list = df['Proud_Pride'].tolist()
        lively_pride_list = df['Lively_Pride'].tolist()
        ashamed_pride_list = df['Ashamed_Pride'].tolist()
        competent_pride_list = df['Competent_Pride'].tolist()
        annoyed_pride_list = df['Annoyed_Pride'].tolist()
        confident_pride_list = df['Confident_Pride'].tolist()
        satisfied_pride_list = df['Satisfied_Pride'].tolist()
        worthless_pride_list = df['Worthless_Pride'].tolist()
        embarrassed_pride_list = df['Embarrassed_Pride'].tolist()
        frustrated_pride_list = df['Frustrated_Pride'].tolist()
        joy_pride_list = df['Joy_Pride'].tolist()

        angry_joy_list = df['Angry_Joy'].tolist()
        proud_joy_list = df['Proud_Joy'].tolist()
        lively_joy_list = df['Lively_Joy'].tolist()
        ashamed_joy_list = df['Ashamed_Joy'].tolist()
        competent_joy_list = df['Competent_Joy'].tolist()
        annoyed_joy_list = df['Annoyed_Joy'].tolist()
        confident_joy_list = df['Confident_Joy'].tolist()
        satisfied_joy_list = df['Satisfied_Joy'].tolist()
        worthless_joy_list = df['Worthless_Joy'].tolist()
        embarrassed_joy_list = df['Embarrassed_Joy'].tolist()
        frustrated_joy_list = df['Frustrated_Joy'].tolist()
        joy_joy_list = df['Joy_Joy'].tolist()

        reported_emotions = dict()

        for i, subject_id in enumerate(subject_id_list):

            if emotion_order_list[i] is np.nan:
                reported_emotions[subject_id] = "No Self Reported data"
            else:
                reported_emotions[subject_id] = dict()
                emotions = emotion_order_list[i].split(', ')
                emotions.append('Joy')
                emotions.append('Neutral')

                for emotion in emotions:
                    reported_emotions[subject_id][emotion] = dict()
                    
                    if emotion == 'Shame':
                        reported_emotions[subject_id][emotion] = {
                            'Angry': angry_shame_list[i].split('. ')[1],
                            'Proud': proud_shame_list[i].split('. ')[1],
                            'Lively': lively_shame_list[i].split('. ')[1],
                            'Ashamed': ashamed_shame_list[i].split('. ')[1],
                            'Competent': competent_shame_list[i].split('. ')[1],
                            'Annoyed': annoyed_shame_list[i].split('. ')[1],
                            'Confident': confident_shame_list[i].split('. ')[1],
                            'Satisfied': satisfied_shame_list[i].split('. ')[1],
                            'Worthless': worthless_shame_list[i].split('. ')[1],
                            'Embarrassed': embarrassed_shame_list[i].split('. ')[1],
                            'Frustrated': frustrated_shame_list[i].split('. ')[1],
                            'Joy': joy_shame_list[i].split('. ')[1]
                        }
                    elif emotion == 'Frustration':
                        reported_emotions[subject_id][emotion] = {
                            'Angry': angry_frustration_list[i].split('. ')[1],
                            'Proud': proud_frustration_list[i].split('. ')[1],
                            'Lively': lively_frustration_list[i].split('. ')[1],
                            'Ashamed': ashamed_frustration_list[i].split('. ')[1],
                            'Competent': competent_frustration_list[i].split('. ')[1],
                            'Annoyed': annoyed_frustration_list[i].split('. ')[1],
                            'Confident': confident_frustration_list[i].split('. ')[1],
                            'Satisfied': satisfied_frustration_list[i].split('. ')[1],
                            'Worthless': worthless_frustration_list[i].split('. ')[1],
                            'Embarrassed': embarrassed_frustration_list[i].split('. ')[1],
                            'Frustrated': frustrated_frustration_list[i].split('. ')[1],
                            'Joy': joy_frustration_list[i].split('. ')[1]
                        }
                    elif emotion == 'Pride':
                        reported_emotions[subject_id][emotion] = {
                            'Angry': angry_pride_list[i].split('. ')[1],
                            'Proud': proud_pride_list[i].split('. ')[1],
                            'Lively': lively_pride_list[i].split('. ')[1],
                            'Ashamed': ashamed_pride_list[i].split('. ')[1],
                            'Competent': competent_pride_list[i].split('. ')[1],
                            'Annoyed': annoyed_pride_list[i].split('. ')[1],
                            'Confident': confident_pride_list[i].split('. ')[1],
                            'Satisfied': satisfied_pride_list[i].split('. ')[1],
                            'Worthless': worthless_pride_list[i].split('. ')[1],
                            'Embarrassed': embarrassed_pride_list[i].split('. ')[1],
                            'Frustrated': frustrated_pride_list[i].split('. ')[1],
                            'Joy': joy_pride_list[i].split('. ')[1]
                        }
                    elif emotion == 'Joy':
                        reported_emotions[subject_id][emotion] = {
                            'Angry': angry_joy_list[i].split('. ')[1],
                            'Proud': proud_joy_list[i].split('. ')[1],
                            'Lively': lively_joy_list[i].split('. ')[1],
                            'Ashamed': ashamed_joy_list[i].split('. ')[1],
                            'Competent': competent_joy_list[i].split('. ')[1],
                            'Annoyed': annoyed_joy_list[i].split('. ')[1],
                            'Confident': confident_joy_list[i].split('. ')[1],
                            'Satisfied': satisfied_joy_list[i].split('. ')[1],
                            'Worthless': worthless_joy_list[i].split('. ')[1],
                            'Embarrassed': embarrassed_joy_list[i].split('. ')[1],
                            'Frustrated': frustrated_joy_list[i].split('. ')[1],
                            'Joy': joy_joy_list[i].split('. ')[1]
                        }
                    elif emotion == 'Neutral':
                        reported_emotions[subject_id][emotion] = {
                                    'Angry': 'Not at all',
                                    'Proud': 'Not at all',
                                    'Lively': 'Not at all',
                                    'Ashamed': 'Not at all',
                                    'Competent': 'Not at all',
                                    'Annoyed': 'Not at all',
                                    'Confident': 'Not at all',
                                    'Satisfied': 'Not at all',
                                    'Worthless': 'Not at all',
                                    'Embarrassed': 'Not at all',
                                    'Frustrated': 'Not at all',
                                    'Joy': 'Not at all'
                                }

        return reported_emotions
    
class FaceReader:

    def __init__(self):
        """
        Initializes the FaceReader class.
        """
        self.features = ['Neutral', 'Happy', 'Sad', 'Angry', 'Surprised', 'Scared', 'Disgusted', 'Valence',
                        'Arousal', 'Pitch', 'Yaw', 'Roll', 'Mouth', 'Left Eye', 'Right Eye', 'Left Eyebrow',
                        'Right Eyebrow', 'Action Unit 01 - Left - Inner Brow Raiser', 'Action Unit 02 - Left - Outer Brow Raiser', 
                        'Action Unit 04 - Left - Brow Lowerer', 'Action Unit 05 - Left - Upper Lid Raiser', 'Action Unit 06 - Left - Cheek Raiser', 
                        'Action Unit 07 - Left - Lid Tightener', 'Action Unit 12 - Left - Lip Corner Puller', 'Action Unit 14 - Left - Dimpler', 
                        'Action Unit 15 - Left - Lip Corner Depressor', 'Action Unit 20 - Left - Lip Stretcher', 'Action Unit 43 - Left - Eyes Closed', 
                        'Action Unit 01 - Right - Inner Brow Raiser', 'Action Unit 02 - Right - Outer Brow Raiser', 'Action Unit 04 - Right - Brow Lowerer', 
                        'Action Unit 05 - Right - Upper Lid Raiser', 'Action Unit 06 - Right - Cheek Raiser', 'Action Unit 07 - Right - Lid Tightener', 
                        'Action Unit 12 - Right - Lip Corner Puller', 'Action Unit 14 - Right - Dimpler', 'Action Unit 15 - Right - Lip Corner Depressor', 
                        'Action Unit 20 - Right - Lip Stretcher', 'Action Unit 43 - Right - Eyes Closed', 'Horizontal position', 
                        'Vertical position', 'Depth position']

    def get_data(self, filename, desired_subject_id, include_neutral=True):
        """
        Gets the data from the FaceReader files.

        Parameters
        ----------
        filename : str
            The path to the FaceReader file.
        desired_subject_id : int
            The desired subject ID.
        include_neutral : bool
            Whether to include the neutral label.
        
        Returns
        -------
        data : dict
            A dictionary containing the data.
        """

        workbook = openpyxl.load_workbook(filename)

        data = dict()

        for sheet_name in workbook.sheetnames:

            sheet = workbook[sheet_name]

            for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
                if row[0] == 'Filename':
                    filename = row[1]
                    match = re.search(r'\\(\d+)_', filename)

                    if match:
                        subject_id = int(match.group(1))

            if subject_id != desired_subject_id:    
                if include_neutral:
                    nb_emotions = 5
                else:
                    nb_emotions = 4

                if len(list(data.keys())) == nb_emotions:
                    break
                else:
                    continue

            if 'frustration' in filename or 'frust' in filename:
                emotion = 'Frustration'
            elif 'joy' in filename:
                emotion = 'Joy'
            elif 'shame' in filename:
                emotion = 'Shame'
            elif 'pride' in filename:
                emotion = 'Pride'

            data[emotion] = dict()

            for row in sheet.iter_rows(min_row=10, max_row=15):
                if row[0].value == 'Video Time':
                    feature_row = row
                    break

            for cell in feature_row:
                if cell.value in self.features:

                    feature = self.preprocess_feature_name(cell.value)

                    data[emotion][feature] = list()

                    for _row in sheet.iter_rows(min_row=cell.row+1, min_col=cell.column, max_col=cell.column):
                        for _cell in _row:
                            if _cell.value == 'Closed' or _cell.value == 'Neutral':
                                feature_value = 0
                            elif _cell.value == 'Open' or _cell.value == 'Lowered':
                                feature_value = 1
                            elif _cell.value == 'Raised':
                                feature_value = 2
                            elif _cell.value == 'FIT_FAILED':
                                if data[emotion][feature] == []:
                                    feature_value = 0
                                else:
                                    feature_value = np.mean(data[emotion][feature])
                            else:
                                feature_value = _cell.value

                            if feature_value is None or type(feature_value) == str:
                                print(f'Error in {subject_id} {emotion} {feature}', type(feature_value), feature_value)
                                break  

                            data[emotion][feature].append(feature_value)
                    
        if include_neutral:
            data['Neutral'] = {key: [0]*len(value) for key, value in data[emotion].items()} 

        return data
    
    
    def preprocess_feature_name(self, feature):
        """
        Preprocess the feature name.

        Parameters
        ----------
        feature : str
            The feature name.
        
        Returns
        -------
        str
            The preprocessed feature name.
        """
        if 'Action Unit' in feature:
            if 'Left' in feature:
                feature = 'left-' + feature.split(' - ')[2].strip()
            elif 'Right' in feature:
                feature = 'right-' + feature.split(' - ')[2].strip()
            else:
                feature = feature.split('-')[1].strip()
        feature = feature.lower().replace(' ', '-')
        if feature == 'pitch':
            feature = 'head-pitch'

        return feature
    
class FEATURES:
    
    def __init__(self, desired_measurement, logfile = None):
        """
        Initialize the class.

        Parameters
        ----------
        desired_measurement : str
            The desired data to extract the features from.
        logfile : str
            The name of the log file.
        """

        self.utilities = UTILITIES()
        self.scaler = StandardScaler()

        nltk.download('punkt', quiet=True)
        nltk.download('averaged_perceptron_tagger', quiet=True)
        nltk.download('stopwords', quiet=True)
        nltk.download('wordnet', quiet=True)
        nltk.download('brown', quiet=True)
        nltk.download('movie_reviews', quiet=True)
        nltk.download('maxent_ne_chunker', quiet = True)
        nltk.download('words', quiet = True)
        nltk.download('vader_lexicon', quiet = True)
        
        logging.getLogger('transformers').setLevel(logging.ERROR)

        self.tokenizer = BertTokenizer.from_pretrained('bert-base-uncased')
        self.model = BertModel.from_pretrained('bert-base-uncased')

        #Facial Landmarks
        self.detector = dlib.get_frontal_face_detector()

        if os.path.exists("models/shape_predictor_68_face_landmarks.dat"):
            self.predictor = dlib.shape_predictor("models/shape_predictor_68_face_landmarks.dat")
        elif os.path.exists("../models/shape_predictor_68_face_landmarks.dat"):
            self.predictor = dlib.shape_predictor("../models/shape_predictor_68_face_landmarks.dat")
        else:
            print("Please download the shape_predictor_68_face_landmarks.dat file from http://dlib.net/files/shape_predictor_68_face_landmarks.dat.bz2 and place it in the models folder.")

        self.landmark_indices_to_distance = [
            ((48,54), 'mouthCornerToMouthCorner'),
            ((62, 66), 'upperLipToLowerLip'),
            ((20, 23), 'leftEyebrowToRightEyebrow'),
            ((20, 39), 'leftEyebrowToLeftEyeInner'),
            ((23, 42), 'rightEyebrowToRightEyeInner'),
            ((18, 36), 'leftEyebrowToLeftEyeOuter'),
            ((25, 45), 'rightEyebrowToRightEyeOuter'),
            ((57, 8), 'bottomLipToChin'),
        ]

        self.landmark_indices_to_angle = [
            ((21, 27, 22), 'leftEyebrow-middleForehead-rightEyebrow'),
            ((48, 57, 54), 'leftMouthCorner-bottomLip-rightMouthCorner'),
        ]

        self.mp_pose = mp.solutions.pose
        self.pose = self.mp_pose.Pose(static_image_mode=False, model_complexity=1, smooth_landmarks=True, min_detection_confidence=0.5, min_tracking_confidence=0.5)

        self.desired_measurement = desired_measurement

        self.positive_emotions = ['Proud', 'Lively', 'Competent', 'Confident', 'Satisfied', 'Joy']
        self.negative_emotions = ['Angry', 'Ashamed', 'Annoyed', 'Worthless', 'Embarrassed', 'Frustrated']
        self.scores = {
            'Not at all': 0,
            'A little bit': 1,
            'Moderately': 2,
            'Quite a bit': 3,
            'Extremely': 4
        }

        self.json_logging_filename = logfile

    def extract_features(self, data, save=False, equalize = False):
        """
        Extract features from the data.
        
        Parameters
        ----------  
        data : dict
            The dictionary that contains all data.
        save : bool
            Whether to save the features to a file.
        equalize : bool
            Whether to equalize the number of samples for each emotion.
        
        Returns
        -------
        all_features : pd.DataFrame
            The extracted features.
        stand_features: pd.DataFrame
            The standardized features.
        
        """

        self.utilities._log_msg(self.json_logging_filename,"Extracting features.")

        pbar = tqdm(total=len(data.keys()), desc="Extracting features")

        features = dict()

        all_emotions = ['Joy', 'Shame', 'Frustration', 'Pride', 'Neutral']
        
        for subject_id in data.keys():

            pbar.set_description(f"Extracting features for subject {subject_id}")

            if 'Empatica' in self.desired_measurement:
                em_empatica = data[subject_id]['Empatica']['temp']["data"]["label"].unique()
                emotion_order = data[subject_id]['Empatica']['temp']['emotion_order']
            else:
                em_empatica = all_emotions
            if 'Audio' in self.desired_measurement:
                em_audio = data[subject_id]['Audio']['data']["label"].unique()
                emotion_order = data[subject_id]['Audio']['emotion_order']
            else:
                em_audio = all_emotions
            if 'Transcript' in self.desired_measurement:
                em_transcript = data[subject_id]['Transcript'].keys()
                emotion_order = em_transcript
            else:
                em_transcript = all_emotions
            if 'Webcam' in self.desired_measurement:
                em_webcam = data[subject_id]['Webcam']["label"].keys()
                emotion_order = data[subject_id]['Webcam']['emotion_order']
            else:
                em_webcam = all_emotions
            if 'GoPro' in self.desired_measurement:
                em_gopro_frontal = data[subject_id]['GoPro']['Frontal']["label"].keys()
                em_gopro_lateral = data[subject_id]['GoPro']['Lateral']["label"].keys()
                em_gopro = list(set(em_gopro_frontal).intersection(em_gopro_lateral))
                emotion_order = data[subject_id]['GoPro']['Frontal']['emotion_order']
            else:
                em_gopro = all_emotions
            if 'SRE' in self.desired_measurement:
                em_sre = data[subject_id]['SRE'].keys()
                emotion_order = em_sre
            else:
                em_sre = all_emotions
            if 'FaceReader' in self.desired_measurement:
                em_facereader = data[subject_id]['FaceReader'].keys()
                emotion_order = em_facereader
            else:
                em_facereader = all_emotions

            # Create a mapping from element to index for the correct order of emotions
            index_map = {value: index for index, value in enumerate(emotion_order)}

            _emotions = list(set(em_empatica).intersection(em_transcript, em_audio, em_webcam, em_gopro, em_sre, em_facereader))
            emotions = sorted(_emotions, key=lambda x: index_map[x])

            inter_dict = dict()

            for emotion in emotions:

                try:
                    intra_features = dict()
                        
                    intra_features['label'] = emotion
                    
                    if 'Empatica' in self.desired_measurement:
                        self._extract_empatica_features(intra_features, emotion, data[subject_id]['Empatica'])
                    if 'Transcript' in self.desired_measurement:
                        self._extract_transcript_features(intra_features, emotion, data[subject_id]['Transcript'], len(emotions))
                    if 'Audio' in self.desired_measurement:
                        self._extract_audio_features(intra_features, emotion, data[subject_id]['Audio'])
                    if 'Webcam' in self.desired_measurement:
                        self._extract_webcam_features(intra_features, emotion, data[subject_id]['Webcam'])
                    if 'GoPro' in self.desired_measurement:
                        self._extract_gopro_frontal_features(intra_features, emotion, data[subject_id]['GoPro']['Frontal'])
                        self._extract_gopro_lateral_features(intra_features, emotion, data[subject_id]['GoPro']['Lateral'])
                    if 'SRE' in self.desired_measurement:
                        self._extract_sre_features(intra_features, emotion, data[subject_id]['SRE'])
                    if 'FaceReader' in self.desired_measurement:
                        self._extract_face_reader_features(intra_features, emotion, data[subject_id]['FaceReader'])
                    
                    inter_dict[emotion] = pd.DataFrame([intra_features]) 

                    if inter_dict[emotion].isnull().values.any():
                        self.utilities._log_msg(self.json_logging_filename,f"    NaN values found for subject {subject_id} and emotion {emotion}.")
                     
                except Exception as e:
                    error_info = {
                        "subject_id": subject_id,
                        "emotion": emotion,
                        "error_message": str(e)
                    }
                    
                    self.utilities._log_msg(self.json_logging_filename, error_info)
                
            features[subject_id] = pd.concat(inter_dict.values(), ignore_index=True)

            pbar.update(1)

        all_features = pd.concat(features.values(), ignore_index=True)
        orig_len = len(all_features)
        all_features.dropna(inplace=True)
        new_len = len(all_features)

        self.utilities._log_msg(self.json_logging_filename,f"    Dropped {orig_len - new_len} rows due to NaN values.")
        
        stand_features = all_features.copy()
        cols_to_standardize = stand_features.columns.difference(['label'])
        stand_features[cols_to_standardize] = self.scaler.fit_transform(stand_features[cols_to_standardize])  

        if equalize:
            all_features = self.utilities._equalize_emotion_segments(all_features)
            stand_features = self.utilities._equalize_emotion_segments(stand_features)
            self.utilities._log_msg(self.json_logging_filename,"    Equalized the number of samples for each emotion.")
        
        pbar.set_description('Extracting features')
        pbar.close()

        self.utilities._log_msg(self.json_logging_filename,"Extraction Completed.")
        self.utilities._log_msg(self.json_logging_filename)

        if save:
            os.makedirs('computed_features', exist_ok=True)
            all_features.to_csv('computed_features/all_features.csv', index=False)
            stand_features.to_csv('computed_features/stand_features.csv', index=False)
            self.utilities._log_msg(self.json_logging_filename, "Features saved to computed_features folder.")
        
        self.utilities._log_msg(self.json_logging_filename)
 
        return all_features, stand_features

    def extract_discrete_time_window_features(self, data, nb_seconds, save=False, equalize = False):
        """
        Extract features from the data by discrete window.
        
        Parameters
        ----------  
        data : dict
            The dictionary that contains all data.
        nb_seconds : int
            The number of seconds to consider for each emotion segment.
        save: bool
            Whether to save the features to a file.
        equalize : bool
            Whether to equalize the number of samples for each emotion.

        Returns
        -------
        all_features : pd.DataFrame
            The extracted features.
        stand_features: pd.DataFrame
            The standardized features.
        """

        self.utilities._log_msg(self.json_logging_filename,"Extracting Discrete Time Window features.")

        pbar = tqdm(total=len(data.keys()), desc="Extracting Discrete Time Window features")

        features = dict()

        all_emotions = ['Joy', 'Shame', 'Frustration', 'Pride', 'Neutral']
        
        for subject_id in data.keys():

            pbar.set_description('Extracting Discrete Time Window features for subject {}'.format(subject_id))

            if 'Empatica' in self.desired_measurement:
                em_empatica = data[subject_id]['Empatica']['temp']["data"]["label"].unique()
                emotion_order = data[subject_id]['Empatica']['temp']['emotion_order']
            else:
                em_empatica = all_emotions
            if 'Audio' in self.desired_measurement:
                em_audio = data[subject_id]['Audio']['data']["label"].unique()
                emotion_order = data[subject_id]['Audio']['emotion_order']
            else:
                em_audio = all_emotions
            if 'Webcam' in self.desired_measurement:
                em_webcam = data[subject_id]['Webcam']["label"].keys()
                emotion_order = data[subject_id]['Webcam']['emotion_order']
            else:
                em_webcam = all_emotions
            if 'GoPro' in self.desired_measurement:
                em_gopro_frontal = data[subject_id]['GoPro']['Frontal']["label"].keys()
                em_gopro_lateral = data[subject_id]['GoPro']['Lateral']["label"].keys()
                em_gopro = list(set(em_gopro_frontal).intersection(em_gopro_lateral))
                emotion_order = data[subject_id]['GoPro']['Frontal']['emotion_order']
            else:
                em_gopro = all_emotions
            if 'FaceReader' in self.desired_measurement:
                em_facereader = data[subject_id]['FaceReader'].keys()
                emotion_order = em_facereader
            else:
                em_facereader = all_emotions

            # Create a mapping from element to index for the correct order of emotions
            index_map = {value: index for index, value in enumerate(emotion_order)}

            _emotions = list(set(em_empatica).intersection(em_audio, em_webcam, em_gopro, em_facereader))
            emotions = sorted(_emotions, key=lambda x: index_map[x])

            emotions.remove('Neutral')

            time_windows = self.utilities._extract_discrete_time_windows(data[subject_id], nb_seconds)

            inter_dict = dict()

            for emotion in emotions:
                
                try:
                    features_list = list()
                    merged_dict = {}
                    merged_dict["label"] = emotion

                    if 'Empatica' in self.desired_measurement:
                        for window in time_windows[emotion]['Empatica']:
                            intra_features = dict()
                            self._extract_empatica_features(intra_features, emotion, data[subject_id]['Empatica'], window)
                            features_list.append(intra_features)
                    if 'Audio' in self.desired_measurement:
                        for window in time_windows[emotion]['Audio']:
                            intra_features = dict()
                            self._extract_audio_features(intra_features, emotion, data[subject_id]['Audio'], window)
                            features_list.append(intra_features)
                    if 'Webcam' in self.desired_measurement:
                        for window in time_windows[emotion]['Webcam']:
                            intra_features = dict()
                            self._extract_webcam_features(intra_features, emotion, data[subject_id]['Webcam'], window, frame_skip=10)
                            features_list.append(intra_features)
                    if 'GoPro' in self.desired_measurement:
                        for window in time_windows[emotion]['GoPro']['Frontal']:
                            intra_features = dict()
                            self._extract_gopro_frontal_features(intra_features, emotion, data[subject_id]['GoPro']['Frontal'], window, frame_skip=10)
                            features_list.append(intra_features)
                        for window in time_windows[emotion]['GoPro']['Lateral']:
                            intra_features = dict()
                            self._extract_gopro_lateral_features(intra_features, emotion, data[subject_id]['GoPro']['Lateral'], window, frame_skip=10)
                            features_list.append(intra_features)
                    if 'FaceReader' in self.desired_measurement:
                        for window in time_windows[emotion]['FaceReader']:
                            intra_features = dict()
                            self._extract_face_reader_features(intra_features, emotion, data[subject_id]['FaceReader'], window)
                            features_list.append(intra_features)
                            
                    for feat in features_list:
                        for feature_name, feature_value in feat.items():
                            if feature_name in merged_dict:
                                merged_dict[feature_name].append(feature_value)
                            else:
                                merged_dict[feature_name] = [feature_value]
                    
                    inter_dict[emotion] = pd.DataFrame([merged_dict])  

                except Exception as e:
                    error_info = {
                        "subject_id": subject_id,
                        "emotion": emotion,
                        "error_message": str(e)
                    }
                    
                    self.utilities._log_msg(self.json_logging_filename, error_info)

                    continue
                
            features[subject_id] = pd.concat(inter_dict.values(), ignore_index=True)

            pbar.update(1)

        all_features = pd.concat(features.values(), ignore_index=True)
        orig_len = len(all_features)
        all_features.dropna(inplace=True)
        new_len = len(all_features)

        self.utilities._log_msg(self.json_logging_filename,f"    Dropped {orig_len - new_len} rows due to NaN values.")
        
        stand_features = all_features.copy()
        cols_to_standardize = stand_features.columns.difference(['label'])
        for col in cols_to_standardize:
            stand_features[col] = stand_features[col].apply(lambda x: self.scaler.fit_transform(np.array(x).reshape(-1, 1)).flatten().tolist())

        if equalize:
            all_features = self.utilities._equalize_emotion_segments(all_features)
            stand_features = self.utilities._equalize_emotion_segments(stand_features)
            self.utilities._log_msg(self.json_logging_filename,"    Equalized the number of samples for each emotion.")

        pbar.set_description('Extracting Discrete Time Window features')
        pbar.close()

        self.utilities._log_msg(self.json_logging_filename,"Extraction Completed.")

        if save:
            os.makedirs('computed_features', exist_ok=True)
            all_features.to_csv(f'computed_features/all_features_windows.csv', index=False)
            stand_features.to_csv(f'computed_features/stand_features_windows.csv', index=False)
            self.utilities._log_msg(self.json_logging_filename, "Discrete window features saved to computed_features folder.")

        self.utilities._log_msg(self.json_logging_filename)

        return all_features, stand_features

    def _extract_empatica_features(self, intra_features, emotion, data, window='All'):
        """
        Extract Empatica features.
        
        Parameters
        ----------
        intra_features : dict
            The dictionary that contains the intra features.    
        emotion : str
            The emotion.
        data : dict
            The dictionary that contains the necessary data.
        window : str or tuple
            The window to extract the features from.
        """
                    
        if 'temp' in data.keys():
            
            d = data['temp']["data"]['Temperature (°C)'][(data['temp']["data"]["label"] == emotion)]
            sample_rate = data['temp']['sample_rate']
            filtered_d = self.utilities._filter_empatica(d, sample_rate, 'temp')

            if window == 'All':
                preprocessed_temp_data  = np.array(filtered_d)
            else:
                start_point, end_point = window
                preprocessed_temp_data  = np.array(filtered_d[start_point:end_point])
        
            temp_features = self._extract_temp_features(preprocessed_temp_data)

            intra_features.update(temp_features)

        if 'eda' in data.keys():
            
            d = data['eda']['data']['EDA (μS)'][(data['eda']['data']["label"] == emotion)]
            sample_rate = data['eda']['sample_rate']
            filtered_d = self.utilities._filter_empatica(d, sample_rate, 'eda')

            if window == 'All':
                preprocessed_eda_data = np.array(filtered_d)
            else:
                start_point, end_point = window
                preprocessed_eda_data = np.array(filtered_d[start_point:end_point])

            eda_features = self._extract_eda_features(preprocessed_eda_data)

            intra_features.update(eda_features)
        
        if 'bvp' in data.keys():
            
            d = data['bvp']["data"]['BVP Value'][(data['bvp']["data"]["label"] == emotion)]
            sample_rate = data['bvp']['sample_rate']
            filtered_d = self.utilities._filter_empatica(d, sample_rate, 'bvp')

            if window == 'All':
                preprocessed_bvp_data = np.array(filtered_d)
            else:
                start_point, end_point = window
                preprocessed_bvp_data = np.array(filtered_d[start_point:end_point])

            bvp_features = self._extract_bvp_features(preprocessed_bvp_data)

            intra_features.update(bvp_features)
        
        if 'acc' in data.keys():
            for axis in ['X', 'Y', 'Z']:
                
                d = data["acc"]['data'][f'acc{axis}'][(data["acc"]['data']["label"] == emotion)]
                sample_rate = data['acc']['sample_rate']
                filtered_d = self.utilities._filter_empatica(d, sample_rate, 'acc')

                if window == 'All':
                    preprocessed_acc_data = np.array(filtered_d)
                else:
                    start_point, end_point = window
                    preprocessed_acc_data = np.array(filtered_d[start_point:end_point])

                acc_features = self._extract_acc_features(preprocessed_acc_data, axis)

                intra_features.update(acc_features)
        
        if 'hr' in data.keys():

            d = data["hr"]['data']['Heart Rate (bpm)'][(data["hr"]['data']["label"] == emotion)]
            sample_rate = data['hr']['sample_rate']
            filtered_d = self.utilities._filter_empatica(d, sample_rate, 'hr')

            if window == 'All':
                preprocessed_hr_data = np.array(filtered_d)
            else:
                start_point, end_point = window
                preprocessed_hr_data = np.array(filtered_d[start_point:end_point])       
            
            hr_features = self._extract_hr_features(preprocessed_hr_data)

            intra_features.update(hr_features) 

    def _extract_transcript_features(self, intra_features, emotion, data, nb_emotions):
        """
        Extract transcript features.
        
        Parameters
        ----------
        intra_features : dict
            The dictionary that contains the intra features.
        emotion : str
            The emotion.
        data : dict
            The dictionary that contains the necessary data.
        """
        
        transcript = data[emotion]["P"]

        transcript_preprocessed = self._preprocess_transcript(transcript)

        # Tokenization
        words = word_tokenize(transcript)
        sentences = sent_tokenize(transcript)

        # Text-based features
        word_count = len(words)
        sentence_count = len(sentences)
        avg_word_length = sum(len(word) for word in words) / word_count
        avg_sentence_length = word_count / sentence_count
        lexical_diversity = len(set(words)) / len(words)

        intra_features.update(
            {
                "word-count": word_count,
                "sentence-count": sentence_count,
                "word-length_mean": avg_word_length,
                "sentence-length_mean": avg_sentence_length,
                "lexical-diversity": lexical_diversity,
            }
        )
        
        blob = TextBlob(transcript_preprocessed) # Calculate sentiment polarity (LLM)
        sentiment = blob.sentiment.polarity

        intra_features.update({"sentiment-score": sentiment})
        
        # Vectorization
        vectorizer = CountVectorizer(stop_words='english')
        data_vectorized = vectorizer.fit_transform([transcript_preprocessed])

        # LDA model
        lda_model = LatentDirichletAllocation(n_components=nb_emotions, random_state=42) #Train a model to extract topics from the transcript
        lda_model.fit(data_vectorized)
        
        text_vectorized = vectorizer.transform([transcript_preprocessed])
        topic_distribution = lda_model.transform(text_vectorized)
        dominant_topic = topic_distribution.argmax()
        
        intra_features.update({"dominant-topic": dominant_topic})
        
        # N-grams (Bigrams as an example)
        bigrams = list(nltk.bigrams(word_tokenize(transcript_preprocessed)))

        intra_features.update({"bigram-count": len(bigrams)})

        # POS tagging
        pos_tags = pos_tag(word_tokenize(transcript_preprocessed))
        noun_count = sum(1 for _, tag in pos_tags if tag in ["NN", "NNS", "NNP", "NNPS"])

        intra_features.update({"noun-count": noun_count})

        # Emotion Words Count
        emotion_words = ["happy", "sad", "angry", "joyful", "fearful", "disgusted", "surprised", "calm"]
        emotion_word_count = sum(words.count(emotion_word) for emotion_word in emotion_words)

        intra_features.update({"emotion-word-count": emotion_word_count})

        # Question Count
        question_count = transcript.count("?")

        intra_features.update({"question-count": question_count})

        # Exclamation Count
        exclamation_count = transcript.count("!")

        intra_features.update({"exclamation-count": exclamation_count})
        
        # Adjective Count
        adjectives = ["JJ", "JJR", "JJS"]
        adj_count = sum(1 for _, tag in pos_tags if tag in adjectives)

        intra_features.update({"adj-count": adj_count})

        # Adverb Count
        adverbs = ["RB", "RBR", "RBS"]
        adv_count = sum(1 for _, tag in pos_tags if tag in adverbs)

        intra_features.update({"adv-count": adv_count})

        # Flesch Reading Ease Score
        flesch_score = textstat.flesch_reading_ease(transcript_preprocessed)

        intra_features.update({"flesch-reading-ease": flesch_score})

        # Modal Verbs Count
        modal_verbs = ["can", "could", "may", "might", "shall", "should", "will", "would", "must"]
        modal_count = sum(words.count(modal) for modal in modal_verbs)

        intra_features.update({"modal_count": modal_count})

        # Negation Words Count
        negations = ["not", "never", "none", "nothing", "nowhere", "neither", "noone", "nobody"]
        negation_count = sum(words.count(negation) for negation in negations)

        intra_features.update({"negation-count": negation_count})

        # Compound Words Count
        compound_count = sum(1 for word in words if "-" in word)

        intra_features.update({"compound-count": compound_count})

        # Unique Words Count
        unique_words_count = len(set(words))

        intra_features.update({"unique-words-count": unique_words_count})

        # POS Tagging for Pronouns
        pos_tags = pos_tag(words)
        pronoun_count = sum(1 for _, tag in pos_tags if tag in ["PRP", "PRP$"])

        intra_features.update({"pronoun-count": pronoun_count})
        
        #LLP
        inputs = self.tokenizer(transcript_preprocessed, return_tensors="pt", truncation=True, padding=True, max_length=512)
        with torch.no_grad():
            output = self.model(**inputs)
        # Mean pooling
        token_embeddings = output['last_hidden_state'][0]
        sentence_embedding = torch.mean(token_embeddings, dim=0)
        embedding = torch.mean(sentence_embedding).item()

        intra_features.update({"bert-aggregated-value": embedding})

        # Advanced Sentiment Analysis with VADER
        sia = SentimentIntensityAnalyzer()
        sentiment_scores = sia.polarity_scores(transcript_preprocessed)

        intra_features.update(
            {
                "vader-neg": sentiment_scores["neg"],
                "vader-neu": sentiment_scores["neu"],
                "vader-pos": sentiment_scores["pos"],
                "vader-compound": sentiment_scores["compound"],
            }
        )

        # Readability Score
        intra_features.update(
            {
                "flesch-kincaid-grade": textstat.flesch_kincaid_grade(transcript_preprocessed),
                "gunning-fog": textstat.gunning_fog(transcript_preprocessed)
            }
        )
    
    def _extract_audio_features(self, intra_features, emotion, data, window = 'All'):
        """
        Extract audio features.
        
        Parameters
        ----------
        intra_features : dict
            The dictionary that contains the intra features.
        emotion : str
            The emotion.
        data : dict
            The dictionary that contains the necessary data.
        window : str or tuple
            The window to extract the features from.
        """

        original_audio = np.array(data['data']['Amplitude'][data['data']['label'] == emotion])
        sample_rate = data['sample_rate']
        filtered_audio = self.utilities._filter_audio(original_audio, sample_rate)

        if window == 'All':
            audio = filtered_audio
        else:
            start_point, end_point = window
            audio = filtered_audio[start_point:end_point]
        
        # Calculate the Mel-frequency cepstral coefficients (MFCCs)
        mfccs = librosa.feature.mfcc(y=audio, sr=sample_rate, n_mfcc=13)
        intra_features["mfccs_mean"] = np.mean(mfccs)
        intra_features["mfccs_var"] = np.mean(np.var(mfccs, axis=1))
        intra_features["total-mfccs_var"] = np.sum(np.var(mfccs, axis=1)) # total variability across all the MFCCs.
        
        # Calculate the chroma features
        chroma_stft = librosa.feature.chroma_stft(y=audio, sr=sample_rate)
        intra_features["chroma-stft_mean"] = np.mean(chroma_stft)
        
        # Calculate the mel-scaled spectrogram
        mel_spectrogram = librosa.feature.melspectrogram(y=audio, sr=sample_rate)
        intra_features["mel-spectrogram_mean"] = np.mean(mel_spectrogram)
        
        # Calculate the tonnetz
        tonnetz = librosa.feature.tonnetz(y=audio, sr=sample_rate)
        intra_features["tonnetz_mean"] = np.mean(tonnetz)

        # Spectral Bandwidth
        spectral_bandwidth = librosa.feature.spectral_bandwidth(y=audio, sr=sample_rate)
        intra_features["spectral-bandwidth_mean"] = np.mean(spectral_bandwidth)

        # Spectral Roll-off
        spectral_rolloff = librosa.feature.spectral_rolloff(y=audio, sr=sample_rate)[0]
        intra_features["spectral-rolloff_mean"] = np.mean(spectral_rolloff)

        # Root Mean Square Energy
        rmse = librosa.feature.rms(y=audio)
        intra_features["rmse_mean"] = np.mean(rmse)

        # Spectral Flatness
        spectral_flatness = librosa.feature.spectral_flatness(y=audio)
        intra_features["spectral-flatness_mean"] = np.mean(spectral_flatness)

        # Pitch - using the pyin algorithm which is a probabilistic variant of YIN.
        pitches, magnitudes = librosa.core.piptrack(y=audio, sr=sample_rate)
        index_of_maxes = magnitudes.argmax(axis=0)
        pitch = pitches[index_of_maxes, range(magnitudes.shape[1])]
        intra_features["pitch_mean"] = np.mean(pitch[pitch > 0])  # This ignores zero pitch values which are placeholders

        # Envelope of the sound
        envelope = np.abs(librosa.onset.onset_strength(y=audio, sr=sample_rate))
        intra_features["envelope_mean"] = np.mean(envelope)
    
        # Zero Crossing Rate
        zcr = librosa.feature.zero_crossing_rate(audio)
        intra_features["zcr_mean"] = np.mean(zcr)

        # Spectral Contrast
        n_bands = 6  # Can be adjusted
        fmin = librosa.note_to_hz('C1')  # Starting frequency for contrast calculation

        # Calculate spectral contrast
        spectral_contrast = librosa.feature.spectral_contrast(y=audio, sr=sample_rate, n_bands=n_bands, fmin=fmin)
        intra_features["spectral-contrast_mean"] = np.mean(spectral_contrast)

        # Temporal Centroid (Sound Sharpness)
        amplitude_envelope = np.abs(librosa.stft(audio))
        amplitude_envelope_sum = np.sum(amplitude_envelope, axis=1)
        temporal_frames = np.arange(len(amplitude_envelope_sum))
        temporal_centroid = np.sum(temporal_frames * amplitude_envelope_sum) / np.sum(amplitude_envelope_sum)
        intra_features["temporal-centroid"] = temporal_centroid

        # Harmonic-to-Noise Ratio
        harmonic, noise = librosa.effects.hpss(audio)
        HNR = np.mean(harmonic) / np.mean(noise)
        intra_features["harmonic-to-noise-ratio"] = HNR

        # Spectral Centroid
        spectral_centroid = librosa.feature.spectral_centroid(y=audio, sr=sample_rate)
        intra_features["spectral-centroid_mean"] = np.mean(spectral_centroid)

        # Rhythm Features
        tempo, _ = librosa.beat.beat_track(y=audio, sr=sample_rate)
        intra_features["tempo"] = tempo

        # Poly Features
        poly_features = librosa.feature.poly_features(y=audio, sr=sample_rate)
        intra_features["poly-features_mean"] = np.mean(np.mean(poly_features, axis=1))

        # RMS Energy
        rms_energy = librosa.feature.rms(y=audio)
        intra_features["rms-energy_mean"] = np.mean(rms_energy)

        # Crest Factor
        crest_factor = max(audio) / np.sqrt(np.mean(np.square(audio)))
        intra_features["crest-factor"] = crest_factor

    def _extract_webcam_features(self, intra_features, emotion, data, window = 'All', frame_reduction_factor=1, frame_skip=30): ## Not used anymore as FaceReader is used now
        """
        Extract webcam features using VideoFileClip.

        Parameters
        ----------
        intra_features : dict
            The dictionary that contains the intra features.
        emotion : str
            The emotion.
        data : dict
            The dictionary that contains the necessary data.
        window : str or tuple
            The window to extract the features from.
        frame_reduction_factor : int
            The frame reduction factor.
        frame_skip : int
            The number of frame to skip.
        """

        if window == 'All':
            start_frame, end_frame = data['label'][emotion]
        else:
            start_frame, end_frame = window
        
        filepath = data['data_path']

        distance_features_dict = {}
        angle_features_dict = {}
        temporal_distance_change_dict = {}
        temporal_angle_change_dict = {}
        left_eye_area_list = []
        right_eye_area_list = []
        mouth_area_list = []
        vertical_face_tilt_list = []

        prev_distances = None
        prev_angles = None

        first_vertical_tilt = None

        blink_counter = 0
        frame_counter = 0

        clip = VideoFileClip(filepath)
        fps = clip.fps
        duration = (end_frame - start_frame) / fps

        for frame_nb in range(start_frame, end_frame + 1, frame_skip):

            frame = clip.get_frame(frame_nb / fps)

            # Speed up the process by reducing the size of the frame 
            frame_height, frame_width = frame.shape[:2]
            resized_width = int(frame_width / frame_reduction_factor)
            resized_height = int(frame_height / frame_reduction_factor)
            frame = cv2.resize(frame, (resized_width, resized_height))

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

            faces = self.detector(gray)

            for face in faces:
                landmarks = self.predictor(gray, face)
                landmarks = np.array([[p.x, p.y] for p in landmarks.parts()])

                distance_features_dict, temporal_distance_change_dict, prev_distances = self._extract_distance_landmarks_points(landmarks, distance_features_dict, temporal_distance_change_dict, prev_distances)
                angle_features_dict, temporal_angle_change_dict, prev_angles = self._extract_angle_landmarks_points(landmarks, angle_features_dict, temporal_angle_change_dict, prev_angles)
                left_eye_area_list, right_eye_area_list, mouth_area_list = self._extract_area_landmarks_points(landmarks, left_eye_area_list, right_eye_area_list, mouth_area_list)

                _vertical_tilt = self._extract_face_tilt(landmarks)

                if first_vertical_tilt is None:
                    first_vertical_tilt = _vertical_tilt

                vertical_tilt = first_vertical_tilt - _vertical_tilt

                vertical_face_tilt_list.append(int(np.mean(vertical_tilt)))

                blink_counter, frame_counter = self._count_blinks(landmarks, blink_counter, frame_counter)

        clip.reader.close()

        for name, distances in distance_features_dict.items():
            intra_features[f'{name}_mean_dist'] = np.mean(distances)
            intra_features[f'{name}_std_dist'] = np.std(distances)
            intra_features[f'{name}_var_dist'] = np.var(distances)

        for name, angles in angle_features_dict.items():
            intra_features[f'{name}_mean_angle'] = np.mean(angles)
            intra_features[f'{name}_std_angle'] = np.std(angles)
            intra_features[f'{name}_var_angle'] = np.var(angles)

        intra_features['leftEye_mean_area'] = np.mean(left_eye_area_list)
        intra_features['leftEye_std_area'] = np.std(left_eye_area_list)
        intra_features['leftEye_var_area'] = np.var(left_eye_area_list)

        intra_features['rightEye_mean_area'] = np.mean(right_eye_area_list)
        intra_features['rightEye_std_area'] = np.std(right_eye_area_list)
        intra_features['rightEye_var_area'] = np.var(right_eye_area_list)

        intra_features['mouth_mean_area'] = np.mean(mouth_area_list)
        intra_features['mouth_std_area'] = np.std(mouth_area_list)
        intra_features['mouth_var_area'] = np.var(mouth_area_list)

        intra_features['verticalFaceTilt_mean'] = np.mean(vertical_face_tilt_list)
        intra_features['verticalFaceTilt_std'] = np.std(vertical_face_tilt_list)
        intra_features['verticalFaceTilt_var'] = np.var(vertical_face_tilt_list)
        
        intra_features['blink_rate'] = (blink_counter / duration) * 60 #Blinks per minute

    def _extract_gopro_frontal_features(self, intra_features, emotion, data, window='All', frame_reduction_factor=1, frame_skip=30):
        """
        Extract Features from the cut GoPro frontal videos

        Parameters
        ----------
        intra_features : dict
            The dictionary that contains the intra features.
        emotion : str
            The emotion.
        data : dict
            The dictionary that contains the necessary data.
        window : str or tuple
            The window to extract the features from.
        frame_reduction_factor : int
            The frame reduction factor.
        frame_skip : int
            The frame skip.
        """

        if window == 'All':
            start_frame, end_frame = data['label'][emotion]
        else:
            start_frame, end_frame = window

        filepath = data['data_path']

        cap = cv2.VideoCapture(filepath)

        cap.set(cv2.CAP_PROP_POS_FRAMES, start_frame)

        frame_counter = 0  

        shoulder_distances = []

        right_wrist_positions = []
        right_wrist_velocities = []

        left_wrist_positions = []
        left_wrist_velocities = []

        symmetry_measurements = []

        while cap.isOpened():
            ret, frame = cap.read()

            if not ret :
                break

            current_frame = int(cap.get(cv2.CAP_PROP_POS_FRAMES))

            if current_frame > end_frame:
                break
                
            if frame_counter % frame_skip != 0:
                frame_counter += 1
                continue

            frame_counter += 1

            frame_height, frame_width = frame.shape[:2]
            resized_width = int(frame_width / frame_reduction_factor)
            resized_height = int(frame_height / frame_reduction_factor)
            frame = cv2.resize(frame, (resized_width, resized_height))

            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            results = self.pose.process(frame_rgb)

            if results.pose_landmarks:
                landmark_list = results.pose_landmarks.landmark

                left_shoulder = [landmark_list[self.mp_pose.PoseLandmark.LEFT_SHOULDER.value].x, landmark_list[self.mp_pose.PoseLandmark.LEFT_SHOULDER.value].y]
                right_shoulder = [landmark_list[self.mp_pose.PoseLandmark.RIGHT_SHOULDER.value].x, landmark_list[self.mp_pose.PoseLandmark.RIGHT_SHOULDER.value].y]
                left_wrist = [landmark_list[self.mp_pose.PoseLandmark.LEFT_WRIST.value].x, landmark_list[self.mp_pose.PoseLandmark.LEFT_WRIST.value].y]
                right_wrist = [landmark_list[self.mp_pose.PoseLandmark.RIGHT_WRIST.value].x, landmark_list[self.mp_pose.PoseLandmark.RIGHT_WRIST.value].y]

                shoulder_dist = np.linalg.norm(
                    np.array(left_shoulder) - np.array(right_shoulder)
                )

                shoulder_distances.append(np.round(shoulder_dist,2))

                right_wrist_position = np.array(right_wrist)
                right_wrist_positions.append(np.round(right_wrist_position,2))

                left_wrist_position = np.array(left_wrist)
                left_wrist_positions.append(np.round(left_wrist_position,2))

                if len(right_wrist_positions) > 1:
                    right_wrist_velocity = np.linalg.norm(right_wrist_positions[-1] - right_wrist_positions[-2])
                    right_wrist_velocities.append(np.round(right_wrist_velocity,3))

                if len(left_wrist_positions) > 1:
                    left_wrist_velocity = np.linalg.norm(left_wrist_positions[-1] - left_wrist_positions[-2])
                    left_wrist_velocities.append(np.round(left_wrist_velocity,3))

                symmetry = np.abs(left_wrist[0] - right_wrist[0])
                symmetry_measurements.append(np.round(symmetry,3))
                
        cap.release()
                
        mean_shoulder_distance = np.mean(shoulder_distances) if shoulder_distances else None
        std_shoulder_distance = np.std(shoulder_distances) if shoulder_distances else None
        var_shoulder_distance = np.var(shoulder_distances) if shoulder_distances else None

        mean_right_wrist_velocity = np.mean(right_wrist_velocities) if right_wrist_velocities else None
        std_right_wrist_velocity = np.std(right_wrist_velocities) if right_wrist_velocities else None
        var_right_wrist_velocity = np.var(right_wrist_velocities) if right_wrist_velocities else None

        mean_left_wrist_velocity = np.mean(left_wrist_velocities) if left_wrist_velocities else None 
        std_left_wrist_velocity = np.std(left_wrist_velocities) if left_wrist_velocities else None
        var_left_wrist_velocity = np.var(left_wrist_velocities) if left_wrist_velocities else None

        mean_symmetry = np.mean(symmetry_measurements) if symmetry_measurements else None
        std_symmetry = np.std(symmetry_measurements) if symmetry_measurements else None
        var_symmetry = np.var(symmetry_measurements) if symmetry_measurements else None

        intra_features.update({
            'shoulderDistance_mean': mean_shoulder_distance,
            'shoulderDistance_std': std_shoulder_distance,
            'shoulderDistance_var': var_shoulder_distance,
            'rightWristVelocity_mean': mean_right_wrist_velocity,
            'rightWristVelocity_std': std_right_wrist_velocity,
            'rightWristVelocity_var': var_right_wrist_velocity,
            'leftWristVelocity_mean': mean_left_wrist_velocity,
            'leftWristVelocity_std': std_left_wrist_velocity,
            'leftWristVelocity_var': var_left_wrist_velocity,
            'symmetry_mean': mean_symmetry,
            'symmetry_std': std_symmetry,
            'symmetry_var': var_symmetry
        })

    def _extract_gopro_lateral_features(self, intra_features, emotion, data, window='All', frame_reduction_factor=1, frame_skip=30):
        """
        Extract Features from the cut GoPro lateral videos

        Parameters
        ----------
        intra_features : dict
            The dictionary that contains the intra features.
        emotion : str
            The emotion.
        data : dict
            The dictionary that contains the necessary data.
        window : str or tuple
            The window to extract the features from.
        frame_reduction_factor : int
            The frame reduction factor.
        frame_skip : int
            The frame skip.
        """

        if window == 'All':
            start_frame, end_frame = data['label'][emotion]
        else:
            start_frame, end_frame = window

        filepath = data['data_path']

        cap = cv2.VideoCapture(filepath)

        cap.set(cv2.CAP_PROP_POS_FRAMES, start_frame)

        frame_counter = 0  

        trunk_inclinations = []
        head_inclinations = []

        first_trunk_angle = None
        first_head_angle = None

        while cap.isOpened():
            ret, frame = cap.read()

            if not ret:
                break

            current_frame = int(cap.get(cv2.CAP_PROP_POS_FRAMES))

            if current_frame > end_frame:
                break
                
            if frame_counter % frame_skip != 0:
                frame_counter += 1
                continue

            frame_counter += 1

            frame_height, frame_width = frame.shape[:2]
            resized_width = int(frame_width / frame_reduction_factor)
            resized_height = int(frame_height / frame_reduction_factor)
            frame = cv2.resize(frame, (resized_width, resized_height))

            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            results = self.pose.process(frame_rgb)

            if results.pose_landmarks:
                landmarks = results.pose_landmarks.landmark
                
                left_shoulder = [landmarks[self.mp_pose.PoseLandmark.LEFT_SHOULDER.value].x, landmarks[self.mp_pose.PoseLandmark.LEFT_SHOULDER.value].y]
                right_shoulder = [landmarks[self.mp_pose.PoseLandmark.RIGHT_SHOULDER.value].x, landmarks[self.mp_pose.PoseLandmark.RIGHT_SHOULDER.value].y]
                left_hip = [landmarks[self.mp_pose.PoseLandmark.LEFT_HIP.value].x, landmarks[self.mp_pose.PoseLandmark.LEFT_HIP.value].y]
                right_hip = [landmarks[self.mp_pose.PoseLandmark.RIGHT_HIP.value].x, landmarks[self.mp_pose.PoseLandmark.RIGHT_HIP.value].y]
                ear = [landmarks[self.mp_pose.PoseLandmark.LEFT_EAR.value].x, landmarks[self.mp_pose.PoseLandmark.LEFT_EAR.value].y]
               
                shoulder_mid = np.mean([left_shoulder, right_shoulder], axis=0)
                hip_mid = np.mean([left_hip, right_hip], axis=0)

                if first_trunk_angle is None and first_head_angle is None:
                    first_trunk_angle = self._compute_trunk_inclination(shoulder_mid, hip_mid)
                    first_head_angle = self._compute_head_inclination(ear, shoulder_mid)
                    
                trunk_inclination_angle =  first_trunk_angle - self._compute_trunk_inclination(shoulder_mid, hip_mid)
                head_inclination_angle = first_head_angle - self._compute_head_inclination(ear, shoulder_mid)

                trunk_inclinations.append(trunk_inclination_angle)
                head_inclinations.append(head_inclination_angle)

        cap.release()

        mean_trunk_inclination = np.mean(trunk_inclinations) if trunk_inclinations else None
        std_trunk_inclination = np.std(trunk_inclinations) if trunk_inclinations else None
        var_trunk_inclination = np.var(trunk_inclinations) if trunk_inclinations else None

        mean_head_inclination = np.mean(head_inclinations) if head_inclinations else None
        std_head_inclination = np.std(head_inclinations) if head_inclinations else None
        var_head_inclination = np.var(head_inclinations) if head_inclinations else None

        intra_features.update({
            'trunkInclination_mean': mean_trunk_inclination,
            'trunkInclination_std': std_trunk_inclination,
            'trunkInclination_var': var_trunk_inclination,
            'headInclination_mean': mean_head_inclination,
            'headInclination_std': std_head_inclination,
            'headInclination_var': var_head_inclination
        })

    def _extract_sre_features(self, intra_features, emotion, data):
        """
        Extract the self reported emotion score

        Parameters
        ----------
        intra_features : dict
            The dictionary that contains the intra features.
        emotion : str
            The emotion.
        data : dict
            The dictionary that contains the necessary data.
        """

        score = 0

        for category in data[emotion].keys():
            if category in self.positive_emotions:
                score += self.scores[data[emotion][category]]
            elif category in self.negative_emotions:
                score -= self.scores[data[emotion][category]]
        
        intra_features.update({
        'sre-score': score
        })

    def _extract_face_reader_features(self, intra_features, emotion, data, window='All'):
        """ 
        Extract the face reader features

        Parameters
        ----------
        intra_features : dict
            The dictionary that contains the intra features.
        emotion : str
            The emotion.
        data : dict
            The dictionary that contains the necessary data.
        window : str or tuple
            The window to extract the features from.
        """

        for feature in data[emotion].keys():

            if window != 'All':
                start_idx, end_idx = window
                d = data[emotion][feature][start_idx:end_idx]
            else:
                d = data[emotion][feature]

            intra_features.update({
                f'{feature}_mean': np.mean(d),
                f'{feature}_std': np.std(d),
                f'{feature}_var': np.var(d)
            })

    def _extract_temp_features(self, preprocessed_temp_data):
        """
        Extract temperature features.

        Parameters
        ----------
        preprocessed_temp_data : np.array
            The preprocessed temperature data.
        
        Returns
        -------
        dict
            A dictionary containing the temperature features.
        """
        mean_temp = np.mean(preprocessed_temp_data)
        std_temp = np.std(preprocessed_temp_data)
        temp_change = preprocessed_temp_data[-1] - preprocessed_temp_data[0]
        max_temp = np.max(preprocessed_temp_data)
        min_temp = np.min(preprocessed_temp_data)
        median_temp = np.median(preprocessed_temp_data)
        temp_range = max_temp - min_temp

        return {
            'temp_mean': mean_temp,
            'temp_std': std_temp,
            'temp_change': temp_change,
            'temp_max': max_temp,
            'temp_min': min_temp,
            'temp_median': median_temp,
            'temp_range': temp_range
        }
    
    def _extract_eda_features(self, preprocessed_eda_data):
        """
        Extract EDA features.

        Parameters
        ----------
        preprocessed_eda_data : np.array
            The preprocessed EDA data.

        Returns
        -------
        dict
            A dictionary containing the EDA features.
        """
            
        # Time-Domain Features
        mean_eda = preprocessed_eda_data.mean()
        std_eda = preprocessed_eda_data.std()
        eda_peaks, _ = find_peaks(preprocessed_eda_data)
        number_of_peaks = len(eda_peaks)
        peak_amplitudes = preprocessed_eda_data[eda_peaks]

        if number_of_peaks == 0:
            mean_peak_amplitude = 0
            max_peak_amplitude = 0
            min_peak_amplitude = 0
        else:
            mean_peak_amplitude = np.mean(peak_amplitudes)
            max_peak_amplitude = np.max(peak_amplitudes)
            min_peak_amplitude = np.min(peak_amplitudes)

        # Additional Statistical Measures
        median_eda = np.median(preprocessed_eda_data)
        eda_range = np.max(preprocessed_eda_data) - np.min(preprocessed_eda_data)

        return {
            'eda_mean': mean_eda,
            'eda_std': std_eda,
            'eda_number_of_peaks': number_of_peaks,
            'eda_mean_peak_amplitude': mean_peak_amplitude,
            'eda_max_peak_amplitude': max_peak_amplitude,
            'eda_min_peak_amplitude': min_peak_amplitude,
            'eda_median': median_eda,
            'eda_range': eda_range
        }
            
    def _extract_bvp_features(self, preprocessed_bvp_data):
        """
        Extract BVP features.

        Parameters
        ----------
        preprocessed_bvp_data : np.array
            The preprocessed BVP data.

        Returns
        -------
        dict
            A dictionary containing the BVP features.
        """

        # Time-Domain Features
        mean_bvp = preprocessed_bvp_data.mean()
        std_bvp = preprocessed_bvp_data.std()
        bvp_peaks, _ = find_peaks(preprocessed_bvp_data)
        number_of_bvp_peaks = len(bvp_peaks)

        if number_of_bvp_peaks == 0:
            peak_intervals = 0
            mean_peak_interval = 0
        else:
            peak_intervals = np.diff(bvp_peaks)
            mean_peak_interval = np.mean(peak_intervals)

        # Additional Statistical Measures
        median_bvp = np.median(preprocessed_bvp_data)
        bvp_range = np.max(preprocessed_bvp_data) - np.min(preprocessed_bvp_data)

        return {
            'bvp_mean': mean_bvp,
            'bvp_std': std_bvp,
            'bvp_number_of_peaks': number_of_bvp_peaks,
            'bvp_mean_peak_interval': mean_peak_interval,
            'bvp_median': median_bvp,
            'bvp_range': bvp_range,
        }
    
    def _extract_acc_features(self, preprocessed_acc_data, axis):
        """
        Extract ACC features.

        Parameters
        ----------
        preprocessed_acc_data : np.array
            The preprocessed ACC data.
        
        axis : str
            The axis of the ACC data.
        """

        mean_acc = preprocessed_acc_data.mean()
        std_acc = preprocessed_acc_data.std()
        peak_acc = preprocessed_acc_data.max()
        energy_acc = np.sqrt((preprocessed_acc_data**2).sum())
        median_acc = np.median(preprocessed_acc_data)
        acc_range = np.max(preprocessed_acc_data) - np.min(preprocessed_acc_data)

        return {
            f'acc{axis}_mean': mean_acc,
            f'acc{axis}_std': std_acc,
            f'acc{axis}_peak': peak_acc,
            f'acc{axis}_energy': energy_acc,
            f'acc{axis}_median': median_acc,
            f'acc{axis}_range': acc_range
        }
    
    def _extract_hr_features(self, preprocessed_hr_data):
        """
        Extract HR features.

        Parameters
        ----------
        preprocessed_hr_data : np.array
            The preprocessed HR data.

        Returns
        -------
        dict
            A dictionary containing the HR features.
        """

        # Time-Domain Features
        mean_hr = preprocessed_hr_data.mean()
        std_hr = preprocessed_hr_data.std()
        max_hr = preprocessed_hr_data.max()
        min_hr = preprocessed_hr_data.min()
        gradient_hr = (preprocessed_hr_data[-1] - preprocessed_hr_data[0]) / len(preprocessed_hr_data)
        hrv_std = np.std(np.diff(preprocessed_hr_data))

        # Additional Statistical Measures
        median_hr = np.median(preprocessed_hr_data)
        hr_range = np.max(preprocessed_hr_data) - np.min(preprocessed_hr_data)

        return {
            'hr_mean': mean_hr,
            'hr_std': std_hr,
            'hr_max': max_hr,
            'hr_min': min_hr,
            'hr_gradient': gradient_hr,
            'hr_variation_std': hrv_std,
            'hr_median': median_hr,
            'hr_range': hr_range
        }

    def _calculate_welch_nperseg(self, data_length, min_nperseg=50, max_nperseg=256):
        """
        Calculate an appropriate nperseg value for Welch's method.

        Parameters
        ----------
        data_length : int
            The length of the data.
        min_nperseg : int
            The minimum nperseg value to use.
        max_nperseg : int
            The maximum nperseg value to use.

        Returns
        -------
        int
            The calculated nperseg value.
        """
        if data_length < min_nperseg:
            return data_length  
        return min(max_nperseg, max(min_nperseg, data_length // 2)) 
        
    def _preprocess_transcript(self, text):
        """
        Preprocess the input text.
        
        Parameters:
        ----------
        text : str
            The text to preprocess.

        Returns:
        -------
        str: 
            Preprocessed text.
        """

        # Lowercasing
        text = text.lower()

        # Remove punctuation
        text = text.translate(str.maketrans('', '', string.punctuation))

        # Tokenization
        tokens = word_tokenize(text)

        # Remove stopwords
        stop_words = set(stopwords.words('english'))
        tokens = [word for word in tokens if word not in stop_words]

        # Lemmatization
        lemmatizer = WordNetLemmatizer()
        tokens = [lemmatizer.lemmatize(word) for word in tokens]

        # Re-join tokens into a single string
        return ' '.join(tokens)
    
    def _calculate_landmarks_distances(self, landmarks, landmark_indices):
        """
        Compute the distances between the landmarks.

        Parameters
        ----------
        landmarks : np.ndarray
            The landmarks.
        landmark_indices : list
            The indices of the landmarks to use.
        
        Returns
        -------
        list
            The distances between the landmarks.
        """

        distances = []

        for indices, name in landmark_indices:
            point1 = landmarks[indices[0]]
            point2 = landmarks[indices[1]]
            distance = np.linalg.norm(point1 - point2)
            distances.append((name, distance))

        return distances

    def _calculate_landmarks_angles(self, landmarks, angle_indices):
        """
        Compute the angles between the landmarks.

        Parameters
        ----------
        landmarks : np.ndarray
            The landmarks.
        angle_indices : list
            The indices of the landmarks to use.

        Returns
        -------
        list
            The angles between the landmarks.
        """

        angles = []

        for indices, name in angle_indices:
            a = landmarks[indices[0]]
            b = landmarks[indices[1]]
            c = landmarks[indices[2]]
            ba = a - b
            bc = c - b
            norm_ba = np.linalg.norm(ba)
            norm_bc = np.linalg.norm(bc)

            if norm_ba > 0 and norm_bc > 0:
                cosine_angle = np.dot(ba, bc) / (norm_ba * norm_bc)

                # Ensure cosine_angle is within valid range
                cosine_angle = max(-1.0, min(1.0, cosine_angle))

                angle = np.arccos(cosine_angle)  # radians
                angles.append((name, np.degrees(angle)))  # degrees
                
        return angles
    
    def _extract_distance_landmarks_points(self, landmarks, distance_features_dict, temporal_distance_change_dict, prev_distances):
        """
        Extract the landmarks points distances .

        Parameters
        ----------
        landmarks : np.ndarray
            The landmarks.
        distance_features_dict : dict
            The dictionary that contains the distance features.
        temporal_distance_change_dict : dict
            The dictionary that contains the temporal distance change features.
        prev_distances : dict
            The dictionary that contains the previous distances computed from the previous frame.
        
        Returns
        -------
        distance_features_dict : dict
            The dictionary that contains the distance features.
        temporal_distance_change_dict : dict
            The dictionary that contains the temporal distance change features.
        prev_distances : dict
            The dictionary that contains the previous distances.
        
        """

        distances = self._calculate_landmarks_distances(landmarks, self.landmark_indices_to_distance)

        for name, distance in distances:
            if name not in distance_features_dict:
                distance_features_dict[name] = []
                temporal_distance_change_dict[name] = []
            distance_features_dict[name].append(distance)
            if prev_distances is not None:
                temporal_distance_change_dict[name].append(abs(distance - prev_distances[name]))

        prev_distances = {name: distance for name, distance in distances}

        return distance_features_dict, temporal_distance_change_dict, prev_distances
    
    def _extract_angle_landmarks_points(self, landmarks, angle_features_dict, temporal_angle_change_dict, prev_angles):
        """
        Extract the landmarks points.

        Parameters
        ----------
        landmarks : np.ndarray
            The landmarks.
        angle_features_dict : dict
            The dictionary that contains the angle features.
        temporal_angle_change_dict : dict
            The dictionary that contains the temporal angle change features.
        prev_angles : dict
            The dictionary that contains the previous angles.
        
        Returns
        -------
        angle_features_dict : dict
            The dictionary that contains the angle features.
        temporal_angle_change_dict : dict
            The dictionary that contains the temporal angle change features.
        prev_angles : dict
            The dictionary that contains the previous angles.
        
        """
        angles = self._calculate_landmarks_angles(landmarks, self.landmark_indices_to_angle)

        for name, angle in angles:
            if name not in angle_features_dict:
                angle_features_dict[name] = []
                temporal_angle_change_dict[name] = []
            angle_features_dict[name].append(angle)
            if prev_angles is not None:
                temporal_angle_change_dict[name].append(abs(angle - prev_angles[name]))

        prev_angles = {name: angle for name, angle in angles}

        return angle_features_dict, temporal_angle_change_dict, prev_angles
    
    def _extract_area_landmarks_points(self, landmarks, left_eye_area_list, right_eye_area_list, mouth_area_list):
        """
        Extract the area between landmarks points.

        Parameters
        ----------
        landmarks : np.ndarray
            The landmarks.
        left_eye_area_list : list
            The list that contains the left eye areas of the previous frames.
        right_eye_area_list : list
            The list that contains the right eye areas of the previous frames.
        mouth_area_list : list
            The list that contains the mouth area of the previous frames.
        
        Returns
        -------
        left_eye_area_list : list
            The list that contains the left eye areas of the previous and current frames.
        right_eye_area_list : list
            The list that contains the right eye areas of the previous and current frames.
        mouth_area_list : list
            The list that contains the mouth areas of the previous and current frames.
        """

        left_eye_points = landmarks[36:42]
        right_eye_points = landmarks[42:48]
        mouth_points = landmarks[60:67]

        left_eye_area = cv2.contourArea(np.array(left_eye_points))
        right_eye_area = cv2.contourArea(np.array(right_eye_points))

        left_eye_area_list.append(left_eye_area)
        right_eye_area_list.append(right_eye_area)

        mouth_area = cv2.contourArea(np.array(mouth_points))
        mouth_area_list.append(mouth_area)

        return left_eye_area_list, right_eye_area_list, mouth_area_list
    
    def _extract_face_tilt(self, landmarks):
        """
        Compute the vertical tilt of the face.

        Parameters
        ----------
        landmarks : np.ndarray
            The landmarks.
        
        Returns
        -------
        vertical_tilt : float
            The vertical tilt of the face.
        """

        left_eye_indices = [36, 37, 38, 39, 40, 41]
        right_eye_indices = [42, 43, 44, 45, 46, 47]
        mouth_indices = [48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59]

        left_eye_center = np.mean(landmarks[left_eye_indices], axis=0)
        right_eye_center = np.mean(landmarks[right_eye_indices], axis=0)

        mouth_center = np.mean(landmarks[mouth_indices], axis=0)

        vertical_tilt = np.arctan2(mouth_center[1] - (left_eye_center[1] + right_eye_center[1]) / 2,
                                    mouth_center[0] - (left_eye_center[0] + right_eye_center[0]) / 2) * 180 / np.pi

        return vertical_tilt
    
    def _extract_eye_aspect_ratio(self, eye_points):
        """
        Compute the eye aspect ratio (EAR).

        Parameters
        ----------
        eye_points : np.ndarray
            The eye points.

        Returns
        -------
        ear : float
            The EAR.
        """

        A = np.linalg.norm(eye_points[1] - eye_points[5])
        B = np.linalg.norm(eye_points[2] - eye_points[4])
        C = np.linalg.norm(eye_points[0] - eye_points[3])
        ear = (A + B) / (2.0 * C)

        return ear

    def _count_blinks(self, landmarks, blink_counter, frame_counter, ear_thres = 0.2, ear_consec_frames = 2):
        """
        Compute the number of blinks.

        Parameters
        ----------
        landmarks : np.ndarray
            The landmarks.
        blink_counter : int
            The number of blinks recorded in the previous frames.
        frame_counter : int
            The frame counter.
        ear_thres : float
            The EAR threshold (minimum value).
        ear_consec_frames : int
            Number of consecutive frames to consider a blink.

        Returns
        ------- 
        blink_counter : int
            The number of blinks recorded in the previous and current frames.
        """

        leftEye = landmarks[36:42]
        rightEye = landmarks[42:48]
        
        leftEAR = self._extract_eye_aspect_ratio(leftEye)
        rightEAR = self._extract_eye_aspect_ratio(rightEye)
        ear = (leftEAR + rightEAR) / 2.0

        if ear < ear_thres:
            frame_counter += 1
        else:
            if frame_counter >= ear_consec_frames:
                blink_counter += 1
            frame_counter = 0

        return blink_counter, frame_counter
    
    def _compute_arm_angle(self, landmark1, landmark2, landmark3):
        """
        Compute the arm angle from given landmarks

        Parameters
        ----------
        landmark1 : np.ndarray
            The first landmark.
        landmark2 : np.ndarray
            The second landmark.
        landmark3 : np.ndarray
            The third landmark.

        Returns
        -------
        angle : float
            The arm angle.
        
        """

        p1 = np.array(landmark1)
        p2 = np.array(landmark2)
        p3 = np.array(landmark3)

        v1 = p1 - p2
        v2 = p3 - p2

        angle = np.math.atan2(np.linalg.det([v1, v2]), np.dot(v1, v2))

        return np.degrees(angle)

    def _compute_trunk_inclination(self, shoulder_mid, hip_mid):
        """
        Compute the trunk inclination.

        Parameters
        ----------
        shoulder_mid : np.ndarray
            The smoothed shoulder midpoint.
        hip_mid : np.ndarray
            The smoothed hip midpoint.

        Returns
        -------
        angle : float
            The trunk inclination angle.
        """

        dy = hip_mid[1] - shoulder_mid[1]
        dx = hip_mid[0] - shoulder_mid[0]
        radians = np.arctan2(dy, dx)
        angle = np.abs(radians * 180.0 / np.pi)

        return angle
    
    def _compute_head_inclination(self, point1, point2):
        """
        Calculate the angle with respect to the vertical axis.

        Parameters
        ----------
        point1 : np.ndarray
            The first point.
        point2 : np.ndarray
            The second point.
        
        Returns
        ------- 
        angle : float
            The angle with respect to the vertical axis.
        """

        dy = point2[1] - point1[1]
        dx = point2[0] - point1[0]
        angle = np.arctan2(dy, dx) * 180.0 / np.pi

        return angle
    
    def visualize_webcam_features(self, data, subject, emotion, draw = 'All', frame_skip = 30):
        """
        Visualize the Webcam features.

        Parameters
        ----------
        data : dict
            The dictionary that contains the necessary data.
        subject : str
            The subject ID.
        emotion : str
            The emotion.
        draw : str or tuple
            Determines what to draw on the frames
        frame_skip : int
            The frame skip.
        """

        blink_counter = 0
        frame_counter = 0

        start_frame, end_frame = data[subject]['Webcam']['label'][emotion]
        data_path = data[subject]['Webcam']['data_path']

        cap = cv2.VideoCapture(data_path)

        # Colors for drawing segments
        segment_colors = {
            'mouthCornerToMouthCorner': (0, 0, 255),  
            'upperLipToLowerLip': (0, 0, 255), 
            'leftMouthCornerToBottomLip': (0, 0, 255),
            'rightMouthCornerToBottomLip': (0, 0, 255), 
            'leftEyebrowToRightEyebrow': (0, 0, 255),
            'leftEyebrowToLeftEyeInner': (0, 0, 255),
            'rightEyebrowToRightEyeInner': (0, 0, 255),
            'leftEyebrowToLeftEyeOuter': (0, 0, 255),
            'rightEyebrowToRightEyeOuter': (0, 0, 255),
            'bottomLipToChin': (0, 0, 255),

            'leftEye-nose-rightEye': (0, 255, 0),  
            'chin-nose-leftEye': (0, 255, 0),  
            'chin-nose-rightEyeOuter': (0, 255, 0),  
            'leftEyebrow-middleForehead-rightEyebrow': (0, 255, 0),  
            'leftMouthCorner-bottomLip-rightMouthCorner': (0, 255, 0),
        }

        cap.set(cv2.CAP_PROP_POS_FRAMES, start_frame)

        first_vertical_tilt = None

        frame_counter = 0

        while cap.isOpened():

            ret, frame = cap.read()
            
            if not ret:
                break

            current_frame = int(cap.get(cv2.CAP_PROP_POS_FRAMES))

            if current_frame > end_frame:
                break

            if frame_counter % frame_skip != 0:
                frame_counter += 1
                continue

            frame_counter += 1

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

            faces = self.detector(gray)

            for face in faces:

                landmarks = self.predictor(gray, face)

                for n in range(landmarks.num_parts):
                    point = (landmarks.part(n).x, landmarks.part(n).y)

                    cv2.circle(frame, point, 1, (255,255,255), -1)  

                for indices, segment_name in self.landmark_indices_to_distance:
                    (x1, y1) = (landmarks.part(indices[0]).x, landmarks.part(indices[0]).y)
                    (x2, y2) = (landmarks.part(indices[1]).x, landmarks.part(indices[1]).y)

                    if draw == 'All' or 'distances' in draw:
                        cv2.line(frame, (x1, y1), (x2, y2), segment_colors[segment_name], 1)

                for indices, segment_name in self.landmark_indices_to_angle:
                    p1 = (landmarks.part(indices[0]).x, landmarks.part(indices[0]).y)
                    p2 = (landmarks.part(indices[1]).x, landmarks.part(indices[1]).y)
                    p3 = (landmarks.part(indices[2]).x, landmarks.part(indices[2]).y)

                    if draw == 'All' or 'angles' in draw:
                        cv2.line(frame, p1, p2, segment_colors[segment_name], 1)
                        cv2.line(frame, p2, p3, segment_colors[segment_name], 1)

                landmarks = np.array([[p.x, p.y] for p in landmarks.parts()])

                left_eye_points = landmarks[36:42]
                right_eye_points = landmarks[42:48]
                mouth_points = landmarks[60:67]
                
                left_eye_contour = cv2.convexHull(np.array(left_eye_points))
                right_eye_contour = cv2.convexHull(np.array(right_eye_points))
                mouth_contour = cv2.convexHull(np.array(mouth_points))

                if draw == 'All' or 'areas' in draw:
                    cv2.drawContours(frame, [left_eye_contour], -1, (0, 255, 255), 1)  
                    cv2.drawContours(frame, [right_eye_contour], -1, (0, 255, 255), 1)             
                    cv2.drawContours(frame, [mouth_contour], -1, (0, 255, 255), 1) 

                _vertical_tilt = self._extract_face_tilt(landmarks)

                if first_vertical_tilt is None:
                    first_vertical_tilt = _vertical_tilt

                vertical_tilt = first_vertical_tilt - _vertical_tilt

                blink_counter, frame_counter = self._count_blinks(landmarks, blink_counter, frame_counter)

                cv2.putText(frame, f'Vertical Tilt: {int(vertical_tilt):.2f} degrees', (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 1)
                cv2.putText(frame, "Blinks: {}".format(blink_counter), (10, 90), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 1)

            cv2.imshow('Video', frame)

            if cv2.waitKey(1) & 0xFF == 27:  # Press 'ESC' to exit
                break

        cap.release()
        cv2.destroyAllWindows()
    
    def visualize_frontal_gopro_features(self, data, subject, emotion, frame_skip = 30):
        """
        Visualize the frontal GoPro features.

        Parameters
        ----------
        data : dict
            The dictionary that contains the necessary data.
        subject : str
            The subject ID.
        emotion : str
            The emotion.
        frame_skip : int
            The frame skip.
        """

        start_frame, end_frame = data[subject]['GoPro']['Frontal']['label'][emotion]
        video_path = data[subject]['GoPro']['Frontal']['data_path']

        frame_counter = 0

        cap = cv2.VideoCapture(video_path)

        cap.set(cv2.CAP_PROP_POS_FRAMES, start_frame)

        right_wrist_positions = []
        left_wrist_positions = []

        while cap.isOpened():
            ret, frame = cap.read()

            if not ret :
                break

            current_frame = int(cap.get(cv2.CAP_PROP_POS_FRAMES))

            if current_frame > end_frame:
                break

            if frame_counter % frame_skip != 0:
                frame_counter += 1
                continue

            frame_counter += 1

            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            results = self.pose.process(frame_rgb)

            if results.pose_landmarks:
                landmark_list = results.pose_landmarks.landmark

                left_shoulder = [landmark_list[self.mp_pose.PoseLandmark.LEFT_SHOULDER.value].x, landmark_list[self.mp_pose.PoseLandmark.LEFT_SHOULDER.value].y]
                right_shoulder = [landmark_list[self.mp_pose.PoseLandmark.RIGHT_SHOULDER.value].x, landmark_list[self.mp_pose.PoseLandmark.RIGHT_SHOULDER.value].y]
                left_elbow = [landmark_list[self.mp_pose.PoseLandmark.LEFT_ELBOW.value].x, landmark_list[self.mp_pose.PoseLandmark.LEFT_ELBOW.value].y]
                right_elbow = [landmark_list[self.mp_pose.PoseLandmark.RIGHT_ELBOW.value].x, landmark_list[self.mp_pose.PoseLandmark.RIGHT_ELBOW.value].y]
                left_wrist = [landmark_list[self.mp_pose.PoseLandmark.LEFT_WRIST.value].x, landmark_list[self.mp_pose.PoseLandmark.LEFT_WRIST.value].y]
                right_wrist = [landmark_list[self.mp_pose.PoseLandmark.RIGHT_WRIST.value].x, landmark_list[self.mp_pose.PoseLandmark.RIGHT_WRIST.value].y]


                shoulder_dist = np.linalg.norm(
                    np.array(left_shoulder) - np.array(right_shoulder)
                )
                shoulder_dist = np.round(shoulder_dist, 2)

                right_elbow_angle = int(self._compute_arm_angle(right_shoulder, right_elbow, right_wrist))

                left_elbow_angle = int(self._compute_arm_angle(left_shoulder, left_elbow, left_wrist))

                right_wrist_position = np.round(np.array(right_wrist),2)
                right_wrist_positions.append(right_wrist_position)

                left_wrist_position = np.round(np.array(left_wrist),2)
                left_wrist_positions.append(left_wrist_position)

                if len(right_wrist_positions) > 1:
                    right_wrist_velocity = np.round(np.linalg.norm(right_wrist_positions[-1] - right_wrist_positions[-2]),3)

                if len(left_wrist_positions) > 1:
                    left_wrist_velocity = np.round(np.linalg.norm(left_wrist_positions[-1] - left_wrist_positions[-2]),3)

                symmetry = np.round(np.abs(left_wrist[0] - right_wrist[0]),3)

                h, w, c = frame.shape

                ls = (int(left_shoulder[0] * w), int(left_shoulder[1] * h))
                rs = (int(right_shoulder[0] * w), int(right_shoulder[1] * h))
                le = (int(left_elbow[0] * w), int(left_elbow[1] * h))
                re = (int(right_elbow[0] * w), int(right_elbow[1] * h))
                lw = (int(left_wrist[0] * w), int(left_wrist[1] * h))
                rw = (int(right_wrist[0] * w), int(right_wrist[1] * h))

                # Draw shoulder line
                cv2.line(frame, ls, rs, (0, 255, 0), 2)

                # Draw elbow angle
                cv2.line(frame, rs, re, (255, 0, 0), 2)
                cv2.line(frame, re, rw, (255, 0, 0), 2)
                cv2.line(frame, ls, le, (255, 0, 0), 2)
                cv2.line(frame, le, lw, (255, 0, 0), 2)

                # Draw wrist position
                cv2.circle(frame, rw, 5, (0, 0, 255), -1)
                cv2.circle(frame, lw, 5, (0, 0, 255), -1)

                cv2.putText(frame, "Right Elbow Angle: {}".format(right_elbow_angle), (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                cv2.putText(frame, "Left Elbow Angle: {}".format(left_elbow_angle), (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                cv2.putText(frame, "Shoulder Distance: {}".format(shoulder_dist), (10, 90), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                cv2.putText(frame, "Right Wrist Position: {}".format(right_wrist_position), (10, 120), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                cv2.putText(frame, "Left Wrist Position: {}".format(left_wrist_position), (10, 150), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                if len(right_wrist_positions) > 1:
                    cv2.putText(frame, "Right Wrist Velocity: {}".format(right_wrist_velocity), (10, 180), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                if len(left_wrist_positions) > 1:
                    cv2.putText(frame, "Left Wrist Velocity: {}".format(left_wrist_velocity), (10, 210), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                cv2.putText(frame, "Symmetry: {}".format(symmetry), (10, 240), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

                frame_height, frame_width = frame.shape[:2]
                resized_width = int(frame_width / 2)
                resized_height = int(frame_height / 2)
                display_frame = cv2.resize(frame, (resized_width, resized_height))

            cv2.imshow('Feature Visualization', display_frame)
            if cv2.waitKey(1) & 0xFF == 27:  # Press 'ESC' to exit
                break
                
        cap.release()
        cv2.destroyAllWindows()

    def visualize_lateral_gopro_videos(self, data, subject, emotion, frame_skip =30):
        """
        Visualize the lateral GoPro features.

        Parameters
        ----------
        data : dict
            The dictionary that contains the necessary data.
        subject : str
            The subject ID.
        emotion : str
            The emotion.
        frame_skip : int
            The frame skip.
        """

        start_frame, end_frame = data[subject]['GoPro']['Lateral']['label'][emotion]
        video_path = data[subject]['GoPro']['Lateral']['data_path']
        cap = cv2.VideoCapture(video_path)
        cap.set(cv2.CAP_PROP_POS_FRAMES, start_frame)

        frame_counter = 0
        first_trunk_angle = None
        first_head_angle = None

        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break

            current_frame = int(cap.get(cv2.CAP_PROP_POS_FRAMES))
            if current_frame > end_frame:
                break

            if frame_counter % frame_skip != 0:
                frame_counter += 1
                continue

            frame_counter += 1

            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            results = self.pose.process(frame_rgb)

            if results.pose_landmarks:
                landmarks = results.pose_landmarks.landmark

                # Extract landmarks
                left_shoulder = [landmarks[self.mp_pose.PoseLandmark.LEFT_SHOULDER.value].x, landmarks[self.mp_pose.PoseLandmark.LEFT_SHOULDER.value].y]
                right_shoulder = [landmarks[self.mp_pose.PoseLandmark.RIGHT_SHOULDER.value].x, landmarks[self.mp_pose.PoseLandmark.RIGHT_SHOULDER.value].y]
                left_hip = [landmarks[self.mp_pose.PoseLandmark.LEFT_HIP.value].x, landmarks[self.mp_pose.PoseLandmark.LEFT_HIP.value].y]
                right_hip = [landmarks[self.mp_pose.PoseLandmark.RIGHT_HIP.value].x, landmarks[self.mp_pose.PoseLandmark.RIGHT_HIP.value].y]

                ear = [landmarks[self.mp_pose.PoseLandmark.LEFT_EAR.value].x, landmarks[self.mp_pose.PoseLandmark.LEFT_EAR.value].y]

                shoulder_mid = np.mean([left_shoulder, right_shoulder], axis=0)
                hip_mid = np.mean([left_hip, right_hip], axis=0)

                if first_trunk_angle is None and first_head_angle is None:
                    first_trunk_angle = self._compute_trunk_inclination(shoulder_mid, hip_mid)
                    first_head_angle = self._compute_head_inclination(ear, shoulder_mid)
                    
                trunk_angle =  first_trunk_angle - self._compute_trunk_inclination(shoulder_mid, hip_mid)
                head_angle = first_head_angle - self._compute_head_inclination(ear, shoulder_mid)

                h, w, c = frame.shape
                cv2.line(frame, (int(shoulder_mid[0] * w), int(shoulder_mid[1] * h)), (int(hip_mid[0] * w), int(hip_mid[1] * h)), (0, 0, 255), 3)
                cv2.putText(frame, f"Trunk Angle: {int(trunk_angle)}", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2, cv2.LINE_AA)
                cv2.line(frame, (int(ear[0] * w), int(ear[1] * h)), (int(shoulder_mid[0] * w), int(shoulder_mid[1] * h)), (255, 0, 0), 3)
                cv2.putText(frame, f"Neck Angle: {int(head_angle)}", (50, 100), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2, cv2.LINE_AA)

            frame_height, frame_width = frame.shape[:2]
            resized_width = int(frame_width / 2)
            resized_height = int(frame_height / 2)
            display_frame = cv2.resize(frame, (resized_width, resized_height))

            cv2.imshow('MediaPipe Pose', display_frame)
            if cv2.waitKey(1) & 0xFF == 27:  # Press 'Esc' to exit
                break

        cap.release()
        cv2.destroyAllWindows()

    def select_features(self, X, y, model, feature_selector, features_names, equalize = False, nb_features=10):
        """
        Selects the most important features using the specified model

        Parameters
        ----------
        X : pandas.DataFrame
            The dataframe containing the features
        y : pandas.Series
            The labels
        model : str
            The model to use for feature selection. Either RF, SVM
        feature selector : str
            The selector to use. Either 'RF' for Random Forest or 'RFE' for Recursive Feature Elimination  using random forest or 'SFS' for forward feature Selection using random forest
        feature_names : dict
            Features names per measurement
        equalize : bool
            Whether to equalize the number of extracted features across measurements or not
        nb_features : int
            The number of features to select

        Returns
        -------
        X_selected : pandas.DataFrame
            The dataframe containing the selected features
        """

        if model == 'RF':
            model = RandomForestClassifier(n_estimators=100, random_state=42)
        elif model == 'SVM':
            model = SVC(kernel='rbf', random_state=42)
        else:
            print("Random Forest model selected by default")
            model = RandomForestClassifier(n_estimators=100, random_state=42)

        if feature_selector == 'RF':
            selector = model
        elif feature_selector == 'RFE':
            selector = RFE(estimator=model, n_features_to_select=nb_features)
        elif feature_selector == 'SFS':
            selector = SequentialFeatureSelector(model, k_features=nb_features, forward=True, floating=False, scoring='accuracy', cv=5)
        else:
            print("Recursive feature Eliminator feature selector selected by default")
            selector = RFE(estimator=model, n_features_to_select=nb_features)

        if equalize:
            X_selected = pd.DataFrame()
            for rec in features_names:
                if rec != 'SRE':
                    if nb_features > len(features_names[rec]):
                        nb_features = len(features_names[rec])
                        print("The number of features to select is greater than the number of features available. The number of features to select is set to {}".format(nb_features))
                    selector.fit(X[features_names[rec]], y)
                    if feature_selector == 'RF':
                        importances = selector.feature_importances_
                        indices = np.argsort(importances)[::-1]
                        top_indices = indices[:nb_features]
                        selected_features = X[features_names[rec]].columns[top_indices]
                    if feature_selector == 'RFE':
                        selected_features = X[features_names[rec]].columns[selector.support_]
                    elif feature_selector == 'SFS':
                        selected_features = X[features_names[rec]].columns[list(selector.k_feature_idx_)]
                    X_selected = pd.concat([X_selected, X[selected_features]], axis=1)
                else:
                    X_selected = pd.concat([X_selected, X[features_names[rec]]], axis=1)
        else:
            if nb_features > len(X.columns):
                nb_features = len(X.columns)
                print("The number of features to select is greater than the number of features available. The number of features to select is set to {}".format(nb_features))
            selector.fit(X, y)
            if feature_selector == 'RF':
                importances = selector.feature_importances_
                indices = np.argsort(importances)[::-1]
                top_indices = indices[:nb_features]
                selected_features = X.columns[top_indices]
            elif feature_selector == 'RFE':
                selected_features = X.columns[selector.support_]
            elif feature_selector == 'SFS':
                selected_features = X.columns[list(selector.k_feature_idx_)]

            X_selected = X[selected_features]

        return X_selected
    
    def get_projection_sets(self, features, proj_obj, features_names, feature_selection=False, nb_selected_features=30):
        """
        Get the sets for the comparison
        
        Parameters
        ----------
        features : pandas.DataFrame
            The dataframe containing the features
        proj_obj : str
            The projection objective to consider
        features_names : dict
            The dictionary containing the features names
        feature_selection : bool
            Whether to perform feature selection or not
        nb_selected_features : int
            The number of features to select

        Returns
        -------
        X : pandas.DataFrame
            The features to use
        y : pandas.Series
            The labels to use
        """
        if proj_obj == 'Neutral vs Non-Neutral':
            X = features.drop('label', axis=1)
            y = features['label']
            y = y.replace({'Frustration': 'Non-Neutral', 'Pride': 'Non-Neutral', 'Joy': 'Non-Neutral', 'Shame': 'Non-Neutral'})
        elif proj_obj == 'Positive vs Negative':
            features = features[features['label'] != 'Neutral']
            X = features.drop('label', axis=1)
            y = features['label']
            y = y.replace({'Frustration': 'Negative', 'Shame': 'Negative', 'Pride': 'Positive', 'Joy': 'Positive'})
        elif proj_obj == 'Shame vs Positive':
            features = features[features['label'] != 'Neutral']
            features = features[features['label'] != 'Frustration']
            X = features.drop('label', axis=1)
            y = features['label']
            y = y.replace({'Pride': 'Positive', 'Joy': 'Positive'})
        elif proj_obj == 'Shame vs Others':
            features = features[features['label'] != 'Neutral']
            X = features.drop('label', axis=1)
            y = features['label']
            y = y.replace({'Frustration': 'Other', 'Pride': 'Other', 'Joy': 'Other'})
        elif proj_obj == 'All Emotions':
            features = features[features['label'] != 'Neutral']
            X = features.drop('label', axis=1)
            y = features['label']

        if feature_selection:
            X = self.select_features(X, y, model = 'RF', feature_selector = 'RFE',features_names = features_names, equalize = False, nb_features=nb_selected_features)
            selected_features = X.columns.tolist()
        else:
            selected_features = None

        return X, y, selected_features
    
    def plot_pca_projection(self, X, y, nb_components=2, dim ='2D', save=False, dir = None, verbose = False):
        """
        Plot the PCA projection of the data

        Parameters
        ----------
        X : pandas.DataFrame
            The dataframe containing the features
        y : pandas.Series
            The labels
        nb_components : int
            The number of components to use for the PCA
        dim : str
            The dimension of the plot. Either '2D' or '3D'
        save : bool
            Whether to save the plot or not.
        dir : str
            The directory to save the plot
        verbose : bool
            Whether to print the progress bar or not.
        """

        pca = PCA(n_components=nb_components)
        pca_result = pca.fit_transform(X)

        pca_features = dict()

        pca_features['pca-one'] = pca_result[:,0]
        pca_features['pca-two'] = pca_result[:,1]
        if nb_components == 3:
            pca_features['pca-three'] = pca_result[:,2]

        if verbose:
            print('Explained variation per principal component: {}'.format(pca.explained_variance_ratio_))

        custom_colors = {
                'Shame': 'magenta',
                'Frustration': 'orange',
                'Joy': 'green',
                'Pride': 'cyan',
                'Neutral': 'black',
                'Non-Neutral': 'grey',
                'Positive': 'blue',
                'Negative': 'red',
                'Other': 'yellow'
            }


        if dim == '2D':
            plt.figure(figsize=(8,5))
            sns.scatterplot(
                hue=y,
                x="pca-one", y="pca-two",
                palette=custom_colors,
                data=pca_features,
                legend="full",
                alpha=1
            )
            plt.title('PCA Results (First 2 Components)')
            if save:
                os.makedirs(dir, exist_ok=True)
                plt.savefig(os.path.join(dir, 'PCA.png'))
                plt.close()
            else:
                plt.show()
        elif dim =='3D':
            def _plot_3d(elev=30, azim=30):
                fig = plt.figure(figsize=(8, 5))
                ax = fig.add_subplot(111, projection='3d')
                
                legend_handles = [plt.Line2D([0], [0], marker='o', color='w', label=label, 
                                            markersize=10, markerfacecolor=color) for label, color in custom_colors.items()]
                
                scatter = ax.scatter(
                    xs=pca_features["pca-one"],
                    ys=pca_features["pca-two"],
                    zs=pca_features["pca-three"],
                    c=[custom_colors[i] for i in y],
                )
                
                ax.view_init(elev=elev, azim=azim)
                
                ax.legend(handles=legend_handles, title="Classes")
                
                ax.set_xlabel('PCA 1')
                ax.set_ylabel('PCA 2')
                ax.set_zlabel('PCA 3')
                plt.title('3D PCA Results')
                if save:
                    os.makedirs(dir, exist_ok=True)
                    plt.savefig(os.path.join(dir, 'PCA.png'))
                    plt.close()
                else:
                    plt.show()

                    elev_slider = widgets.IntSlider(min=0, max=90, step=1, value=30, description='Elev')
                    azim_slider = widgets.IntSlider(min=0, max=360, step=1, value=30, description='Azim')

                    widgets.interactive(_plot_3d, elev=elev_slider, azim=azim_slider)
    
    def plot_tsne_projection(self, X, y, nb_component, verbose, perplexity=40, n_iter=300, save=False, dir = None):
        """
        Plot the t-SNE projection of the data

        Parameters
        ----------
        X : pandas.DataFrame
            The dataframe containing the features
        y : pandas.Series
            The labels
        nb_component : int
            The number of components to use for the t-SNE
        verbose : bool
            Whether to print the progress bar or not.
        perplexity : int
            The perplexity parameter for the t-SNE
        n_iter : int
            The number of iterations for the t-SNE
        save : bool
            Whether to save the plot or not.
        dir : str
            The directory to save the plot
        """

        tsne = TSNE(n_components=nb_component, verbose=verbose, perplexity=perplexity, n_iter=n_iter)
        tsne_results = tsne.fit_transform(X)

        tne_features = dict()

        tne_features['tsne-one'] = tsne_results[:,0]
        tne_features['tsne-two'] = tsne_results[:,1]

        plt.figure(figsize=(8,5))
        sns.scatterplot(
            x="tsne-one", y="tsne-two",
            hue=y,
            palette=sns.color_palette("Set1", n_colors=5),
            data=tne_features,
            legend="full",
            alpha=1
        )
        plt.title('t-SNE Results')

        if save:
            if dir is not None:
                os.makedirs(dir, exist_ok=True)
                plt.savefig(os.path.join(dir, 'TSNE.png'))
                plt.close()
        else:
            plt.show()

    
    def plot_sammon_projection(self, X, y, nb_component=2, max_iter=100, eps=1e-3, n_init=10, random_state=42, save=False, dir = None):
        """
        Plot the Sammon projection of the data

        Parameters
        ----------
        X : pandas.DataFrame
            The dataframe containing the features
        y : pandas.Series
            The labels
        nb_component : int
            The number of components to use for the Sammon
        max_iter : int
            The maximum number of iterations for the Sammon
        eps : float
            The epsilon parameter for the Sammon
        n_init : int
            The number of initializations for the Sammon
        random_state : int
            The random state for the Sammon
        save : bool
            Whether to save the plot or not.
        dir : str
            The directory to save the plot
        """

        mds = MDS(n_components=nb_component, max_iter=max_iter, eps=eps, n_init= n_init, random_state=random_state)
        mds_results = mds.fit_transform(X)

        plt.figure(figsize=(8, 5))
        for label in y.unique():
            mask = y == label
            plt.scatter(mds_results[mask, 0], mds_results[mask, 1], label=label)

        plt.title('MDS Mapping of Features')
        plt.xlabel('MDS Dimension 1')
        plt.ylabel('MDS Dimension 2')
        plt.legend()
        if save:
            os.makedirs(dir, exist_ok=True)
            plt.savefig(os.path.join(dir, 'Sammon.png'))
            plt.close()
        else:
            plt.show()

    def plot_subjects_pca_projections(self, features_grouping):
        """
        Plot the PCA projection colored by subject
        
        Parameters
        ----------
        features_grouping : dict
            The dictionary containing the features grouped by emotion.
        """
        emotion = list(features_grouping.keys())
        df = pd.concat(features_grouping.values(), axis=0)
            
        df.sort_index(inplace=True)

        X = df.drop('subject', axis=1)
        y = df['subject']

        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)

        pca = PCA(n_components=2)
        X_pca = pca.fit_transform(X_scaled)

        pca_df = pd.DataFrame(data = X_pca, columns = ['PC1', 'PC2'])
        pca_df['Subject'] = y

        custom_palette = [
            "#E6194B", "#3CB44B", "#FFE119", "#0082C8", "#F58231",
            "#911EB4", "#46F0F0", "#F032E6", "#D2F53C", "#FABEBE",
            "#008080", "#E6BEFF", "#AA6E28", "#FFFAC8", "#800000",
            "#AAFFC3", "#808000", "#FFD8B1", "#000080", "#808080"
        ]

        plt.figure(figsize=(10, 8))
        sns.scatterplot(x='PC1', y='PC2', hue='Subject', data=pca_df, palette=custom_palette, legend='full', alpha=0.7)

        for subject, group_df in pca_df.groupby('Subject'):
            self.utilities.draw_ellipse(group_df[['PC1', 'PC2']].mean(), np.cov(group_df[['PC1', 'PC2']].T), alpha=0.2, color='gray')

        plt.title('PCA Projection by Subject')
        plt.xlabel('Principal Component 1')
        plt.ylabel('Principal Component 2')
        plt.legend(title='Subject', bbox_to_anchor=(1.05, 1), loc='upper left')
        plt.show()

    def plot_subject_coding_projections(self, features_grouping, subject_coding):
        """
        Plot the PCA projection colored by subject coding

        Parameters
        ----------
        features_grouping : dict
            The dictionary containing the features grouped by emotion.
        subject_coding : dict
            The dictionary containing the subject coding.
        """
        emotions = list(features_grouping.keys())
        df = pd.concat([features_grouping[emotions[0]], features_grouping[emotions[1]], features_grouping[emotions[2]], features_grouping[emotions[3]]], axis=0)
        df.sort_index(inplace=True)

        scaler = StandardScaler()
        scaled_features = scaler.fit_transform(df.iloc[:, 1:])

        df['Category'] = df['subject'].apply(lambda x: next((k for k, v in subject_coding.items() if x in v), 'Unknown'))

        pca = PCA(n_components=2)
        pca_results = pca.fit_transform(scaled_features)

        categories = list(subject_coding.keys())
        sns.set(style='whitegrid') 

        feature_names = df.columns[1:-1]
        top_features_per_component = pd.DataFrame(pca.components_.T, index=feature_names, columns=['PC1', 'PC2'])
        top_features_pc1 = top_features_per_component['PC1'].abs().sort_values(ascending=False).head(15)
        top_features_pc2 = top_features_per_component['PC2'].abs().sort_values(ascending=False).head(15)

        print("Top contributing features for PC1:", top_features_pc1.index.tolist())
        print("Top contributing features for PC2:", top_features_pc2.index.tolist())

        plt.figure(figsize=(10, 7))

        for category, color in zip(categories, sns.color_palette("hsv", len(categories))):
            indices = df['Category'] == category
            plt.scatter(pca_results[indices, 0], pca_results[indices, 1], label=category, s=50, alpha=0.7)

        plt.title('PCA Projection Colored by Consistency across emotions', fontsize=16)
        plt.xlabel('PCA Component 1', fontsize=14)
        plt.ylabel('PCA Component 2', fontsize=14)
        plt.legend(title='Subject Category', title_fontsize='13', fontsize='12')

        plt.show()

    def pca_feature_importance(self, data, pca):
        """
        Perform PCA on the dataset and return the features that best describe each principal component, ordered by their contribution.
        
        Parameters
        ----------
        data : pd.DataFrame
            The dataset.
        pca : PCA
            The PCA model.

        Returns
        -------
        dict
            The dictionary containing the feature contributions for each principal component.
        """

        components = pca.components_
        feature_names = data.columns
        pca_feature_contributions = {}

        for i, component in enumerate(components):
            feature_contributions = sorted(zip(feature_names, component), key=lambda x: np.abs(x[1]), reverse=True)
            pca_feature_contributions[i + 1] = feature_contributions
        
        return pca_feature_contributions
    
    def get_all_projections(self, features, features_names, groups=None, nb_selected_features=30, save=False):
        """
        Get all the projections

        Parameters
        ----------
        features : pandas.DataFrame
            The dataframe containing the features
        features_names : dict
            The dictionary containing the features names
        groups : dict
            The dictionary containing the groups of subjects
        nb_selected_features : int
            The number of features to select
        save : bool
            Whether to save the plots or not.

        Returns
        -------
        selected_features : dict
            The dictionary containing the selected features
        """
        desired_projections = ['Neutral vs Non-Neutral', 'Positive vs Negative', 'Shame vs Others', 'All Emotions']

        selected_features = dict()

        pbar = tqdm(total=len(groups.keys()) * len(desired_projections))

        for group in groups.keys():

            features_to_use = self.utilities.select_features_from_subjects(features, groups[group], include_neutral=True)

            selected_features[group] = dict()

            for projection in desired_projections: 
                pbar.set_description(f'Computing projection: {projection}')

                X, y, selected_features_names = self.get_projection_sets(features_to_use, projection, features_names, feature_selection=True, nb_selected_features=nb_selected_features)

                selected_features[group][projection] = selected_features_names

                dir = os.path.join('Analysis', 'Projections', group, projection)

                self.plot_pca_projection(X, y, nb_components=2, dim='2D', save=save, dir=dir)

                if len(groups[group]) < 10:
                    self.plot_tsne_projection(X, y, nb_component=2, verbose=False, perplexity=len(groups[group]), n_iter=500, save=save, dir=dir)
                else:
                    self.plot_tsne_projection(X, y, nb_component=2, verbose=False, perplexity=10, n_iter=500, save=save, dir=dir)
                self.plot_sammon_projection(X, y, nb_component=2, max_iter=3000, eps=1e-3, n_init=500, random_state=42, save=save, dir=dir)

                pbar.update(1)
        pbar.close()

        return selected_features

    def get_subjects_clusters(self, features_grouping, analysis, desired_emotion = None, nb_clusters=None, method = 'Elbow', projections='PCA', windows = False, plot = False, save = False):
        """
        Get the clusters of the subjects based on the features

        Parameters
        ----------
        features_grouping : dict
            The dictionary containing the features per emotion
        analysis : str
            The type of analysis to perform. Either 'single' or 'pair' or 'all
        desired_emotion : str or list
            The desired emotion to analyze. Str if one emotion and list if two emotions
        nb_clusters : int
            The number of clusters to use
        method : str
            The method to use for clustering. Either 'Elbow' or 'Silhouette'
        projections : str
            The type of projection to use. Either 'PCA' or 'Sammon'
        windows : bool
            Whether to features are divided into windows or not
        plot : bool
            Whether to plot the elbow method or not.
        save : bool
            Whether to save the plots or not.

        Returns
        -------
        clusters_dict : dict
            The dictionary containing the clusters of the subjects
        """
        clusters_dict = dict()
        feature_importance = dict()

        if analysis == 'single':
            if desired_emotion is None:
                emotions = features_grouping.keys()
            else:
                if isinstance(desired_emotion, str):
                    emotions = [desired_emotion]
                else:
                    print("Desired emotion should be a string with a single analysis")
                    return
        elif analysis == 'pair':
            if desired_emotion is None:
                emotions = list(combinations(list(features_grouping.keys()), 2))
            else:
                if isinstance(desired_emotion, list):
                    emotions = [desired_emotion]
                else:
                    print("Desired emotion should be a list with a pair of emotions")
                    return
        elif analysis == 'all':
            emotions = [list(features_grouping.keys())]
        else:
            print("Analysis type not recognized. Please select either 'single' or 'pair' or 'all'")
            return

        for emotion in emotions:
            if analysis == 'single':
                df = features_grouping[emotion].copy()
                em = emotion
            elif analysis == 'pair':
                df = pd.concat([features_grouping[emotion[0]], features_grouping[emotion[1]]], axis=0)
                df.sort_index(inplace=True)
                em = f'{emotion[0]}-{emotion[1]} Pair'
            elif analysis == 'all':
                df = pd.concat([features_grouping[emotion[0]], features_grouping[emotion[1]], features_grouping[emotion[2]], features_grouping[emotion[3]]], axis=0)
                df.sort_index(inplace=True)
                em = 'All Emotions'

            clusters_dict[em] = dict()
            feature_importance[em] = dict()

            if windows:
                for feature in df.columns:
                    if feature != 'subject':
                        for row in range(df.shape[0]):
                            df[feature].iloc[row] = np.mean(df[feature].iloc[row])

            scaler = StandardScaler()
            scaled_features = scaler.fit_transform(df.iloc[:, 1:])

            if method == 'Elbow':
                inertia = []
                for k in range(1, 6):
                    kmeans = KMeans(n_clusters=k, random_state=42)
                    kmeans.fit(scaled_features)
                    inertia.append(kmeans.inertia_)
            elif method == 'Silhouette':
                silhouette_scores = []
                cluster_range = range(2, 6)
                for k in cluster_range:
                    kmeans = KMeans(n_clusters=k, random_state=42)
                    clusters = kmeans.fit_predict(scaled_features)
                    silhouette_avg = silhouette_score(scaled_features, clusters)
                    silhouette_scores.append(silhouette_avg)

            if nb_clusters is not None:
                optimal_clusters = nb_clusters
            else:
                optimal_clusters =  list(cluster_range)[silhouette_scores.index(max(silhouette_scores))]

            kmeans = KMeans(n_clusters=optimal_clusters, init='k-means++', n_init=10, max_iter=1000, tol=1e-5, random_state=42)
            clusters = kmeans.fit_predict(scaled_features)

            feature_names = df.columns[1:]
            
            if projections == 'PCA':
                pca = PCA(n_components=2)
                principal_components = pca.fit_transform(scaled_features)
                top_features_per_component = pd.DataFrame(pca.components_.T, index=feature_names, columns=['PC1', 'PC2'])
                top_features_pc1 = top_features_per_component['PC1'].abs().sort_values(ascending=False).head(15)
                top_features_pc2 = top_features_per_component['PC2'].abs().sort_values(ascending=False).head(15)
                feature_importance[em]['PC1'] = top_features_pc1.index.tolist()
                feature_importance[em]['PC2'] = top_features_pc2.index.tolist()
            elif projections == 'Sammon':
                mds = MDS(n_components=2, max_iter=5000, eps=1e-3, n_init=1000, random_state=42)
                principal_components = mds.fit_transform(scaled_features)

            df['cluster'] = clusters

            fig, axs = plt.subplots(1, 2, figsize=(20, 6))  

            if method == 'Elbow':
                axs[0].plot(range(1, 6), inertia, color='crimson', linestyle='-', linewidth=2)
                axs[0].set_title('Elbow Method', fontsize=16)
                axs[0].set_xlabel('Number of clusters', fontsize=14)
                axs[0].set_ylabel('Inertia', fontsize=14)
                axs[0].tick_params(labelsize=12)
                axs[0].grid(True, which='both', linestyle='--', linewidth=0.5)
                axs[0].axvline(optimal_clusters, color='black', linestyle='--', linewidth=1)
            elif method == 'Silhouette':
                axs[0].plot(range(2, 6), silhouette_scores, color='crimson', linestyle='-', linewidth=2)
                axs[0].set_title('Silhouette Method', fontsize=16)
                axs[0].set_xlabel('Number of clusters', fontsize=14)
                axs[0].set_ylabel('Silhouette Score', fontsize=14)
                axs[0].tick_params(labelsize=12)
                axs[0].grid(True, which='both', linestyle='--', linewidth=0.5)
                axs[0].axvline(optimal_clusters, color='black', linestyle='--', linewidth=1)
                axs[0].xaxis.set_major_locator(ticker.MaxNLocator(integer=True))

            scatter = axs[1].scatter(principal_components[:, 0], principal_components[:, 1], c=clusters, cmap='coolwarm')
            axs[1].set_title('Cluster Visualization', fontsize=16)
            if projections == 'PCA':
                axs[1].set_xlabel('Principal Component 1', fontsize=14)
                axs[1].set_ylabel('Principal Component 2', fontsize=14)
            elif projections == 'Sammon':
                axs[1].set_xlabel('Sammon Component 1', fontsize=14)
                axs[1].set_ylabel('Sammon Component 2', fontsize=14)
            axs[1].tick_params(labelsize=12)
            axs[1].grid(True, which='both', linestyle='--', linewidth=0.5)

            plt.suptitle(f'{em} Analysis', fontsize=20)
            plt.tight_layout(rect=[0, 0.03, 1, 0.95])

            if plot:
                plt.show()

            for cluster_nb in range(optimal_clusters):
                if analysis == 'single':
                    clusters_dict[em][f'cluster{cluster_nb}'] = list(df['subject'][df['cluster']==cluster_nb])
                elif analysis =='pair':
                    subject_counts = df['subject'][df['cluster']==cluster_nb].value_counts()
                    subjects_to_keep = subject_counts[subject_counts == 2].index
                    removed_subjects = subject_counts[subject_counts != 2].index
                    print(f"Removed subjects for {em} and cluster {cluster_nb}: {list(removed_subjects)}")
                    clusters_dict[em][f'cluster{cluster_nb}'] = list(np.sort(list(set(df['subject'][df['cluster']==cluster_nb]).intersection(subjects_to_keep))))
                elif analysis == 'all':
                    subject_counts = df['subject'][df['cluster']==cluster_nb].value_counts()
                    subjects_to_keep = subject_counts[subject_counts > 2].index
                    removed_subjects = subject_counts[subject_counts < 3].index
                    print(f"Removed subjects for {em} and cluster {cluster_nb}: {list(removed_subjects)}")
                    clusters_dict[em][f'cluster{cluster_nb}'] = list(np.sort(list(set(df['subject'][df['cluster']==cluster_nb]).intersection(subjects_to_keep))))

            if save:
                os.makedirs(f'Analysis/Cluster Analysis/{projections}', exist_ok=True)
                plt.savefig(f'Analysis/Cluster Analysis/{projections}/{em}.png')
                plt.close()

        return clusters_dict, feature_importance

    def map_subjects_to_code(self, clusters_dict, subject_coding):
        """
        Map the subjects to their respective codes

        Parameters
        ----------
        clusters_dict : dict
            The dictionary containing the clusters of the subjects
        subject_coding : dict
            The dictionary containing the subject coding
        
        Returns
        -------
        dict
            The dictionary containing the codes of the subjects
        """
        coding_dict = dict()
        for emotion in clusters_dict.keys():
            coding_dict[emotion] = dict()
            for cluster in clusters_dict[emotion].keys():
                coding_dict[emotion][cluster] = list()
                for subject in clusters_dict[emotion][cluster]:
                    for code in subject_coding.keys():
                        if subject in subject_coding[code]:
                            coding_dict[emotion][cluster].append(code)
                            break
                
        return coding_dict
    def plot_intersections(self, clusters_dict, nb_emotions =2, save = False):
        """
        Plot the intersections of the clusters

        Parameters
        ----------
        clusters_dict : dict
            The dictionary containing the clusters of the subjects
        nb_emotions : int
            The number of emotions to consider for the intersections
        save : bool
            Whether to save the plots or not.
        """
        emotion_pairs = list(combinations(list(clusters_dict.keys()),nb_emotions))
        
        plt.figure(figsize=(12, 6 * len(emotion_pairs)))

        for pair in emotion_pairs:
            plt.figure(figsize=(12, 6))
    
            plt.subplot(1, 2, 1) 
            if nb_emotions == 2:
                venn = venn2([set(clusters_dict[i]['cluster0']) for i in pair], set_labels=pair, set_colors=('grey', 'red'))
            elif nb_emotions == 3 and len(pair) == 3:
                venn = venn3([set(clusters_dict[i]['cluster0']) for i in pair], set_labels=pair, set_colors=('black', 'grey', 'red'))
            else:
                print("Venn diagram for more than 3 sets is not supported.")
                continue  
            for patch in venn.patches:
                if patch is not None:
                    patch.set_alpha(0.8)
            plt.title(f'Cluster 0')
 
            plt.subplot(1, 2, 2)
            if nb_emotions == 2:
                venn = venn2([set(clusters_dict[i]['cluster1']) for i in pair], set_labels=pair, set_colors=('grey', 'red'))
            elif nb_emotions == 3 and len(pair) == 3:
                venn = venn3([set(clusters_dict[i]['cluster1']) for i in pair], set_labels=pair, set_colors=('black', 'grey', 'red'))
            for patch in venn.patches:
                if patch is not None:
                    patch.set_alpha(0.8)
            plt.title(f'Cluster 1')
            
            if nb_emotions == 2:
                plt.suptitle(f'Intersections of the clusters for {pair[0]} and {pair[1]}', fontsize=15)
            elif nb_emotions == 3:
                plt.suptitle(f'Intersections of the clusters for {pair[0]}, {pair[1]} and {pair[2]}', fontsize=15)

            plt.tight_layout(rect=[0, 0.03, 1, 0.95])

            if save:
                os.makedirs(f'Analysis/Cluster Analysis/Intersections/{nb_emotions} Emotions', exist_ok=True)
                if nb_emotions == 2:
                    plt.savefig(f'Analysis/Cluster Analysis/Intersections/{nb_emotions} Emotions/{pair[0]}&{pair[1]}.png')
                elif nb_emotions == 3:
                    plt.savefig(f'Analysis/Cluster Analysis/Intersections/{nb_emotions} Emotions/{pair[0]}&{pair[1]}&{pair[2]}.png')
                plt.close()
            else:
                plt.show()

    def get_subject_consistency(self, clusters_dict, emotions = None, verbose=True, min_nb_emotions=3):
        """
        Get the consistent subjects across the clusters for at least 3 emotions

        Parameters
        ----------
        clusters_dict : dict
            The dictionary containing the clusters of the subjects
        verbose : bool
            Whether to print the results or not.
        min_nb_emotions : int
            The minimum number of emotions to consider for the consistency
        
        Returns
        ------- 
        dict
            The dictionary containing the consistent subjects across the clusters
        """

        clusters = dict()
        count = dict()

        if emotions is None:
            emotions = list(clusters_dict.keys())

        for emotion in emotions:
            if emotion not in clusters_dict.keys():
                print(f"Emotion {emotion} not found in the clusters dictionary.")
                continue
            for cluster in clusters_dict[emotion].keys():
                for subject1 in clusters_dict[emotion][cluster]:
                    for subject2 in clusters_dict[emotion][cluster]:
                        if subject1 != subject2:
                            if tuple([subject1, subject2]) not in count.keys():
                                if tuple([subject2, subject1]) not in count.keys():
                                    count[tuple([subject1, subject2])] = 1
                            else:
                                count[tuple([subject1, subject2])] += 1
        i = 0
        for pair in count:
            if count[pair] >= min_nb_emotions:
                cluster_found = False
                for k in clusters.keys():
                    if pair[0] in clusters[k] and pair[1] not in clusters[k]:
                        clusters[k].append(pair[1])
                        cluster_found = True
                        break
                    elif pair[1] in clusters[k] and pair[0] not in clusters[k]:
                        clusters[k].append(pair[0])
                        cluster_found = True
                        break
                    elif pair[0] in clusters[k] and pair[1] in clusters[k]:
                        cluster_found = True
                        break
                if not cluster_found:
                    clusters[f'Cluster {i}'] = [pair[0], pair[1]]
                    i += 1

        if verbose:
            for cluster in clusters.keys():
                print(f"{cluster}: {clusters[cluster]}")

        return clusters, count
    
    def save_all_clustering_results(self, features, analysis='single', method='Silhouette'):
        """
        Save all the clustering results

        Parameters
        ----------
        features : pandas.DataFrame
            The dataframe containing the features
        analysis : str
            The type of analysis to perform. Either 'single' or 'pair' or 'all
        method : str
            The method to use for clustering. Either 'Elbow' or 'Silhouette'
        """
        features_grouping = self.utilities.group_features_by_label(features, affect_subject=True)
        pbar = tqdm(total=2, desc="Saving Clustering Results")
        pbar.set_description(f"Saving Clustering Results")
        for projections in ['Sammon', 'PCA']:
            pbar.set_description(f"Extracting the {projections} projections")
            clusters_dict, feature_importance = self.get_subjects_clusters(features_grouping, analysis=analysis, projections=projections, method=method, save=True)
            clusters_dict_consistency, _ = self.get_subjects_clusters(features_grouping, analysis=analysis, projections=projections, nb_clusters=2, method=method, save=False)
            clusters, _ = self.get_subject_consistency(clusters_dict_consistency, verbose=False, min_nb_emotions=4)
            plt.ioff()
            pbar.update(1)
        
        clusters_json = json.dumps(clusters_dict, indent=4, default=self.utilities.default_converter)
        clusters_json = clusters_json.replace('\n            ', '').replace('\n        ]', ' ]')
        with open(f'Analysis/Cluster Analysis/SubjectClustering.json', 'w') as file:
            file.write(clusters_json)
        with open(f'Analysis/Cluster Analysis/SubjectConsistency.json', 'w') as file:
            json.dump(clusters, file, indent=4, default=self.utilities.default_converter)
        with open(f'Analysis/Cluster Analysis/FeatureImportance.json', 'w') as file:
            json.dump(feature_importance, file, indent=4, default=self.utilities.default_converter)    
        for nb_emotions in [2,3]:
            pbar.set_description(f'Extracting the intersections for {nb_emotions} emotions')
            self.plot_intersections(clusters_dict, nb_emotions=nb_emotions, save=True)
            plt.ioff()
        pbar.close()

class GENERAL:

    def __init__(self, path, include_neutral_emotion, desired_measurement, subjects = 'All', subjects_to_remove = [], verbose = False, increase_empatica_emotion_window=False):

        """
        Initiliaze the general class

        Parameters
        ----------
        path : str
            The path to the data.
        include_neutral_emotion : bool
            Whether to include the neutral emotion or not.
        desired_measurement : list
            The list of the data types to extract the features from.
        subjects : list or str
            The list of the subjects to consider or 'All' if all of them are considered.
        subjects_to_remove : list 
            The list of the subjects to not consider.
        verbose : bool
            Whether to print the progress bar or not.
        """

        self.empatica = EMPATICA()
        self.transcript = TRANSCRIPT()
        self.audio = AUDIO()
        self.webcam = WEBCAM()
        self.gopro = GOPRO()
        self.sre = SRE()
        self.face_reader = FaceReader()
        self.utilities =  UTILITIES()

        self.path = path
        self.include_neutral_emotion = include_neutral_emotion
        self.desired_measurement = desired_measurement
        self.min_vid_duration = 60 # In seconds
        self.increase_empatica_emotion_window = increase_empatica_emotion_window

        self.subjects = subjects
        self.subjects_to_remove = subjects_to_remove
        self.verbose = verbose

        timestamp_path = os.path.join(self.path, 'Timestamp', 'osher_time_stamps_final.xlsx')

        self.emotion_timing_frontal, self.go_pro_clap_times_frontal = self.utilities._get_timings(timestamp_path, 'frontal')
        self.emotion_timing_lateral, self.go_pro_clap_times_lateral = self.utilities._get_timings(timestamp_path, 'left')

        self.reported_emotions = self.sre._extract_data(self.path)

        if os.path.exists("metadata/webcam_clap_time.json") and os.path.exists("metadata/audio_clap_time.json") and os.path.exists("metadata/empatica_shift.json"):
            with open("metadata/webcam_clap_time.json", 'r') as json_file:
                self.webcam_clap_time = json.load(json_file)
            
            with open("metadata/audio_clap_time.json", 'r') as json_file:
                self.audio_clap_time = json.load(json_file)

            with open("metadata/empatica_shift.json", 'r') as json_file:
                self.empatica_shift = json.load(json_file)
        elif os.path.exists("../metadata/webcam_clap_time.json") and os.path.exists("../metadata/audio_clap_time.json") and os.path.exists("../metadata/empatica_shift.json"):
            with open("../metadata/webcam_clap_time.json", 'r') as json_file:
                self.webcam_clap_time = json.load(json_file)
            
            with open("../metadata/audio_clap_time.json", 'r') as json_file:
                self.audio_clap_time = json.load(json_file)

            with open("../metadata/empatica_shift.json", 'r') as json_file:
                self.empatica_shift = json.load(json_file)
        else:
            print("The metadata files are missing.")

        os.makedirs('logs', exist_ok=True)

        datetime_str = datetime.now().strftime("%Y%m%d-%H%M%S")

        self.json_logging_filename = os.path.join('logs', f'Logs_{datetime_str}.json')
    
    def get_data(self):
        """
        Get the data from the files

        Returns
        -------
        data : dict
            The dictionary that contains the data.
        tags : dict
            The dictionary that contains the tags.
        """

        self.utilities._log_msg(self.json_logging_filename, "Loading The Data... ")
        
        data = dict()
        tags = dict()

        keys_to_del = []

        if self.subjects == 'All':
            subjects = self.emotion_timing_frontal.keys()
        else:
            subjects = self.subjects
        
        if self.subjects_to_remove != []:
            self.utilities._log_msg(self.json_logging_filename, f"   The following subjects will not be considered: {self.subjects_to_remove} ")
            subjects = [subject for subject in subjects if subject not in self.subjects_to_remove]

        nb_subjects = len(subjects)

        pbar = tqdm(total=nb_subjects, desc="Loading Data")

        for subject_id in subjects:

            pbar.set_description('Loading Data for Subject {}'.format(subject_id))
            
            data[subject_id] = dict()
            tags[subject_id] = dict()

            if 'Empatica' in self.desired_measurement:

                empatica_dir = os.path.join(self.path, str(subject_id), 'Empatica')
                
                if os.path.exists(empatica_dir) and os.listdir(empatica_dir) != []:
                    empatica_data_path = os.path.join(empatica_dir, os.listdir(empatica_dir)[0])
                    data[subject_id]['Empatica'], tags[subject_id]['Empatica'] = self.empatica.get_data(empatica_data_path, self.verbose)
                else:
                    keys_to_del.append(subject_id)
                    self.utilities._log_msg(self.json_logging_filename, f"   The Empatica data for Subject {subject_id} doesn't exists. Erasing Subject..")
                    nb_subjects -= 1
                    pbar.total = nb_subjects
                    pbar.refresh()
                    continue

            if 'Transcript' in self.desired_measurement:
                
                transcript_dir = os.path.join(self.path, str(subject_id), 'Transcript')
                
                if os.path.exists(transcript_dir) and os.listdir(transcript_dir) != []:
                    transcript_data_path = os.path.join(transcript_dir, os.listdir(transcript_dir)[0])
                    data[subject_id]['Transcript'] = self.transcript.get_data(transcript_data_path, include_neutral=self.include_neutral_emotion)
                else:
                    keys_to_del.append(subject_id)
                    self.utilities._log_msg(self.json_logging_filename, f"   The Transcript data for Subject {subject_id} doesn't exists. Erasing Subject..")
                    nb_subjects -= 1
                    pbar.total = nb_subjects
                    pbar.refresh()
                    continue
            
            if 'Audio' in self.desired_measurement:
                audio_dir = os.path.join(self.path,str(subject_id),'Audio Recorder')
                
                if os.path.exists(audio_dir) and os.listdir(audio_dir) != []:
                    filename = os.path.join(audio_dir, os.listdir(audio_dir)[0])
                    data[subject_id]['Audio'] = self.audio.get_data(filename, self.emotion_timing_frontal[subject_id])
                else:
                    keys_to_del.append(subject_id)
                    self.utilities._log_msg(self.json_logging_filename, f"   The audio data for Subject {subject_id} doesn't exists. Erasing Subject..")
                    nb_subjects -= 1
                    pbar.total = nb_subjects
                    pbar.refresh()
                    continue

            if 'Webcam' in self.desired_measurement:

                webcam_dir = os.path.join(self.path, str(subject_id), 'Webcam')

                if os.path.exists(webcam_dir) and os.listdir(webcam_dir) != []:
                    webcam_data_path = os.path.join(webcam_dir, os.listdir(webcam_dir)[0])
                    if '._' in webcam_data_path:
                        webcam_data_path = webcam_data_path.replace('._', '')
                    data[subject_id]['Webcam'] = self.webcam.get_data(webcam_data_path)
                    if data[subject_id]['Webcam']['duration'] < self.min_vid_duration:
                        keys_to_del.append(subject_id)
                        self.utilities._log_msg(self.json_logging_filename, f"   The Webcam data for Subject {subject_id} is less than {self.min_vid_duration} seconds. Erasing Subject..")
                        nb_subjects -= 1
                        pbar.total = nb_subjects
                        pbar.refresh()
                        continue
                else:
                    keys_to_del.append(subject_id)
                    self.utilities._log_msg(self.json_logging_filename, f"   The Webcam data for Subject {subject_id} doesn't exists. Erasing Subject..")
                    nb_subjects -= 1
                    pbar.total = nb_subjects
                    pbar.refresh()
                    continue
            
            if 'GoPro' in self.desired_measurement:

                gopro_dir = os.path.join(self.path, str(subject_id), 'GoPros')

                if os.path.exists(gopro_dir) and os.listdir(gopro_dir) != []:
                    data[subject_id]['GoPro'] = dict()

                    gopro_frontal_filepath = os.path.join(gopro_dir, 'Frontal.mp4')
                    gopro_lateral_filepath = os.path.join(gopro_dir, 'Lateral - Left.mp4')

                    if os.path.exists(gopro_frontal_filepath):
                        data[subject_id]['GoPro']['Frontal'] = self.gopro.get_data(gopro_frontal_filepath)
                    else:
                        keys_to_del.append(subject_id)
                        self.utilities._log_msg(self.json_logging_filename, f"   The GoPro Frontal data for Subject {subject_id} doesn't exists. Erasing Subject..")
                        nb_subjects -= 1
                        pbar.total = nb_subjects
                        pbar.refresh()
                        continue
                        
                    if os.path.exists(gopro_lateral_filepath):
                        data[subject_id]['GoPro']['Lateral'] = self.gopro.get_data(gopro_lateral_filepath)
                    else:
                        keys_to_del.append(subject_id)
                        self.utilities._log_msg(self.json_logging_filename, f"   The GoPro Lateral - Left data for Subject {subject_id} doesn't exists. Erasing Subject..")
                        nb_subjects -= 1
                        pbar.total = nb_subjects
                        pbar.refresh()
                        continue

                else:
                    keys_to_del.append(subject_id)
                    self.utilities._log_msg(self.json_logging_filename, f"   The GoPro data for Subject {subject_id} doesn't exists. Erasing Subject..")
                    nb_subjects -= 1
                    pbar.total = nb_subjects
                    pbar.refresh()
                    continue

            if 'SRE' in self.desired_measurement:

                data[subject_id]['SRE'] = self.reported_emotions[subject_id]

                if data[subject_id]['SRE'] == "No Self Reported data":
                    keys_to_del.append(subject_id)
                    self.utilities._log_msg(self.json_logging_filename, f"   The Self Reported Emotions data for Subject {subject_id} doesn't exists. Erasing Subject..")
                    nb_subjects -= 1
                    pbar.total = nb_subjects
                    pbar.refresh()
                    continue

            if 'FaceReader' in self.desired_measurement:

                face_reader_dir = os.path.join(self.path, 'facereaderexcelfiles')

                if os.path.exists(face_reader_dir) and os.listdir(face_reader_dir) != []:

                    face_reader_file_path = None

                    filename = 'MONSCE facereader analysis redone 3.2.24.xlsx'

                    face_reader_file_path = os.path.join(face_reader_dir, filename)
                    
                    data[subject_id]['FaceReader'] = self.face_reader.get_data(face_reader_file_path, subject_id, include_neutral=self.include_neutral_emotion)

                    if self.include_neutral_emotion:
                        nb_emotions = 5
                    else:
                        nb_emotions = 4

                    if len(list(data[subject_id]['FaceReader'].keys()))!=nb_emotions:
                        print('here', len(list(data[subject_id]['FaceReader'].keys())), list(data[subject_id]['FaceReader'].keys()))
                        keys_to_del.append(subject_id)
                        self.utilities._log_msg(self.json_logging_filename, f"   The FaceReader data for Subject {subject_id} is not complete. Erasing Subject..")
                        nb_subjects -= 1
                        pbar.total = nb_subjects
                        pbar.refresh()
                    
                    if face_reader_file_path is None:
                        keys_to_del.append(subject_id)
                        self.utilities._log_msg(self.json_logging_filename, f"   The FaceReader data for Subject {subject_id} doesn't exists. Erasing Subject..")
                        nb_subjects -= 1
                        pbar.total = nb_subjects
                        pbar.refresh()
                        continue

            pbar.update(1)

        for key in keys_to_del:
            del data[key]

        pbar.set_description('Loading Data')

        pbar.close()

        self.utilities._log_msg(self.json_logging_filename, "Loading Completed")
        self.utilities._log_msg(self.json_logging_filename)

        return data, tags
    
    def label_data(self, data):
        """
        Label the data.
        
        Parameters
        ----------
        data : dict
            The dictionary that contains the data.
        
        """
        self.utilities._log_msg(self.json_logging_filename, "Labelling The Data... ")

        keys_to_del = []
        subjects = data.keys()
        nb_subjects = len(subjects)

        pbar = tqdm(total=nb_subjects, desc="Labelling Data")

        for subject_id in data.keys():
            pbar.set_description('Labelling Data for Subject {}'.format(subject_id))
            if 'Empatica' in data[subject_id].keys():
                self.empatica.label_data(data[subject_id]['Empatica'], self.emotion_timing_frontal[subject_id], self.empatica_shift[str(subject_id)], include_neutral=self.include_neutral_emotion, verbose=self.verbose, increase_window=self.increase_empatica_emotion_window)
                for measurement in data[subject_id]['Empatica'].keys():
                    if len(list(data[subject_id]['Empatica'][measurement]['data']['label'].unique())) != 5:
                        keys_to_del.append(subject_id)
                        self.utilities._log_msg(self.json_logging_filename, f"   The Empatica data for Subject {subject_id} is not complete. Erasing Subject..")
                        nb_subjects -= 1
                        pbar.total = nb_subjects
                        pbar.refresh()
                        break
                if subject_id in keys_to_del:
                    continue
            if 'Audio' in data[subject_id].keys():
                self.audio.label_data(data[subject_id]['Audio'], self.emotion_timing_frontal[subject_id], self.audio_clap_time[str(subject_id)], self.go_pro_clap_times_frontal[subject_id], include_neutral=self.include_neutral_emotion, verbose=self.verbose)
                if len(list(data[subject_id]['Audio']['data']['label'].unique())) != 5:
                    keys_to_del.append(subject_id)
                    self.utilities._log_msg(self.json_logging_filename, f"   The Audio data for Subject {subject_id} is not complete. Erasing Subject..")
                    nb_subjects -= 1
                    pbar.total = nb_subjects
                    pbar.refresh()
                    continue
            if 'Webcam' in data[subject_id].keys():
                self.webcam.label_data(data[subject_id]['Webcam'], self.emotion_timing_frontal[subject_id], self.webcam_clap_time[str(subject_id)], self.go_pro_clap_times_frontal[subject_id], include_neutral=self.include_neutral_emotion, verbose=self.verbose)
                if len(list(data[subject_id]['Webcam']['label'].keys())) != 5:
                    keys_to_del.append(subject_id)
                    self.utilities._log_msg(self.json_logging_filename, f"   The Webcam data for Subject {subject_id} is not complete. Erasing Subject..")
                    nb_subjects -= 1
                    pbar.total = nb_subjects
                    pbar.refresh()
                    continue
            if 'GoPro' in data[subject_id].keys():
                if 'Frontal' in data[subject_id]['GoPro'].keys():
                    self.gopro.label_data(data[subject_id]['GoPro']['Frontal'], self.emotion_timing_frontal[subject_id], self.include_neutral_emotion)
                    if len(list(data[subject_id]['GoPro']['Frontal']['label'].keys())) != 5:
                        keys_to_del.append(subject_id)
                        self.utilities._log_msg(self.json_logging_filename, f"   The GoPro Frontal data for Subject {subject_id} is not complete. Erasing Subject..")
                        nb_subjects -= 1
                        pbar.total = nb_subjects
                        pbar.refresh()
                        continue
                if 'Lateral' in data[subject_id]['GoPro'].keys():
                    self.gopro.label_data(data[subject_id]['GoPro']['Lateral'], self.emotion_timing_lateral[subject_id], self.include_neutral_emotion)
                    if len(list(data[subject_id]['GoPro']['Lateral']['label'].keys())) != 5:
                        keys_to_del.append(subject_id)
                        self.utilities._log_msg(self.json_logging_filename, f"   The GoPro Lateral data for Subject {subject_id} is not complete. Erasing Subject..")
                        nb_subjects -= 1
                        pbar.total = nb_subjects
                        pbar.refresh()
                        continue
            
            pbar.update(1)
        
        for key in keys_to_del:
            del data[key]

        pbar.set_description('Labelling Data')
        pbar.close()

        self.utilities._log_msg(self.json_logging_filename, "Labelling Completed")
        self.utilities._log_msg(self.json_logging_filename)

class CORRELATION():

    def __init__(self, desired_measurement):
        """
        Initialize the utilities class
        """

        self.features = FEATURES(desired_measurement)
        self.utilities = UTILITIES()

        self.correlation_thresholds = {
            "No or Negligible correlation": [0, 0.1],
            "Weak correlation": [0.1, 0.3],
            "Moderate correlation": [0.3, 0.5],
            "Good correlation": [0.5, 0.7],
            "Strong correlation": [0.7, 1],
        }

        self.correlation_thresholds_grouped = {
            "No or Negligible correlation": [0, 0.1],
            "Weak correlation": [0.1, 0.3],
            "Moderate correlation": [0.3, 0.5],
            "Good correlation": [0.5, 1],
            "Strong correlation": [0.7, 1],
        }

        self.correlations_to_study = self.utilities._group_desired_measurement_for_correlation(desired_measurement)

        self.emotion_valence = {
            'Pride': 1,
            'Joy': 1,
            'Frustration': -1,
            'Shame': -1,
            'Neutral': 0
        }

        self.subject_mapping =  self.utilities._get_subject_mapping()
    
    def _get_correlated_features(self, correlation_matrix, threshold_min=0.5, threshold_max = 1):
        """
        Get the features that are correlated above a certain threshold.

        Parameters
        ----------
        correlation_matrix : pandas.DataFrame
            The correlation matrix
        threshold_min : float
            The minimum threshold
        threshold_max : float
            The maximum threshold

        Returns
        -------
        features_to_consider : dict
            The dict of the features to consider

        """

        features_to_consider = dict()

        for df1_feature in correlation_matrix.index:
            for df2_feature in correlation_matrix.columns:
                if threshold_min <= abs(correlation_matrix.loc[df1_feature, df2_feature]) < threshold_max and df1_feature != df2_feature:
                    if (df2_feature, df1_feature) not in list(features_to_consider.keys()) and (df1_feature, df2_feature) not in list(features_to_consider.keys()):
                        if df1_feature.split('_')[0] != df2_feature.split('_')[0]: #To avoid having the same feature selected
                            features_to_consider[(df1_feature, df2_feature)] = correlation_matrix.loc[df1_feature, df2_feature]
                    
        return features_to_consider
    
    def group_correlated_features(self, corr_matrices, desired_measurement, group, group_thres = False, valence_analysis = False, time_windows = False, save=False):
        """
        Group the correlated features depending on their correlation coefficient

        Parameters
        ----------
        corr_matrices : dict
            The dictionary containing the correlation matrices
        desired_measurement : list
            The list of the data types to extract the features from.
        group : str
            The group to consider for the correlation
        group_thres : bool
            Whether to group the threshold groups (Good Correlation would be from 0.5 ot 1)
        valence_analysis : bool
            Whether the analysis include the valence or not
        time_windows : bool
            Whether the analysis include the time windows or not
        save : bool
            Whether to save the grouped correlated features or not

        Returns
        -------
        features_correlation : dict
            The dictionary containing the grouped correlated features
        """

        features_correlation = dict()

        if not time_windows:
            for correlation_type in self.correlation_thresholds.keys():
                features_correlation[correlation_type] = dict()
                if valence_analysis:
                    for data_to_study in ['Empatica', 'Transcript', 'Audio', 'Webcam', 'GoPro', 'SRE']:
                        correlation_matrix = corr_matrices[data_to_study]
                        min_thres, max_thres = self.correlation_thresholds[correlation_type]
                        features_correlation[correlation_type][data_to_study] = self._get_correlated_features(correlation_matrix, min_thres, max_thres)
                else:
                    for cor_to_study in self.correlations_to_study.keys():
                        correlation_matrix = corr_matrices[cor_to_study]
                        min_thres, max_thres = self.correlation_thresholds[correlation_type]
                        if group_thres:
                            max_thres
                        features_correlation[correlation_type][cor_to_study] = self._get_correlated_features(correlation_matrix, min_thres, max_thres)
        else:
            for emotion in corr_matrices.keys():
                features_correlation[emotion] = dict()
                for subject_id in corr_matrices[emotion].keys():
                    features_correlation[emotion][subject_id] = dict()
                    correlation_matrix = corr_matrices[emotion][subject_id]
                    for correlation_type in self.correlation_thresholds.keys():
                        min_thres, max_thres = self.correlation_thresholds[correlation_type]
                        features_correlation[emotion][subject_id][correlation_type] = self._get_correlated_features(correlation_matrix, min_thres, max_thres)

        if save:
            if type(desired_measurement) is list or type(desired_measurement) is tuple:
                desired_measurement = desired_measurement[0] + '-' + desired_measurement[1]
                    
            os.makedirs(f'Analysis/Correlation/{group}/{desired_measurement}', exist_ok=True)

            with open(f'Analysis/Correlation/{group}/{desired_measurement}/features_correlation.json', 'w') as file:
                json.dump(self.utilities._convert_keys_to_string(features_correlation), file, indent=4)

        return features_correlation

    def get_discrete_windows_correlation_matrices(self, features_grouping, feature_names, include_pairs = False, method='kendall'):
        """
        Get correlation matrices for each emotion and each window

        Parameters
        ----------
        features_grouping : dict
            Dictionary with features grouped by label 
        feature_names : dict
            The dictionary containing the names of the features for each modality
        include_pairs : bool
            Whether to group the measurement into pairs or not.
        method : str
            The method to use for the correlation. Either 'pearson', 'spearman' or 'kendall'.     

        Returns
        -------
        dict
            Dictionary with correlation matrices for each emotion and each window
        """

        corr_matrices = dict()

        for emotion in features_grouping.keys():
            corr_matrices[emotion] = dict()
            for row in range(len(features_grouping[emotion])):

                df1_features = features_grouping[emotion].iloc[row].drop('label')
                df2_features = features_grouping[emotion].iloc[row].drop('label')

                if include_pairs:
                    measurements = list(feature_names.keys())
                    df1_features = df1_features.loc[df1_features.index.intersection(feature_names[measurements[0]])]
                    df2_features = df2_features.loc[df2_features.index.intersection(feature_names[measurements[1]])]                    
                    
                correlation_matrix = pd.DataFrame(index=df1_features.index, columns=df2_features.index)

                for df1_feature in df1_features.index:
                    for df2_feature in df2_features.index:
                        try:
                            
                            f1 = df1_features[df1_feature]
                            f2 = df2_features[df2_feature]
                                
                            if method == 'pearson':
                                correlation, _ = pearsonr(f1, f2)
                            elif method == 'spearman':
                                correlation, _ = spearmanr(f1, f2)
                            elif method == 'kendall':
                                correlation, _ = kendalltau(f1, f2)

                            correlation_matrix.loc[df1_feature, df2_feature] = correlation

                        except Exception as e:
                            print(row)
                            print('Error for features: ', df1_feature, df2_feature)
                            print(e)
                            correlation_matrix.loc[df1_feature, df2_feature] = float('nan')

                corr_matrices[emotion][self.subject_mapping[row]] = correlation_matrix.astype(float)

        return corr_matrices

    def get_pair_count_across_subjects(self, features_correlation, desired_measurement, desired_thresholds, group, group_cat = False, method = 'kendall', save=False):
        """
        Extract and count the number of iteration of every pair of features for each emotion and each threshold accross the subjects.

        Parameters
        ----------
        features_correlation : dict
            Dictionary with the correlated features
        desired_measurement : str
            The desired measurement
        desired_thresholds : list
            The desired thresholds
        group : str
            The group of the subjects
        group_cat : bool
            Adds the feature that are strongly correlated to the one that have a "good correlation".
        method : str
            The method to use for the correlation. Either 'pearson', 'spearman' or 'kendall'
        save : bool
            Whether to save the results or not

        Returns
        -------
        pair_count_dict : dict
            Dictionary with the pair counts
        """
        
        if type(desired_measurement) is list or type(desired_measurement) is tuple:
            desired_measurement = desired_measurement[0] + '-' + desired_measurement[1]
        else:
            desired_measurement = desired_measurement

        pair_count_dict = dict()

        if group_cat:
            for emotion in features_correlation.keys():
                for subject_id in features_correlation[emotion].keys():
                    for pair, corr_coeff in features_correlation[emotion][subject_id]['Strong correlation'].items():
                        if pair not in features_correlation[emotion][subject_id]['Good correlation'].keys():
                            features_correlation[emotion][subject_id]['Good correlation'].update({pair: corr_coeff})

        for emotion in features_correlation.keys():

            pair_count_dict[emotion] = dict()

            if save:
                os.makedirs(f'Analysis/Correlation/{group}/{desired_measurement}/PairCount/AcrossSubjects/{emotion}', exist_ok=True)

            for threshold in desired_thresholds:
                l = list()
                for subject_id in features_correlation[emotion].keys():
                    for pair in features_correlation[emotion][subject_id][threshold]:
                        l.append(pair)
                _pair_counts = Counter(l)
                pair_counts = dict()
                for pair in _pair_counts.keys():
                    if pair[0] in pair_counts.keys():
                        pair_counts[pair[0]].update({pair[1]: _pair_counts[pair]})
                    else:
                        pair_counts[pair[0]] = {pair[1]: _pair_counts[pair]}

                for feature in pair_counts.keys():
                    pair_counts[feature] = dict(sorted(pair_counts[feature].items(), key=lambda item: item[1], reverse=True))

                pair_count_dict[emotion][threshold] = pair_counts

                if save:
                    with open(f'Analysis/Correlation/{group}/{desired_measurement}/PairCount/AcrossSubjects/{emotion}/{threshold.replace(" ", "_")}.json', 'w') as file:
                        file.write(json.dumps(pair_counts, indent=4))

        return pair_count_dict
    
    def get_pair_count_across_emotions(self, features_correlation, desired_measurement, desired_thresholds, group, emotions = ['Shame', 'Pride', 'Frustration', 'Joy'], group_cat= False, save=False):
        """
        Get the count of pairs of features that are correlated across all emotions for each subject

        Parameters
        ----------
        features_correlation : dict
            Dictionary containing the correlation matrices for each emotion and each subject
        desired_measurement : str
            The desired measurement
        desired_thresholds : slist
            The desired thresholds
        group : str
            The group of the subjects
        group_cat : bool
            Adds the feature that are strongly correlated to the one that have a "good correlation".
        save : bool, optional
            Whether to save the result in a file or not

        Returns
        -------
        pair_count_across_emotions : dict
            Dictionary containing the count of pairs of features that are correlated across all emotions for each subject
        """

        if type(desired_measurement) is list or type(desired_measurement) is tuple:
            desired_measurement = desired_measurement[0] + '-' + desired_measurement[1]
        else:
            desired_measurement = desired_measurement
        
        subjects = set()
        for emotion in emotions:
            if subjects == set():
                subjects = set(list(features_correlation[emotion].keys()))
            else:
                subjects = subjects.intersection(list(features_correlation[emotion].keys()))
        subjects = np.sort(list(subjects))

        pair_count_across_emotions = dict()

        if group_cat:
            for emotion in emotions:
                for subject in subjects:
                    for pair, corr_coeff in features_correlation[emotion][subject]['Strong correlation'].items():
                        if pair not in features_correlation[emotion][subject]['Good correlation'].keys():
                            features_correlation[emotion][subject]['Good correlation'].update({pair: corr_coeff})

        for threshold in desired_thresholds:
            pair_count_across_emotions[threshold] = dict()
            for subject in subjects:
                pair_count_across_emotions[threshold][subject] = dict()
                l = list()
                for emotion in emotions:
                    for pair in features_correlation[emotion][subject][threshold].keys():
                        l.append(pair)
                _pair_counts = Counter(l)
                pair_counts = dict()
                for pair in _pair_counts.keys():
                    if pair[0] in pair_counts.keys():
                        pair_counts[pair[0]].update({pair[1]: _pair_counts[pair]})
                    else:
                        pair_counts[pair[0]] = {pair[1]: _pair_counts[pair]}

                for feature in pair_counts.keys():
                    pair_counts[feature] = dict(sorted(pair_counts[feature].items(), key=lambda item: item[1], reverse=True))

                pair_count_across_emotions[threshold][subject] = pair_counts

            if save:
                pair_count_across_emotions_converted = {str(key): value for key, value in pair_count_across_emotions[threshold].items()}
                os.makedirs(f'Analysis/Correlation/{group}/{desired_measurement}/PairCount/AcrossEmotions', exist_ok=True)
                with open(f'Analysis/Correlation/{group}/{desired_measurement}/PairCount/AcrossEmotions/{threshold.replace(" ", "_")}.json', 'w') as file:
                    json.dump(pair_count_across_emotions_converted, file, indent=4)

        return pair_count_across_emotions
    
    def get_correlation_consistency(self, features_correlation, group, group_cat = False, desired_measurement = None, save=False):
        """
        Get the consistency of the correlation across the emotions for each pair of features

        Parameters
        ----------
        features_correlation : dict
            Dictionary containing the correlation scores for each emotion and each subject
        group : str
            The group of the subjects
        group_cat : bool
            Adds the feature that are strongly correlated to the one that have a "good correlation".
        desired_measurement : str
            The desired measurement
        save : bool
            Whether to save the result in a file or not
        
        Returns
        -------
        correlation_consistency : dict
            Dictionary containing the consistency of the correlation across the emotions for each pair of features
        """

        if group_cat:
            for emotion in features_correlation.keys():
                for subject_id in features_correlation[emotion].keys():
                    for pair, corr_coeff in features_correlation[emotion][subject_id]['Strong correlation'].items():
                        if pair not in features_correlation[emotion][subject_id]['Good correlation'].keys():
                            features_correlation[emotion][subject_id]['Good correlation'].update({pair: corr_coeff})
            desired_threshold = 'Good correlation'
        else:
            desired_threshold = 'Strong correlation'

        correlation_consistency = dict()
        for subject_id in features_correlation['Shame'].keys():
            for emotion in features_correlation.keys():
                for pair in features_correlation[emotion][subject_id][desired_threshold].keys():
                    pair = f"{pair[0].split('_')[0]}--{pair[1].split('_')[0]}"
                    if 'accX' in pair:
                        pair = pair.replace('accX', 'acc')
                    if 'accY' in pair:
                        pair = pair.replace('accY', 'acc')
                    if 'accZ' in pair:
                        pair = pair.replace('accZ', 'acc')
                    if pair == 'acc--acc':
                        continue
                    if pair in correlation_consistency.keys():
                        if subject_id in correlation_consistency[pair].keys():
                            if emotion not in correlation_consistency[pair][subject_id]:
                                correlation_consistency[pair][subject_id].append(emotion)
                        else:
                            correlation_consistency[pair][subject_id] = [emotion]
                    else:
                        correlation_consistency[pair] = dict()
                        correlation_consistency[pair][subject_id] = [emotion]
        pairs_to_remove = []
        for pair in correlation_consistency.keys():
            grouped_emotions = dict()
            for subject_id, emotions in correlation_consistency[pair].items():
                sorted_emotions = tuple(sorted(emotions))
                if 'Frustration' in sorted_emotions and 'Shame' in sorted_emotions:
                    if 'Negative Valence' not in grouped_emotions.keys():
                        grouped_emotions['Negative Valence'] = [subject_id]
                    else:
                        grouped_emotions['Negative Valence'].append(subject_id)
                elif 'Pride' in sorted_emotions and 'Joy' in sorted_emotions:
                    if 'Positive Valence' not in grouped_emotions.keys():
                        grouped_emotions['Positive Valence'] = [subject_id]
                    else:
                        grouped_emotions['Positive Valence'].append(subject_id)

            if grouped_emotions != dict():
                correlation_consistency[pair] = grouped_emotions
            else:
                pairs_to_remove.append(pair)

        for pair in pairs_to_remove:
            del correlation_consistency[pair]

        if save:
            if type(desired_measurement) is list or type(desired_measurement) is tuple:
                desired_measurement = desired_measurement[0] + '-' + desired_measurement[1]
            else:
                desired_measurement = desired_measurement

            os.makedirs(f'Analysis/Correlation/{group}/{desired_measurement}', exist_ok=True)
            with open(f'Analysis/Correlation/{group}/{desired_measurement}/correlation_consistency.json', 'w') as json_file:
                json.dump(correlation_consistency, json_file, indent=4)

        return correlation_consistency
    
    def select_features_for_correlation(self, all_features, desired_measurement, n_estimators=500, n_splits=8):
        """
        Select the features that represent best the data for each measurement device

        Parameters
        ----------
        all_features : pd.DataFrame
            DataFrame containing all the features
        desired_measurement : list
            List of the desired features
        n_estimators : int
            Number of trees in the forest
        n_splits : int
            Number of folds in the cross-validation

        Returns
        -------
        empatica_selected_features : list
            List of the selected features for the Empatica
        audio_selected_features : list
            List of the selected features for the audio
        webcam_selected_features : list
            List of the selected features for the webcam
        gopro_selected_features : list
            List of the selected features for the GoPro
        facereader_selected_features : list
            List of the selected features for the FaceReader
        """

        features_names = self.utilities.get_feature_names(all_features, desired_measurement)

        estimator = RandomForestClassifier(n_estimators=n_estimators, random_state=42, n_jobs=-1) 

        cv_strategy = StratifiedKFold(n_splits, shuffle=True, random_state=42)

        rfecv = RFECV(estimator=estimator, step=1, cv=cv_strategy, scoring='accuracy')

        if 'Empatica' in desired_measurement:
            empatica_selected_features = self._select_features_for_correlation_per_measurement(all_features, features_names['Empatica'], rfecv, empatica=True)
        else:
            empatica_selected_features = list()
        if 'Audio' in desired_measurement:
            audio_selected_features = self._select_features_for_correlation_per_measurement(all_features, features_names['Audio'], rfecv)
        else:
            audio_selected_features = list()
        if 'Webcam' in desired_measurement:
            webcam_selected_features = self._select_features_for_correlation_per_measurement(all_features, features_names['Webcam'], rfecv)
        else:
            webcam_selected_features = list()
        if 'GoPro' in desired_measurement:
            gopro_selected_features = self._select_features_for_correlation_per_measurement(all_features, features_names['GoPro'], rfecv)
        else:
            gopro_selected_features = list() 
        if 'FaceReader' in desired_measurement:
            facereader_selected_features = self._select_features_for_correlation_per_measurement(all_features, features_names['FaceReader'], rfecv)
        else:
            facereader_selected_features = list()

        all_selected_features = list()
        all_selected_features.extend(empatica_selected_features)
        all_selected_features.extend(audio_selected_features)
        all_selected_features.extend(webcam_selected_features)
        all_selected_features.extend(gopro_selected_features)
        all_selected_features.extend(facereader_selected_features)

        features_per_measurement = [empatica_selected_features, audio_selected_features, webcam_selected_features, gopro_selected_features, facereader_selected_features]

        return all_selected_features, features_per_measurement

    def _select_features_for_correlation_per_measurement(self, all_features, features_names, model, empatica=False):
        """
        Select the features that represent best the data in entry

        Parameters
        ----------
        all_features : pd.DataFrame
            DataFrame containing all the features
        features_names : list
            List containing the names of the features of the measurement given in entry
        model : RFECV
            Model used for the feature selection
        empatica : bool
            If True, the function will select the features for the Empatica

        Returns
        -------
        selected_features : list
            List of the selected features for the the measurement given in entry
        """

        if empatica:

            temp_features_names = []
            bvp_features_names = []
            eda_features_names = []
            acc_features_names = []
            hr_features_names = []

            for name in features_names:
                if 'temp' in name:
                    temp_features_names.append(name)
                elif 'bvp' in name:
                    bvp_features_names.append(name)
                elif 'eda' in name:
                    eda_features_names.append(name)
                elif 'acc' in name:
                    acc_features_names.append(name)
                elif 'hr' in name:
                    hr_features_names.append(name)
                else:
                    print(f'Unrecognized feature name: {name}')

            empatica_all_features = all_features[['label'] + temp_features_names + bvp_features_names + eda_features_names + acc_features_names + hr_features_names]
            empatica_features_names = [temp_features_names, bvp_features_names, eda_features_names, acc_features_names, hr_features_names]

            all_features_averaged = empatica_all_features.copy()

            for feature in empatica_all_features.columns:
                if feature != 'label':
                    for row in range(empatica_all_features.shape[0]):
                        all_features_averaged[feature].iloc[row] = np.mean(ast.literal_eval(empatica_all_features[feature].iloc[row]))

            empatica_selected_features = list()

            for empatica_name in empatica_features_names:
                X = all_features_averaged[empatica_name]
                y = all_features_averaged['label']

                model.fit(X, y)

                selected_features = X.columns[model.support_]

                empatica_selected_features.extend(selected_features)

            return empatica_selected_features
        else:
            all_features = all_features[['label']+features_names]
            all_features_averaged = all_features.copy()

            for feature in all_features.columns:
                if feature != 'label':
                    for row in range(all_features.shape[0]):
                        all_features_averaged[feature].iloc[row] = np.mean(ast.literal_eval(all_features[feature].iloc[row]))

            X = all_features_averaged.drop('label', axis=1)
            y = all_features_averaged['label']

            model.fit(X, y)

            selected_features = X.columns[model.support_]

            return selected_features
        
    def plot_correlation_heatmap(self, corr_matrices, cor_to_study):
        """
        Plot the correlation heatmap for a specific correlation to study.

        Parameters
        ----------
        corr_matrices : dict
            The dictionary containing the correlation matrices
        cor_to_study : str
            The correlation to study
        """

        data_to_study = self.correlations_to_study[cor_to_study]
        correlation_matrix = corr_matrices[cor_to_study]

        plt.figure(figsize=(24, 30))
        sns.heatmap(correlation_matrix, annot=True, fmt=".2f", cmap="coolwarm", cbar=True, cbar_kws={"shrink": 0.5}, square=True)
        plt.title(f' {data_to_study[0]} x {data_to_study[1]} - Feature Correlation Matrix - Shape: {correlation_matrix.shape}')
        plt.show()

    def plot_pair_count_across_emotions(self, correlation, desired_measurement, group, corr_threshold = 'Good correlation', save=False):
        """
        Plot the distribution of the correlation pair counts across emotions

        Parameters
        ----------
        correlation : dict
            Dictionary containing the correlation matrices for each emotion and each subject
        desired_measurement : str or list
            List of the desired feature(s)
        group : str
            The group to consider
        corr_threshold : str
            The correlation threshold to consider
        save : bool
            Whether to save the plots or not
        """
        if type(desired_measurement) is list or type(desired_measurement) is tuple:
            desired_measurement = desired_measurement[0] + '-' + desired_measurement[1]
        else:
            desired_measurement = desired_measurement

        fig, axs = plt.subplots(2, 2, figsize=(20, 12))
        axs = axs.flatten()  

        global_max = 0
        for nb_emotions in [1, 2, 3, 4]:
            for id in correlation[corr_threshold].keys():
                total = 0
                for pair in correlation[corr_threshold][id].keys():
                    for key in correlation[corr_threshold][id][pair].keys():
                        if correlation[corr_threshold][id][pair][key] == nb_emotions:
                            total += 1
                if total > global_max:
                    global_max = total

        for idx, nb_emotions in enumerate([1,2,3,4]):
            d = dict()

            for id in correlation[corr_threshold].keys():
                t = 0
                for pair in correlation[corr_threshold][id].keys():
                    for key in correlation[corr_threshold][id][pair].keys():
                        if correlation[corr_threshold][id][pair][key] == nb_emotions:
                            t += 1
                d[id] = t

            sorted_ids = sorted(d.keys(), key=lambda x: str(x))
            sorted_values = [d[id] for id in sorted_ids]

            ax = axs[idx]
            bars = ax.bar(range(len(sorted_ids)), sorted_values, color='#DC143C')
            ax.set_xticks(range(len(sorted_ids)))
            ax.set_xticklabels(sorted_ids)
            ax.set_xlabel('Subject ID', fontsize=12)
            ax.set_ylabel('Number of correlated pairs', fontsize=12)
            ax.set_title(f'Number of correlated pairs present in {nb_emotions} emotions', fontsize=14)
            ax.tick_params(axis='x', rotation=45, labelsize=10)
            ax.grid(axis='y', linestyle='--', alpha=0.7)

            ax.set_ylim([0, global_max+5])

            for bar in bars:
                yval = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2, yval, int(yval), ha='center', va='bottom', fontsize=10)

        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        fig.suptitle(f'{desired_measurement} - Distribution of Correlation Pair Counts Per Presence in Emotions - Coeff > 0.5', fontsize=16, fontweight='bold')
        
        if save:
            os.makedirs(f'Analysis/Correlation/{group}/{desired_measurement}/PairCount/AcrossEmotions/', exist_ok=True)
            plt.savefig(f'Analysis/Correlation/{group}/{desired_measurement}/PairCount/AcrossEmotions/PairCountAcrossEmotions.png', format='png', dpi=300)
            plt.close()
        else:
            plt.show()

    def plot_pair_count_across_emotions_per_emotion_pair(self, features_correlation, desired_measurement, group, corr_threshold = 'Good correlation', save=False):
        """
        Plot the distribution of the correlation pair counts across emotions

        Parameters
        ----------
        correlation : dict
            Dictionary containing the correlation matrices for each emotion and each subject
        desired_measurement : str or list
            List of the desired feature(s)
        group : str
            The group to consider
        corr_threshold : str
            The correlation threshold to consider
        save : bool
            Whether to save the plots or not
        """
        if type(desired_measurement) is list or type(desired_measurement) is tuple:
            desired_measurement = desired_measurement[0] + '-' + desired_measurement[1]
        else:
            desired_measurement = desired_measurement

        nrows, ncols = 2, 3
        fig, axs = plt.subplots(nrows, ncols, figsize=(20,12))  
        axs = axs.flatten() 
        emotion_combinations = list(combinations(['Shame', 'Joy', 'Pride', 'Frustration'], 2))

        global_max = 0
        for emotion_combination in emotion_combinations:
            correlation = self.get_pair_count_across_emotions(features_correlation, desired_measurement, ['Good correlation', 'Strong correlation'], group, emotions=emotion_combination, group_cat=True, save=False)
            
            for id in correlation[corr_threshold].keys():
                total_pairs = 0
                for pair in correlation[corr_threshold][id].keys():
                    for key in correlation[corr_threshold][id][pair].keys():
                        if correlation[corr_threshold][id][pair][key] == 2:
                            total_pairs += 1
                if total_pairs > global_max:
                    global_max = total_pairs

        for idx, emotion_combination in enumerate(emotion_combinations):
            correlation = self.get_pair_count_across_emotions(features_correlation, desired_measurement, ['Good correlation', 'Strong correlation'], group, emotions=emotion_combination, group_cat=True, save=False)
            
            d = dict()
            for id in correlation[corr_threshold].keys():
                t = 0
                for pair in correlation[corr_threshold][id].keys():
                    for key in correlation[corr_threshold][id][pair].keys():
                        if correlation[corr_threshold][id][pair][key] == 2:
                            t += 1
                d[id] = t

            sorted_ids = sorted(d.keys(), key=lambda x: str(x))
            sorted_values = [d[id] for id in sorted_ids]

            axs[idx].bar(range(len(sorted_ids)), sorted_values, color='#DC143C')
            axs[idx].set_xticks(range(len(sorted_ids)))
            axs[idx].set_xticklabels(sorted_ids, rotation=45, fontsize=10)
            axs[idx].set_xlabel('Subject ID', fontsize=12)
            axs[idx].set_ylabel('Number of correlated pairs', fontsize=12)
            axs[idx].set_title(f'{desired_measurement} - Pair Counts Across Emotions ({emotion_combination[0]}-{emotion_combination[1]})', fontsize=14)
            axs[idx].grid(axis='y', linestyle='--', alpha=0.7)

            axs[idx].set_ylim([0, global_max+5])  

            for bar in axs[idx].containers[0]:
                yval = bar.get_height()
                axs[idx].text(bar.get_x() + bar.get_width()/2, yval, int(yval), ha='center', va='bottom', fontsize=10)

        # Adjust layout for better fit and display
        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        fig.suptitle(f'{desired_measurement} - Distribution of Correlation Pair Counts Across Emotion Pairs - Coeff > 0.5', fontsize=16, fontweight='bold')

        if save:
            os.makedirs(f'Analysis/Correlation/{group}/{desired_measurement}/PairCount/AcrossEmotions/', exist_ok=True)
            plt.savefig(f'Analysis/Correlation/{group}/{desired_measurement}/PairCount/AcrossEmotions/PairCountAcrossEmotionsPairs.png', format='png', dpi=300)
            plt.close()
        else:
            plt.show()

    def plot_pair_count_across_subjects(self, correlation, desired_measurement, group, corr_threshold = 'Good correlation', save=False):
        """
        Plot the distribution of the correlation pair counts across subjects

        Parameters
        ----------
        correlation : dict
            Dictionary containing the correlation matrices for each emotion and each subject
        desired_measurement : str or list
            List of the desired feature(s)
        group : str
            The group to consider
        corr_threshold : str
            The correlation threshold to consider
        save : bool
            Whether to save the plots or not
        """

        if type(desired_measurement) is list or type(desired_measurement) is tuple:
            desired_measurement = desired_measurement[0] + '-' + desired_measurement[1]
        else:
            desired_measurement = desired_measurement

        nb_subjects_list = [5, 10, 15, 20]

        fig, axs = plt.subplots(2, 2, figsize=(20, 12))
        axs = axs.flatten() 

        global_max = 0
        for nb_subjects in nb_subjects_list:
            for emotion in correlation.keys():
                corr = correlation[emotion][corr_threshold]
                total_pairs = sum(corr[pair][key] >= nb_subjects for pair in corr for key in corr[pair])
                global_max = max(global_max, total_pairs)

        for idx, nb_subjects in enumerate(nb_subjects_list):
            d = dict()
            for emotion in correlation.keys():
                corr = correlation[emotion][corr_threshold]
                t = 0
                for pair in corr.keys():
                    for key in corr[pair].keys():
                        if corr[pair][key] >= nb_subjects:            
                            t += 1

                d[emotion] = t

            ax = axs[idx]
            bars = ax.bar(d.keys(), d.values(), color='#DC143C')
            ax.set_xlabel('Emotion', fontsize=12)
            ax.set_ylabel('Number of correlated pairs', fontsize=12)
            ax.set_title(f'Number of correlated pairs present in at least {nb_subjects} subjects', fontsize=14)
            ax.tick_params(axis='x', labelrotation=45, labelsize=10)
            ax.grid(axis='y', linestyle='--', alpha=0.7)
            ax.set_ylim([0, global_max+5])

            for bar in bars:
                yval = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2, yval, int(yval), ha='center', va='bottom', fontsize=10)

        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        fig.suptitle(f'{desired_measurement} - Distribution of Correlation Pair Counts Across Subjects - Coeff > 0.5', fontsize=16, fontweight='bold')
        
        if save:
            os.makedirs(f'Analysis/Correlation/{group}/{desired_measurement}/PairCount/AcrossSubjects/', exist_ok=True)
            plt.savefig(f'Analysis/Correlation/{group}/{desired_measurement}/PairCount/AcrossSubjects/PairCountAcrossSubjects.png', format='png', dpi=300)
            plt.close()
        else:
            plt.show()

    def save_all_correlation_results(self, desired_measurement, subject_groups, include_pairs = False, only_pairs = False, select_features = False, method='kendall'):
        """
        Save the correlation results for each emotion measurement and each correlation method

        Parameters
        ----------
        desired_measurement : list
            List of the desired features
        subject_group : list
            List of the subjects to consider
        include_pairs : bool
            Whether to group the measurement into pairs or not.
        only_pairs : bool
            Whether to consider only pairs.
        select_features : bool
            Whether to run feature selection or not
        method : str
            The method to use for the correlation. Either 'pearson', 'spearman' or 'kendall'
        """

        _desired_measurement = desired_measurement.copy()

        if include_pairs:
            unique_pairs = list(combinations(_desired_measurement,2))
            _desired_measurement.extend(unique_pairs)

        pbar = tqdm(total=len(_desired_measurement)*len(subject_groups), desc="Extracting Correlation Results")

        for measurement in _desired_measurement:
            if measurement in ['Empatica', 'Transcript', 'Audio', 'Webcam', 'GoPro', 'SRE', 'FaceReader']:
                if only_pairs:
                    continue
                pbar.set_description(f'Extracting Correlation Results for {measurement}...')
                if os.path.exists(f'../computed_features/{measurement}/stand_features_windows.csv'):
                    stand_features = pd.read_csv(f'../computed_features/{measurement}/stand_features_windows.csv')
                else:
                    stand_features = pd.read_csv(f'computed_features/{measurement}/stand_features_windows.csv')
                feature_names = self.utilities.get_feature_names(stand_features, [measurement])
                include_pairs = False
            else:
                pbar.set_description(f'Extracting Correlation Results for {measurement[0]}-{measurement[1]}...')
                if os.path.exists(f'../computed_features/stand_features_windows.csv'):
                    stand_features = pd.read_csv(f'../computed_features/stand_features_windows.csv')
                else:
                    stand_features = pd.read_csv(f'computed_features/stand_features_windows.csv')
                feature_names = self.utilities.get_feature_names(stand_features, measurement)
                include_pairs = True

            for group in subject_groups:
                pbar.set_description(f'Extracting Correlation Results for {group}...')
                features_to_use = self.utilities.select_features_from_subjects(stand_features, subject_groups[group])
                if select_features:
                    all_selected_features, _ = self.select_features_for_correlation(stand_features, measurement, n_estimators=500, n_splits=8)
                    stand_features = stand_features[['label'] + all_selected_features]

                features_grouping = self.utilities.group_features_by_label(features_to_use, windows=True)
                corr_matrices = self.get_discrete_windows_correlation_matrices(features_grouping, feature_names, method = method, include_pairs=include_pairs)
                features_correlation = self.group_correlated_features(corr_matrices, measurement, group, time_windows = True, save=True)
                self.save_correlation_plots_per_subject_per_emotion(features_correlation, measurement, group, save=True)
                pair_count_across_subjects = self.get_pair_count_across_subjects(features_correlation, measurement, ['Good correlation', 'Strong correlation'], group, group_cat = True, save=True)
                self.plot_pair_count_across_subjects(pair_count_across_subjects, measurement, group, save=True)
                pair_count_across_emotions = self.get_pair_count_across_emotions(features_correlation, measurement, ['Good correlation', 'Strong correlation'], group, group_cat = True, save=True)
                self.plot_pair_count_across_emotions(pair_count_across_emotions, measurement, group, save=True)
                self.plot_pair_count_across_emotions_per_emotion_pair(features_correlation, measurement, group, save=True)
                _= self.get_correlation_consistency(features_correlation, group, group_cat=True, desired_measurement=measurement, save=True)

                pbar.update(1)

        pbar.close()

    def save_correlation_plots_per_subject_per_emotion(self, correlation, desired_measurement, group, save=False):
        """
        Saves the correlation plots per subject and per emotion

        Parameters
        ----------
        correlation : dict
            Dictionary containing the correlation matrices for each emotion and each subject
        desired_measurement : str or list
            List of the desired feature(s)
        group : str
            The group of subjects to consider
        save : bool
            Whether to save the plots or not
        """
        if type(desired_measurement) is list or type(desired_measurement) is tuple:
            desired_measurement = desired_measurement[0] + '-' + desired_measurement[1]

        fig, axs = plt.subplots(2, 2, figsize=(15, 10))
        axs = axs.flatten()

        global_max = 0
        for emotion in correlation.keys():
            for subject in correlation[emotion].keys():
                total_count = sum(len(correlation[emotion][subject][correlation_type]) for correlation_type in ['Good correlation', 'Strong correlation'])
                if total_count > global_max:
                    global_max = total_count

        for ax, emotion in zip(axs, correlation.keys()):
            d = dict()
            
            for subject in correlation[emotion].keys():
                d[subject] = dict()
                for correlation_type in ['Good correlation', 'Strong correlation']:
                    d[subject][correlation_type] = len(correlation[emotion][subject][correlation_type])
            
            subject_ids = sorted(d.keys(), key=lambda x: str(x)) 
            good_correlation_counts = [d[subj]['Good correlation'] for subj in subject_ids]
            strong_correlation_counts = [d[subj]['Strong correlation'] for subj in subject_ids]

            positions = range(len(subject_ids))

            ax.bar(positions, good_correlation_counts, label='Good correlation', color='#DC143C')
            ax.bar(positions, strong_correlation_counts, bottom=good_correlation_counts, label='Strong correlation', color='#A30D22')

            ax.set_xticks(positions)
            ax.set_xticklabels(subject_ids, rotation=45, ha="right")
            ax.set_xlabel('Subject ID', fontsize=10)
            ax.set_ylabel('Number of Correlations', fontsize=10)
            ax.set_title(f'Number of Good and Strong Correlations per Subject for {emotion}', fontsize=12)
            ax.tick_params(axis='x', rotation=45)
            ax.legend()

            ax.set_ylim([0, global_max+5])

        fig.suptitle(f'{desired_measurement} - Correlation Counts per Subject and Per emotion', fontsize=16, fontweight='bold')

        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        if save:
            os.makedirs(f'Analysis/Correlation/{group}/{desired_measurement}', exist_ok=True)
            plt.savefig(f'Analysis/Correlation/{group}/{desired_measurement}/CountPerSubjectPerEmotion.png', format='png', dpi=300)
            plt.close()
        else:
            plt.show()

        fig, axs = plt.subplots(2, 2, figsize=(12, 10))
        axs = axs.flatten() 

        for idx, (emotion, ax) in enumerate(zip(correlation.keys(), axs)):
            good_correlation_counts = []
            strong_correlation_counts = []

            for subject in correlation[emotion].keys():
                good_correlation_counts.append(len(correlation[emotion][subject]['Good correlation']))
                strong_correlation_counts.append(len(correlation[emotion][subject]['Strong correlation']))
            
            data = [good_correlation_counts, strong_correlation_counts]

            bplot = ax.boxplot(data, patch_artist=True)
            
            for patch, color in zip(bplot['boxes'], ['#DC143C', '#A30D22']):
                patch.set_facecolor(color)

            for component in ['whiskers', 'caps', 'medians', 'fliers']:
                plt.setp(bplot[component], color='black') 

            ax.set_xticklabels(['Good correlation', 'Strong correlation'])
            ax.set_title(f'{emotion} Correlation Counts')
            ax.set_ylabel('Number of Features')
            ax.grid(axis='y', linestyle='--', alpha=0.6)

        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        fig.suptitle(f'{desired_measurement} - Distribution of Correlation Feature Counts per Emotion', fontsize=16, fontweight='bold')
        if save:
            os.makedirs(f'Analysis/Correlation/{group}/{desired_measurement}', exist_ok=True)
            plt.savefig(f'Analysis/Correlation/{group}/{desired_measurement}/DistributionPerEmotion.png', format='png', dpi=300)
            plt.close()
        else:
            plt.show()
        
class CrossCorrelation:

    def __init__(self, desired_measurement):
        
        self.utilities = UTILITIES()
        self.correlation = CORRELATION(desired_measurement)
        self.subject_mapping =  self.utilities._get_subject_mapping()
        self.desired_measurement = desired_measurement
        
    def get_cross_correlations(self, features_grouping, features_to_cross_corr):
        """
        Compute the cross-correlation between all features for each emotion

        Parameters
        ----------
        features_grouping : dict
            Dictionary of features grouped by emotion
        features_to_cross_corr : list
            List of features to cross-correlate
        
        Returns
        -------
        cross_corr_vectors : dict
            Dictionary of cross-correlation vectors for each emotion
        """

        cross_corr_vectors = dict()

        for emotion in features_grouping.keys():
            cross_corr_vectors[emotion] = dict()
            for row in range(len(features_grouping[emotion])):
                subject_id = self.subject_mapping[row]
                cross_corr_vectors[emotion][subject_id] = dict()
                for pair_to_corr in features_to_cross_corr:
                    try:
                        f1 = features_grouping[emotion][pair_to_corr[0]].iloc[row]
                        f2 = features_grouping[emotion][pair_to_corr[1]].iloc[row]

                        cross_correlation = np.correlate(f1, f2, mode='full')
                        lags = np.arange(-len(f2)+1, len(f1))
                        abs_cross_correlation = np.abs(cross_correlation)

                        max_index = np.argmax(abs_cross_correlation)

                        cross_corr_vectors[emotion][subject_id][f"{pair_to_corr[0]}--{pair_to_corr[1]}"] = {
                            "cross_correlation": cross_correlation,
                            "lags": lags,
                            "max_cross_correlation": abs_cross_correlation[max_index], 
                            "max_lag": lags[max_index]
                        }

                    except Exception as e:
                        print(subject_id)
                        print('Error for features: ', pair_to_corr[0], pair_to_corr[1])
                        print(e)
                        cross_corr_vectors[emotion][subject_id][f"{pair_to_corr[0]}--{pair_to_corr[1]}"] = float('nan')

        return cross_corr_vectors

    def get_features_to_cross_correlate(self, features_names, pair_comparison=False, selected_features=None):
        """
        Get the features to cross correlate.

        Parameters:
        ----------
        features_names : dict
            Dictionary with the features names for each emotion measurement.
        selected_features : list
            List of the selected features.
        
        Returns:
        -------
        features_to_cross_correlate : list
            List with the features to cross correlate. 
        """
        features_to_cross_correlate = list()

        if 'Empatica' in features_names.keys():
            empatica_features = ['temp_mean', 'eda_mean', 'bvp_mean', 'hr_mean', 'accX_mean', 'accY_mean', 'accZ_mean']
        else:
            empatica_features = list()
        if 'Audio' in features_names.keys():
            audio_features = [feature for feature in features_names['Audio'] if 'mean' in feature]
        else:
            audio_features = list()
        if 'Webcam' in features_names.keys():
            webcam_features = [feature for feature in features_names['Webcam'] if 'mean' in feature]
        else:
            webcam_features = list()
        if 'GoPro' in features_names.keys():
            gopro_features = [feature for feature in features_names['GoPro'] if 'mean' in feature] 
        else:
            gopro_features = list()
        if 'FaceReader' in features_names.keys():
            facereader_features = [feature for feature in features_names['FaceReader'] if 'mean' in feature]
        else:
            facereader_features = list()

        if selected_features is None:
            selected_features = list()
            if 'Empatica' in features_names.keys():
                selected_features.extend(empatica_features)
            if 'Audio' in features_names.keys():
                selected_features.extend(audio_features)
            if 'Webcam' in features_names.keys():
                selected_features.extend(webcam_features)
            if 'GoPro' in features_names.keys():
                selected_features.extend(gopro_features)
            if 'FaceReader' in features_names.keys():
                selected_features.extend(facereader_features)

        for feature in empatica_features:
            if feature in selected_features:
                if not pair_comparison:
                    for empatica_feature in empatica_features:
                        if empatica_feature in selected_features and feature != empatica_feature:
                            if [feature, empatica_feature] not in features_to_cross_correlate and [empatica_feature, feature] not in features_to_cross_correlate:
                                features_to_cross_correlate.append([feature, empatica_feature])
                for audio_feature in audio_features:
                    if audio_feature in selected_features:
                        if [feature, audio_feature] not in features_to_cross_correlate and [audio_feature, feature] not in features_to_cross_correlate:
                            features_to_cross_correlate.append([feature, audio_feature])
                for webcam_feature in webcam_features:
                    if webcam_feature in selected_features:
                        if [feature, webcam_feature] not in features_to_cross_correlate and [webcam_feature, feature] not in features_to_cross_correlate:
                            features_to_cross_correlate.append([feature, webcam_feature])
                for gopro_feature in gopro_features:
                    if gopro_feature in selected_features:
                        if [feature, gopro_feature] not in features_to_cross_correlate and [gopro_feature, feature] not in features_to_cross_correlate:
                            features_to_cross_correlate.append([feature, gopro_feature])
                for facereader_feature in facereader_features:
                    if facereader_feature in selected_features:
                        if [feature, facereader_feature] not in features_to_cross_correlate and [facereader_feature, feature] not in features_to_cross_correlate:
                            features_to_cross_correlate.append([feature, facereader_feature])

        for feature in audio_features:
            if feature in selected_features:
                if not pair_comparison:
                    for audio_feature in audio_features:
                        if audio_feature in selected_features and feature != audio_feature:
                            if [feature, audio_feature] not in features_to_cross_correlate and [audio_feature, feature] not in features_to_cross_correlate:
                                features_to_cross_correlate.append([feature, audio_feature])
                for webcam_feature in webcam_features:
                    if webcam_feature in selected_features:
                        if [feature, webcam_feature] not in features_to_cross_correlate and [webcam_feature, feature] not in features_to_cross_correlate:
                            features_to_cross_correlate.append([feature, webcam_feature])
                for gopro_feature in gopro_features:
                    if gopro_feature in selected_features:
                        if [feature, gopro_feature] not in features_to_cross_correlate and [gopro_feature, feature] not in features_to_cross_correlate:
                            features_to_cross_correlate.append([feature, gopro_feature])
                for facereader_feature in facereader_features:
                    if facereader_feature in selected_features:
                        if [feature, facereader_feature] not in features_to_cross_correlate and [facereader_feature, feature] not in features_to_cross_correlate:
                            features_to_cross_correlate.append([feature, facereader_feature])

        for feature in webcam_features:
            if feature in selected_features:
                if not pair_comparison:
                    for webcam_feature in webcam_features:
                        if webcam_feature in selected_features and feature != webcam_feature:
                            if [feature, webcam_feature] not in features_to_cross_correlate and [webcam_feature, feature] not in features_to_cross_correlate:
                                features_to_cross_correlate.append([feature, webcam_feature])
                for gopro_feature in gopro_features:
                    if gopro_feature in selected_features:
                        if [feature, gopro_feature] not in features_to_cross_correlate and [gopro_feature, feature] not in features_to_cross_correlate:
                            features_to_cross_correlate.append([feature, gopro_feature])
                for facereader_feature in facereader_features:
                    if facereader_feature in selected_features:
                        if [feature, facereader_feature] not in features_to_cross_correlate and [facereader_feature, feature] not in features_to_cross_correlate:
                            features_to_cross_correlate.append([feature, facereader_feature])

        for feature in gopro_features:
            if feature in selected_features:
                if not pair_comparison:
                    for gopro_feature in gopro_features:
                        if gopro_feature in selected_features and feature != gopro_feature:
                            if [feature, gopro_feature] not in features_to_cross_correlate and [gopro_feature, feature] not in features_to_cross_correlate:
                                features_to_cross_correlate.append([feature, gopro_feature])
                for facereader_feature in facereader_features:
                    if facereader_feature in selected_features:
                        if [feature, facereader_feature] not in features_to_cross_correlate and [facereader_feature, feature] not in features_to_cross_correlate:
                            features_to_cross_correlate.append([feature, facereader_feature])

        for feature in facereader_features:
            if feature in selected_features:
                if not pair_comparison:
                    for facereader_feature in facereader_features:
                        if facereader_feature in selected_features and feature != facereader_feature:
                            if [feature, facereader_feature] not in features_to_cross_correlate and [facereader_feature, feature] not in features_to_cross_correlate:
                                features_to_cross_correlate.append([feature, facereader_feature])

        return features_to_cross_correlate
    
    def compare_max_lags(self, cross_correlations, feature_names, group, intra_comparison = True, save=False):
        """
        Compare the max lags for each emotion and each feature pair

        Parameters
        ----------
        cross_correlations : dict
            Dictionary of cross-correlation vectors for each emotion
        feature_names : dict
            Dictionary with the feature names for each emotion measurement
        intra_comparison : bool
            Whether to compare the max lags for each feature pair within one measurement type or not.
        save : bool
            Whether to save the max lags or not
        
        Returns
        -------
        lags_dict : dict
            Dictionary with the max lags for each emotion and each feature pair
        """

        lags_dict = dict()
        for emotion in cross_correlations.keys():
            lags_dict[emotion] = dict()
            for subject_id in cross_correlations[emotion].keys():
                for feature_pair in cross_correlations[emotion][subject_id].keys():
                    if len(feature_names.keys())>1 and intra_comparison == False:
                        same_measurement = False
                        for measurment in feature_names.keys():
                            if feature_pair.split('--')[0] in feature_names[measurment] and feature_pair.split('--')[1] in feature_names[measurment]:
                                same_measurement = True
                        if not same_measurement:
                            if feature_pair not in lags_dict[emotion].keys():
                                lags_dict[emotion][feature_pair] = list()
                            lags_dict[emotion][feature_pair].append(int(cross_correlations[emotion][subject_id][feature_pair]['max_lag']))
                    else:
                        if feature_pair not in lags_dict[emotion].keys():
                            lags_dict[emotion][feature_pair] = list()
                        lags_dict[emotion][feature_pair].append(int(cross_correlations[emotion][subject_id][feature_pair]['max_lag']))

        if save:

            desired_measurements = list(feature_names.keys())
            if len(desired_measurements)>1:
                folder_name = ''
                for measurement in desired_measurements:
                    if measurement == desired_measurements[-1]:
                        folder_name += measurement
                    else:
                        folder_name += measurement + '-'
            else:
                folder_name = desired_measurements[0]

            os.makedirs(f'Analysis/Cross-Correlation/{group}/{folder_name}', exist_ok=True)

            class SingleLineArrayEncoder(json.JSONEncoder):
                def iterencode(self, o, _one_shot=False):
                    list_lvl = 0
                    for s in super(SingleLineArrayEncoder, self).iterencode(o, _one_shot=_one_shot):
                        if s.startswith('['):
                            list_lvl += 1
                        elif s.startswith(']'):
                            list_lvl -= 1

                        if list_lvl > 0:
                            s = s.replace('\n', '').replace(' ', '')
                        yield s

            with open(f"Analysis/Cross-Correlation/{group}/{folder_name}/Lags.json", 'w') as fp:
                json.dump(lags_dict, fp, cls=SingleLineArrayEncoder, indent=4)

        return lags_dict

    def get_lag_stats(self, lags_dict, feature_names, group, p_value = 0.05, method= 'wilcoxon', verbose = False, save = False):
        """
        Get the pairs of features with statistically significant lags across subjects.

        Parameters
        ----------
        lags_dict : dict
            Dictionary with the max lags for each emotion and each feature pair
        feature_names : dict
            Dictionary with the feature names for each emotion measurement
        group : str
            The group to consider
        p_value : float
            The p-value to use for the t-test
        method: str
            The method to use for the t-test. Either 'ttest' or 'wilcoxon'
        verbose : bool
            Whether to print the results or not
        save : bool
            Whether to save the results or not
        
        Returns
        -------
        shifted_feature_pairs : dict
            Dictionary with the feature pairs with statistically significant lags
        centered_feature_pairs : dict
            Dictionary with the feature pairs with lags not statistically significant
        """
        centered_feature_pairs = dict()
        shifted_feature_pairs = dict()

        for emotion in lags_dict.keys():

            centered_feature_pairs[emotion] = dict()
            shifted_feature_pairs[emotion] = dict()

            for feature_pair, lags in lags_dict[emotion].items():
                
                lags = np.array(lags)
                
                if method == 'ttest':
                    t_stat, p_val = ttest_1samp(lags, 0)
                elif method == 'wilcoxon':
                    t_stat, p_val = wilcoxon(lags, zero_method= 'zsplit', alternative='two-sided')
                else:
                    raise ValueError('The method is not valid. Please use either "ttest" or "wilcoxon"')

                mean_lag = np.mean(lags)
                std_lag = np.std(lags)
                
                if p_val < p_value:
                    shifted_feature_pairs[emotion][feature_pair] = {
                        "Mean": mean_lag, 
                        "Std": std_lag
                    }
                    if verbose:
                        print("The lag is statistically significantly different from zero.")
                        print(f'Feature Pair: {feature_pair}')
                        print(f'Mean Lag: {mean_lag}')
                        print(f'Standard Deviation of Lag: {std_lag}')
                        print(f'T-statistic: {t_stat}, P-value: {p_val}')
                        print("\n" + "="*50 + "\n")
                else:
                    centered_feature_pairs[emotion][feature_pair] = {
                        "Mean": mean_lag, 
                        "Std": std_lag
                    }
                    if verbose:
                        print("The lag is not statistically significantly different from zero.")
                        print(f'Feature Pair: {feature_pair}')
                        print(f'Mean Lag: {mean_lag}')
                        print(f'Standard Deviation of Lag: {std_lag}')
                        print(f'T-statistic: {t_stat}, P-value: {p_val}')
                        print("\n" + "="*50 + "\n")

        consistent_shifted_pairs = self.get_consistent_shift_pairs(shifted_feature_pairs, verbose = False)

        if save:

            desired_measurements = list(feature_names.keys())
            if len(desired_measurements)>1:
                folder_name = ''
                for measurement in desired_measurements:
                    if measurement == desired_measurements[-1]:
                        folder_name += measurement
                    else:
                        folder_name += measurement + '-'
            else:
                folder_name = desired_measurements[0]

            os.makedirs(f'Analysis/Cross-Correlation/{group}/{folder_name}', exist_ok=True)

            with open(f'Analysis/Cross-Correlation/{group}/{folder_name}/centered_feature_pairs.json', 'w') as f:
                json.dump(centered_feature_pairs, f, indent=4)
            with open(f'Analysis/Cross-Correlation/{group}/{folder_name}/shifted_feature_pairs.json', 'w') as f:
                json.dump(shifted_feature_pairs, f, indent=4)
            with open(f'Analysis/Cross-Correlation/{group}/{folder_name}/consistent_shifted_pairs.json', 'w') as f:
                json.dump(consistent_shifted_pairs, f, indent=4)

        return shifted_feature_pairs, centered_feature_pairs
    
    def get_shifts_stats(self, feature_pairs, feature_names, group, save = False):
        """
        This function takes the feature_pairs dictionary and returns three dictionaries with the positive, negative and no shifts

        Parameters:
        ----------------
        feature_pairs: dict
            Dictionary with the feature pairs and their statistics
        feature_names: dict
            Dictionary with the feature names for each emotion measurement
        group: str
            The group to consider
        save: bool
            Whether to save the results or not
        
        Returns:
        ----------------
        positive_shifts: dict
            Dictionary with the positive shifts
        negative_shifts: dict
            Dictionary with the negative shifts
        no_shifts: dict
            Dictionary with the no shifts

        """

        positive_shifts = dict()
        negative_shifts = dict()
        no_shifts = dict()

        for emotion in feature_pairs.keys():
            positive_shifts[emotion] = dict()
            negative_shifts[emotion] = dict()
            no_shifts[emotion] = dict()
            for pair in feature_pairs[emotion].keys():
                if 'angle' in pair:
                    if pair[-5:] == 'angle':
                        feature1 = pair.split('--')[0]
                        feature2 = pair.replace(feature1 +'-', '')
                        feature2 += '_angle'
                    else:
                        feature1 = pair.split("_angle--")[0]
                        feature1+= '_angle'
                        feature2 = pair.split("_angle--")[1]
                else:
                    feature1 = pair.split("--")[0]
                    feature2 = pair.split("--")[1]
                if feature_pairs[emotion][pair]['Mean'] > 0:
                    if feature1 not in positive_shifts[emotion].keys():
                        positive_shifts[emotion][feature1] = dict()
                    positive_shifts[emotion][feature1][feature2] = str(np.round(feature_pairs[emotion][pair]['Mean']*10,2)) + "s"
                elif feature_pairs[emotion][pair]['Mean'] < 0:
                    if feature1 not in negative_shifts[emotion].keys():
                        negative_shifts[emotion][feature1] = dict()
                    negative_shifts[emotion][feature1][feature2] = str(np.round(feature_pairs[emotion][pair]['Mean']*10,2)) + "s"
                elif feature_pairs[emotion][pair]['Mean'] == 0:
                    if feature1 not in no_shifts.keys():
                        no_shifts[emotion][feature1] = dict()
                    no_shifts[emotion][feature1][feature2] = str(np.round(feature_pairs[emotion][pair]['Mean']*10,2)) + "s"

        if save:

            desired_measurements = list(feature_names.keys())
            if len(desired_measurements)>1:
                folder_name = ''
                for measurement in desired_measurements:
                    if measurement == desired_measurements[-1]:
                        folder_name += measurement
                    else:
                        folder_name += measurement + '-'
            else:
                folder_name = desired_measurements[0]

            os.makedirs(f'Analysis/Cross-Correlation/{group}/{folder_name}', exist_ok=True)
                
            with open(f'Analysis/Cross-Correlation/{group}/{folder_name}/positive_shifts.json', 'w') as f:
                json.dump(positive_shifts, f, indent=4)
            with open(f'Analysis/Cross-Correlation/{group}/{folder_name}/negative_shifts.json', 'w') as f:
                json.dump(negative_shifts, f, indent=4)
            with open(f'Analysis/Cross-Correlation/{group}/{folder_name}/no_shifts.json', 'w') as f:
                json.dump(no_shifts, f, indent=4)

        return positive_shifts, negative_shifts, no_shifts
    
    def get_consistent_shift_pairs(self, shifted_feature_pairs, verbose = False):
        """
        Extract the pairs that are consistent with positive and negative valence emotions

        Parameters:
        ----------------
        shifted_feature_pairs: dict
            Dictionary with the feature pairs and their statistics
        verbose: bool
            Whether to print the results or not

        Returns:
        ----------------
        consistent_pairs_positive_valence: list
            List with the pairs that are consistent with positive valence emotions
        consistent_pairs_negative_valence: list
            List with the pairs that are consistent with negative valence emotions
        """
        positive_valence_consistent = {}
        negative_valence_consistent = {}
        for emotion in shifted_feature_pairs.keys():
            for pair in shifted_feature_pairs[emotion].keys():
                pair = pair.replace('accX', 'acc')
                pair = pair.replace('accY', 'acc')
                pair = pair.replace('accZ', 'acc')
                if emotion == 'Pride' or emotion == 'Joy':
                    if pair not in positive_valence_consistent.keys():
                        positive_valence_consistent[pair] = 1
                    else:
                        positive_valence_consistent[pair] += 1
                elif emotion == 'Shame' or emotion == 'Frustration':
                    if pair not in negative_valence_consistent.keys():
                        negative_valence_consistent[pair] = 1
                    else:
                        negative_valence_consistent[pair] += 1

        consistent_pairs_positive_valence = [pair for pair in positive_valence_consistent.keys() if positive_valence_consistent[pair] ==2]
        consistent_pairs_negative_valence = [pair for pair in negative_valence_consistent.keys() if negative_valence_consistent[pair] ==2]

        if verbose:
            print('Pairs that are consistent with positive valence emotions:', consistent_pairs_positive_valence)
            print('Pairs that are consistent with negative valence emotions:', consistent_pairs_negative_valence)

        return consistent_pairs_positive_valence, consistent_pairs_negative_valence
        
    def compare_shift_across_emotions(self, shifts, group, save = False, filename=None):
        """
        Compare the shifts across emotions to see if there are any common shifts between them.

        Parameters
        ----------
        shifts : dict
            Dictionary containing the shifts for each emotion.
        
        Returns
        -------
        shifts_across_emotions : dict
            Dictionary containing the shifts across emotions.

        """        

        emotions_to_compare = ['Shame', 'Pride', 'Joy', 'Frustration']
        pairs = list(combinations(emotions_to_compare, 2))
        triads = list(combinations(emotions_to_compare, 3))

        shifts_across_emotions = dict()

        for pair in pairs:
            shifts_across_emotions[pair[0]+'-'+pair[1]] = dict()
            for key1 in shifts[pair[0]].keys():
                for key2 in shifts[pair[1]].keys():
                    if key1 == key2:
                        for value in shifts[pair[0]][key1]:
                            if value in shifts[pair[1]][key2]:
                                if key1 not in shifts_across_emotions[pair[0]+'-'+pair[1]].keys():
                                    shifts_across_emotions[pair[0]+'-'+pair[1]][key1] = [value]
                                else:
                                    shifts_across_emotions[pair[0]+'-'+pair[1]][key1].append(value)

        for triad in triads:
            shifts_across_emotions[triad[0]+'-'+triad[1]+'-'+triad[2]] = dict()
            for key1 in shifts[triad[2]].keys():
                for key2 in shifts_across_emotions[triad[0]+'-'+triad[1]].keys():
                    if key1 == key2:
                        for value in shifts[triad[2]][key1]:
                            if value in shifts_across_emotions[triad[0]+'-'+triad[1]][key2]:
                                if key1 not in shifts_across_emotions[triad[0]+'-'+triad[1]+'-'+triad[2]].keys():
                                    shifts_across_emotions[triad[0]+'-'+triad[1]+'-'+triad[2]][key1] = [value]
                                else:
                                    shifts_across_emotions[triad[0]+'-'+triad[1]+'-'+triad[2]][key1].append(value)

        for emotion in emotions_to_compare:
            if emotion not in triads[-1]:
                shifts_across_emotions[triad[0]+'-'+triad[1]+'-'+triad[2]+'-'+emotion] = dict()
                for key1 in shifts[emotion].keys():
                    for key2 in shifts_across_emotions[triad[0]+'-'+triad[1]+'-'+triad[2]].keys():
                        if key1 == key2:
                            for value in shifts[emotion][key1]:
                                if value in shifts_across_emotions[triad[0]+'-'+triad[1]+'-'+triad[2]][key2]:
                                    if key1 not in shifts_across_emotions[triad[0]+'-'+triad[1]+'-'+triad[2]+'-'+emotion].keys():
                                        shifts_across_emotions[triad[0]+'-'+triad[1]+'-'+triad[2]+'-'+emotion][key1] = [value]
                                    else:
                                        shifts_across_emotions[triad[0]+'-'+triad[1]+'-'+triad[2]+'-'+emotion][key1].append(value)

        if save:

            os.makedirs(f'Analysis/Cross-Correlation/{group}', exist_ok=True)

            with open(f'Analysis/Cross-Correlation/{group}/{filename}.json', 'w') as f:
                json.dump(shifts_across_emotions, f, indent=4)

        return shifts_across_emotions
    
    def plot_shift_stats(self, positive_shifts, negative_shifts, no_shifts, desired_measurement, group, save_path = False):
        """
        Plots the shifts between features for a given emotion.

        Parameters:
        -----------
        positive_shifts : dict
            Dictionary containing the positive shifts for each emotion.
        negative_shifts : dict
            Dictionary containing the negative shifts for each emotion.
        no_shifts : dict
            Dictionary containing the no shifts for each emotion.
        emotion : str
            Emotion for which to plot the shifts.
        desired_measurement : list
            List containing the desired measurement for the features.
        save_path : bool
            Whether to save the plot or not.
        """

        if type(desired_measurement) is tuple or type(desired_measurement) is list:
            path = f'Analysis/Cross-Correlation/{group}/{desired_measurement[0]}-{desired_measurement[1]}'
            meas = f'{desired_measurement[0]}-{desired_measurement[1]}'
            os.makedirs(path, exist_ok=True)
        else:
            path = f'Analysis/Cross-Correlation/{group}/{desired_measurement}'
            meas = desired_measurement
            os.makedirs(path, exist_ok=True)

        emotions = ['Shame', 'Frustration', 'Pride', 'Joy']

        fig, axs = plt.subplots(2, 2, figsize=(16, 12))  
        axs = axs.flatten() 

        global_max = float('-inf')
        global_min = float('inf')

        for emotion in emotions:
            for shift_dict in [positive_shifts[emotion], negative_shifts[emotion], no_shifts[emotion]]:
                for feature1 in shift_dict.keys():
                    for _, value in shift_dict[feature1].items():
                        value_float = float(value.replace('s', ''))
                        global_max = max(global_max, value_float)
                        global_min = min(global_min, value_float) 

        for idx, emotion in enumerate(emotions):
            labels_values = list()
            for feature1 in positive_shifts[emotion].keys():
                for feature2, value in positive_shifts[emotion][feature1].items():
                    label = f"{feature1.replace('_mean', '')}--{feature2.replace('_mean', '')}"
                    value_float = float(value.replace('s', ''))
                    labels_values.append((label, value_float))
            for feature1 in negative_shifts[emotion].keys():
                for feature2, value in negative_shifts[emotion][feature1].items():
                    label = f"{feature1.replace('_mean', '')}--{feature2.replace('_mean', '')}"
                    value_float = float(value.replace('s', ''))
                    labels_values.append((label, value_float))
            for feature1 in no_shifts[emotion].keys():
                for feature2, value in no_shifts[emotion][feature1].items():
                    label = f"{feature1.replace('_mean', '')}--{feature2.replace('_mean', '')}"
                    value_float = float(value.replace('s', ''))
                    labels_values.append((label, value_float))
                    
            labels_values.sort(key=lambda x: x[1])
    
            if labels_values:
                labels, values = zip(*labels_values)
                ax = axs[idx]  
                
                ax.bar(labels, values, color='#DC143C')
                ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=10)
                ax.tick_params(axis='y', labelsize=10)
                ax.set_ylabel('Lag (s)', fontsize=12)
                ax.set_ylim([global_min-5, global_max+5])
                
                if isinstance(desired_measurement, (list, tuple)):
                    measurements = " and ".join(desired_measurement)
                    plot_title = f'{emotion} - Shifts between {measurements} features'
                else:
                    plot_title = f'{emotion} - Shifts between {desired_measurement} features'
                ax.set_title(plot_title, fontsize=12)
                ax.grid(axis='y', linestyle='--', alpha=0.7)
            else:
                axs[idx].plot([], [])
                axs[idx].set_title(emotion)

        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        fig.suptitle(f'{meas} - Shifts between features across emotions', fontsize=16, fontweight='bold')

        if save_path:
            png_filename = "Shifts_Across_Emotions.png"
            png_full_path = os.path.join(path, png_filename)
            plt.savefig(png_full_path, format='png', dpi=300)
            plt.close()
        else:
            plt.show()
    
    def plot_cross_correlation(self, cross_correlations, subject_id, emotion, features_to_plot, pdf=None):
        """
        Plot the cross-correlation between two features for a given emotion

        Parameters
        ----------
        cross_correlations : dict
            Dictionary of cross-correlation vectors for each emotion
        subject_id : int
            Subject ID
        emotion : str
            Emotion to plot
        features_to_plot : str
            Features to plot in the format "feature1-feature2"
        pdf : PdfPages object, optional
            PdfPages object to save the plot to a pdf file
        """

        cross_correlation = cross_correlations[emotion][subject_id][features_to_plot]['cross_correlation']
        lags = cross_correlations[emotion][subject_id][features_to_plot]['lags']
        max_corr = cross_correlations[emotion][subject_id][features_to_plot]['max_cross_correlation']
        max_lag = cross_correlations[emotion][subject_id][features_to_plot]['max_lag']
        features = features_to_plot.split('--')

        plt.figure(figsize=(10, 5)) 
        plt.plot(lags, cross_correlation, label='Cross-Correlation')
        plt.axvline(max_lag, color='r', linestyle='--', label=f'Max Correlation at lag={max_lag}')
        plt.axhline(max_corr, color='g', linestyle='--', label=f'Peak Correlation={max_corr:.2f}')
        plt.title(f'Cross-Correlation between {features[0]} and {features[1]} for {emotion} Emotion (Subject {subject_id})') 
        plt.xlabel('Lag')
        plt.ylabel('Cross-Correlation')
        plt.grid(True)
        plt.legend()

        if pdf is not None:
            pdf.savefig()
            plt.close()
        else:
            plt.show()

    def save_cross_corr_graphs_to_pdf(self, cross_correlations, feature_names, group):
        """
        Save the raw empatica data to a PDF file.

        Parameters
        ----------
        cross_correlations : dict
            Dictionary of cross-correlation vectors for each emotion
        feature_names : dict
            Dictionary of feature names for each emotion
        group : str
            Group name
        """

        desired_measurements = list(feature_names.keys())
        if len(desired_measurements)>1:
            folder_name = ''
            for measurement in desired_measurements:
                if measurement == desired_measurements[-1]:
                    folder_name += measurement
                else:
                    folder_name += measurement + '-'
        else:
            folder_name = desired_measurements[0]

        os.makedirs(f'Analysis/Cross-Correlation/{group}/{folder_name}', exist_ok=True)

        filename = f'Analysis/Cross-Correlation/{group}/{folder_name}/cross_correlations.pdf'

        for emotion in cross_correlations.keys():
            name = filename.replace('.pdf', f'_{emotion}.pdf')
            with PdfPages(name) as pdf:
                for subject_id in cross_correlations[emotion].keys():
                    plt.figure(figsize=(11, 8.5))  
                    plt.text(0.5, 0.5, f'Subject {subject_id}', ha='center', va='center', size=24)
                    plt.axis('off')
                    pdf.savefig() 
                    plt.close() 
                    for features in cross_correlations[emotion][subject_id].keys():
                        self.plot_cross_correlation(cross_correlations, subject_id, emotion, features, pdf)

    def save_all_shift_stat_graphs(self, desired_measurement, subject_groups):
        """
        Plots the shifts between features for each emotion.

        Parameters:
        -----------
        desired_measurement : list
            List containing the desired measurement for the features.
        subject_groups : dict
            dict of the desired subjects to use
        """

        pairs = list(combinations(desired_measurement, 2))
        desired_measurement.extend(pairs)
        
        for measurement in desired_measurement:
            print(f'Computing Shift Stats graphs for {measurement}...')
            if type(measurement) is str:
                if os.path.exists(f'../computed_features/{measurement}/stand_features_windows.csv'):
                    stand_features = pd.read_csv(f'../computed_features/{measurement}/stand_features_windows.csv')
                else:
                    stand_features = pd.read_csv(f'computed_features/{measurement}/stand_features_windows.csv')
                features_names = self.utilities.get_feature_names(stand_features, [measurement])
            else:
                if os.path.exists('../computed_features/stand_features_windows.csv'):
                    stand_features = pd.read_csv(f'../computed_features/stand_features_windows.csv')
                else:
                    stand_features = pd.read_csv('computed_features/stand_features_windows.csv')
                features_names = self.utilities.get_feature_names(stand_features, measurement)

            for group in subject_groups.keys():
                features_to_use = self.utilities.select_features_from_subjects(stand_features, subject_groups[group])

                features_to_cross_correlate = self.get_features_to_cross_correlate(features_names)

                features_grouping = self.utilities.group_features_by_label(features_to_use)

                cross_correlations = self.get_cross_correlations(features_grouping, features_to_cross_correlate)
                lags_dict = self.compare_max_lags(cross_correlations, features_names, group, intra_comparison = False, save = False)
                shifted_feature_pairs, centered_feature_pairs = self.get_lag_stats(lags_dict, features_names, group, p_value = 0.05, verbose = False, save = False) 
                positive_shifts, negative_shifts, no_shifts = self.get_shifts_stats(shifted_feature_pairs, features_names, group, save = False)
                self.plot_shift_stats(positive_shifts, negative_shifts, no_shifts, measurement, group, save_path = True)

    def save_all_cross_correlation_results(self, desired_measurement, subject_groups, include_pairs=False, only_pairs = False, save_cross_corr_graphs=False):
        """
        Save the results of the cross-correlation analysis for all the measurements and all the window sizes.

        Parameters
        ----------
        desired_measurement : list
            List of the desired measurements to use
        subject_groups : dict
            dict of the desired subjects to use
        include_pairs : bool
            Whether to include the pairs of measurements in the cross-correlation analysis or not
        only_pairs : bool
            Whether to only include the pairs of measurements in the cross-correlation analysis or not
        save_cross_corr_graphs : bool
            Whether to save the cross-correlation graphs to a PDF file or not
        """

        _desired_measurement = desired_measurement.copy()
        
        if include_pairs:
            unique_pairs = list(combinations(_desired_measurement,2))
            _desired_measurement.extend(unique_pairs)

        pbar = tqdm(total=len(_desired_measurement)*len(subject_groups), desc="Extracting Cross-Correlation Results")

        for measurement in _desired_measurement:

            if type(measurement) is str:
                if only_pairs:
                    continue
                pbar.set_description(f'Computing Cross-Correlation for {measurement}...')
                if os.path.exists(f'../computed_features/{measurement}/stand_features_windows.csv'):
                    stand_features = pd.read_csv(f'../computed_features/{measurement}/stand_features_windows.csv')
                else:
                    stand_features = pd.read_csv(f'computed_features/{measurement}/stand_features_windows.csv')
                features_names = self.utilities.get_feature_names(stand_features, [measurement])
                features_to_cross_correlate = self.get_features_to_cross_correlate(features_names, pair_comparison=False)
            else:
                pbar.set_description(f'Computing Cross-Correlation for {measurement[0]}-{measurement[1]}...')
                if os.path.exists('../computed_features/stand_features_windows.csv'):
                    stand_features = pd.read_csv('../computed_features/stand_features_windows.csv')
                else:
                    stand_features = pd.read_csv('computed_features/stand_features_windows.csv')
                features_names = self.utilities.get_feature_names(stand_features, measurement)
                features_to_cross_correlate = self.get_features_to_cross_correlate(features_names, pair_comparison=True)

            for group in subject_groups:
                features_to_use = self.utilities.select_features_from_subjects(stand_features, subject_groups[group])

                features_grouping = self.utilities.group_features_by_label(features_to_use, windows=True)

                cross_correlations = self.get_cross_correlations(features_grouping, features_to_cross_correlate)
                lags_dict = self.compare_max_lags(cross_correlations, features_names, group, intra_comparison = False, save = True)
                shifted_feature_pairs, centered_feature_pairs = self.get_lag_stats(lags_dict, features_names, group, p_value = 0.05, verbose = False, save = True) 
                positive_shifts, negative_shifts, no_shifts = self.get_shifts_stats(shifted_feature_pairs, features_names, group, save = True)
                if save_cross_corr_graphs:
                    self.save_cross_corr_graphs_to_pdf(cross_correlations, features_names, group)
                self.plot_shift_stats(positive_shifts, negative_shifts, no_shifts, measurement, group, save_path=True)

                pbar.update(1)
        
        pbar.close()

class CLASSIFIER:
    def __init__(self):

        self.utilities = UTILITIES()
        pass

    def get_classification_sets(self, features, class_obj, group, augment_data=False, feature_selection=False, selected_features = None, nb_features = 30, reduce_dim=False, reduce_dim_method='PCA', include_neutral=True):
        """
        Get the sets for the comparison
        
        Parameters
        ----------
        features : pandas.DataFrame
            The dataframe containing the features
        class_obj : str
            The classification objective to consider
        group : list
            The group of subjects to consider
        augment_data : bool
            Whether to augment the data or not
        feature_selection : bool
            Whether to perform feature selection or not
        selected_features : list
            The list of desired features to use
        reduce_dim : bool
            Whether to reduce the dimensionality of the data or not
        reduce_dim_method : str
            The method to use for dimensionality reduction
        include_neutral : bool
            Whether to include the neutral class or not
            
        Returns
        -------
        X_train : pandas.DataFrame
            The training set
        X_test : pandas.DataFrame
            The test set
        y_train : pandas.DataFrame
            The training labels
        y_test : pandas.DataFrame
            The test labels
        """

        features_to_use = self.utilities.select_features_from_subjects(features, group, include_neutral=include_neutral)

        if class_obj == 'Neutral vs Non-Neutral':
            X = features_to_use.drop('label', axis=1)
            y = features_to_use['label']
            y = y.replace({'Frustration': 'Non-Neutral', 'Pride': 'Non-Neutral', 'Joy': 'Non-Neutral', 'Shame': 'Non-Neutral'})
        elif class_obj == 'Positive vs Negative':

            X = features_to_use.drop('label', axis=1)
            y = features_to_use['label']
            y = y.replace({'Frustration': 'Negative', 'Shame': 'Negative', 'Pride': 'Positive', 'Joy': 'Positive'})
        elif class_obj == 'Shame vs Others':
            X = features_to_use.drop('label', axis=1)
            y = features_to_use['label']
            y = y.replace({'Frustration': 'Other', 'Pride': 'Other', 'Joy': 'Other'})
        elif class_obj == 'All Emotions':
            X = features_to_use.drop('label', axis=1)
            y = features_to_use['label']

        label_encoder = LabelEncoder()
        y = label_encoder.fit_transform(y)

        if augment_data:
            smote = SMOTE(random_state=42)
            X, y = smote.fit_resample(X, y)

        y = to_categorical(y)
        label_names = list(label_encoder.classes_)

        if feature_selection:
            if selected_features is None:
                # selector = RFECV(RandomForestClassifier(n_estimators=100, random_state=42), step=1, cv=2)
                # selector.fit(X, y)
                # selected_features = X[X.columns[selector.support_]]
                # X = selector.transform(X)
                
                model = RandomForestClassifier(n_estimators=100, random_state=42)
                selector = RFE(estimator=model, n_features_to_select=nb_features)
                if nb_features > len(X.columns):
                    nb_features = len(X.columns)
                    print("The number of features to select is greater than the number of features available. The number of features to select is set to {}".format(nb_features))
                selector.fit(X, y)
                selected_features = X.columns[selector.support_]
                X = X[selected_features]
            else:
                X = X[selected_features]

        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.25, random_state=42)

        if reduce_dim:
            if reduce_dim_method == 'PCA':
                pca = PCA(random_state=42) 
                X_train = pca.fit_transform(X_train)
                X_test = pca.transform(X_test)
            elif reduce_dim_method == 'Sammon':
                mds = MDS(random_state=42)
                X_train = mds.fit_transform(X_train)

        return X_train, y_train, X_test, y_test, selected_features, label_names
    

    def get_classification_performances(self, stand_features, groups, augment_data=False, feature_selection=False, nb_features=15, reduce_dim=False, verbose=False, save=False):
        """
        Get the classification performances for the different groups and objectives.

        Parameters
        ----------
        stand_features : pd.DataFrame
            The standardized features
        groups : dict
            The dictionary containing the groups of subjects
        augment_data : bool
            Whether to augment the data or not
        feature_selection : bool
            Whether to perform feature selection or not
        nb_features : int
            The number of features to select
        reduce_dim : bool
            Whether to reduce the dimensionality of the data or not
        verbose : bool
            Whether to print the results or not
        save : bool
            Whether to save the results or not

        Returns
        -------
        classifier_performance : dict
            Dictionary containing the classification performances
        """
        classifier_performance = dict()
        pbar = tqdm(total=len(groups)*4, desc="Computing Classification Performances")
        for group in groups.keys():
            if verbose:
                print('Group: ', group)
            classifier_performance[group] = dict()
            for objective in ['Neutral vs Non-Neutral', 'Positive vs Negative', 'Shame vs Others', 'All Emotions']:
                pbar.set_description(f'Computing Classification Performances for {group} - {objective}')
                if verbose:
                    print('Objective: ', objective)
                classifier_performance[group][objective] = dict()

                X_train, y_train, X_test, y_test, selected_features, label_names = self.get_classification_sets(stand_features, objective, groups[group], augment_data=augment_data, feature_selection=feature_selection, nb_features=nb_features, reduce_dim=reduce_dim)
                classifier_performance[group][objective]['Selected Features'] = selected_features.tolist()

                colors = ["grey", "crimson"] 
                n_bins = 100  
                cmap_name = "my_custom_cmap"
                cm = LinearSegmentedColormap.from_list(cmap_name, colors, N=n_bins)

                best_rf_params = {'n_estimators': 50, 'max_depth': 15, 'max_features': 'sqrt', 'min_samples_split': 3, 'criterion': 'gini'}
                rf_model = RandomForestClassifier(**best_rf_params, random_state=42)
                rf_model.fit(X_train, np.argmax(y_train, axis=1))
                rf_predictions = rf_model.predict(X_test)
                rf_accuracy = accuracy_score(np.argmax(y_test, axis=1), rf_predictions)
                classifier_performance[group][objective]['Random Forest'] = str(round(rf_accuracy*100,1))+'%'
                pbar.set_description(f'Computing Classification Performances for {group} - {objective}')
                if verbose:
                    print('Random Forest: ', round(rf_accuracy*100, 1), '%')
                cm_rf = confusion_matrix(np.argmax(y_test, axis=1), rf_predictions)
                plt.figure(figsize=(10, 7))
                sns.heatmap(cm_rf, annot=True, fmt='g', cmap=cm, xticklabels=label_names, yticklabels=label_names, annot_kws={"size": 16})

                plt.title(f'Confusion Matrix for Random Forest')
                plt.xlabel('Predicted Labels')
                plt.ylabel('True Labels')
                if save:
                    os.makedirs(f'Analysis/Classification/{group}/{objective}', exist_ok=True)
                    plt.savefig(f'Analysis/Classification/{group}/{objective}/Confusion_Matrix_RF.png')
                else:
                    plt.show()
                plt.close()


                best_svm_params = {'C': 1000, 'kernel': 'rbf', 'gamma': 'scale'}
                svm_model = SVC(**best_svm_params, probability=True, random_state=42)
                svm_model.fit(X_train, np.argmax(y_train, axis=1))
                svm_predictions = svm_model.predict(X_test)
                svm_accuracy = accuracy_score(np.argmax(y_test, axis=1), svm_predictions)
                classifier_performance[group][objective]['SVM'] = str(round(svm_accuracy*100,1))+'%'
                if verbose:
                    print('SVM: ', round(svm_accuracy*100,1), '%')
                cm_svm = confusion_matrix(np.argmax(y_test, axis=1), svm_predictions)
                plt.figure(figsize=(10, 7))
                sns.heatmap(cm_svm, annot=True, fmt='g', cmap=cm, xticklabels=label_names, yticklabels=label_names, annot_kws={"size": 16})
                plt.title(f'Confusion Matrix for SVM')
                plt.xlabel('Predicted Labels')
                plt.ylabel('True Labels')
                if save:
                    os.makedirs(f'Analysis/Classification/{group}/{objective}', exist_ok=True)
                    plt.savefig(f'Analysis/Classification/{group}/{objective}/Confusion_Matrix_SVM.png')
                else:
                    plt.show()
                plt.close()

                xgb_model = XGBClassifier(
                    n_estimators=5,
                    learning_rate=0.1,
                    max_depth=1,
                    min_child_weight=1,
                    gamma=0,
                    subsample=0.8,
                    colsample_bytree=0.8,
                    objective='binary:logistic',  # or 'multi:softprob' for multiclass and set num_class
                    reg_alpha=0.005,
                    random_state=42
                )
                xgb_model.fit(X_train, np.argmax(y_train, axis=1))
                xgb_predictions = xgb_model.predict(X_test)
                xgb_accuracy = accuracy_score(np.argmax(y_test, axis=1), xgb_predictions)
                classifier_performance[group][objective]['XGBoost'] = str(round(xgb_accuracy*100,1))+'%'
                if verbose:
                    print('XGBoost: ', round(xgb_accuracy*100,1), '%')
                cm_xgb = confusion_matrix(np.argmax(y_test, axis=1), xgb_predictions)
                plt.figure(figsize=(10, 7))
                sns.heatmap(cm_xgb, annot=True, fmt='g', cmap=cm, xticklabels=label_names, yticklabels=label_names, annot_kws={"size": 16})
                plt.title(f'Confusion Matrix for XGB')
                plt.xlabel('Predicted Labels')
                plt.ylabel('True Labels')
                if save:
                    os.makedirs(f'Analysis/Classification/{group}/{objective}', exist_ok=True)
                    plt.savefig(f'Analysis/Classification/{group}/{objective}/Confusion_Matrix_XGB.png')
                else:
                    plt.show()
                plt.close()

                ada_model = AdaBoostClassifier(
                    estimator=DecisionTreeClassifier(max_depth=6),
                    n_estimators=50,
                    learning_rate=1,
                    random_state=42
                )
                ada_model.fit(X_train, np.argmax(y_train, axis=1))
                ada_predictions = ada_model.predict(X_test)
                ada_accuracy = accuracy_score(np.argmax(y_test, axis=1), ada_predictions)
                classifier_performance[group][objective]['AdaBoost'] = str(round(ada_accuracy*100,1))+'%'
                if verbose:
                    print('AdaBoost: ', round(ada_accuracy*100,1), '%')
                cm_ada = confusion_matrix(np.argmax(y_test, axis=1), ada_predictions)
                plt.figure(figsize=(10, 7))
                sns.heatmap(cm_ada, annot=True, fmt='g', cmap=cm, xticklabels=label_names, yticklabels=label_names, annot_kws={"size": 16})
                plt.title(f'Confusion Matrix for Ada Boost')
                plt.xlabel('Predicted Labels')
                plt.ylabel('True Labels')
                if save:
                    os.makedirs(f'Analysis/Classification/{group}/{objective}', exist_ok=True)
                    plt.savefig(f'Analysis/Classification/{group}/{objective}/Confusion_Matrix_ADA.png')
                else:
                    plt.show()
                plt.close()

                pbar.update(1)
        
        pbar.close()
        if save:
            os.makedirs('Analysis/Classification', exist_ok=True)
            with open('Analysis/Classification/classifier_performance.json', 'w') as file:
                json.dump(classifier_performance, file, indent=4)
        
        return classifier_performance
        
class UTILITIES:

    def __init__(self):
        """
        Initialize the utilities class.
        """

        self.scaler = StandardScaler()
        self.subject_mapping =  self._get_subject_mapping()

        return
    
    def group_features_by_label(self, features, windows=False, clean=True, replace_nan=True, affect_subject = False):
        """
        Group features by label and clean them if clean=True

        Parameters
        ----------
        features : pd.DataFrame
            Dataframe with all the features
        windows : bool, optional
            If True, the features divided into discrete time windows.
        clean : bool, optional
            If True, clean the features, by default True
        replace_nan : bool, optional
            If True, replace the nan values by the mean of non NaN values, by default True
        affect_subject : bool, optional
            If True, return the subject mapping, by default False
            
        Returns
        -------
        list
            List of pd.DataFrame with features grouped by label
        """
        
        features_grouping = {
            'Shame': features[features['label'] == 'Shame'].copy(),
            'Pride': features[features['label'] == 'Pride'].copy(),
            'Joy': features[features['label'] == 'Joy'].copy(),
            'Frustration': features[features['label'] == 'Frustration'].copy(),
        }

        for emotion in features_grouping.keys():
            rows_to_remove = list()
            for row in range(len(features_grouping[emotion])):
                for feature in features_grouping[emotion].iloc[row].index:
                    if feature != 'label':
                        if windows:
                            if type(features_grouping[emotion].iloc[row][feature]) == str:
                                if 'nan' in features_grouping[emotion].iloc[row][feature]:
                                    list_str = features_grouping[emotion].iloc[row][feature].replace('nan', '"nan"')
                                else:
                                    list_str = features_grouping[emotion].iloc[row][feature]
                                features_grouping[emotion].iloc[row][feature] = [float('nan') if x == 'nan' else x for x in eval(list_str)]

                            if clean:
                                if len(features_grouping[emotion].iloc[row][feature]) == 1:
                                    if int(row) not in rows_to_remove:
                                        rows_to_remove.append(int(row))
                if affect_subject:
                    features_grouping[emotion]['label'].iloc[row] = self.subject_mapping[row]
                    
                if clean and windows:
                    self._clean_features(features_grouping[emotion].iloc[row], replace_nan)

            if clean and windows:
                for row in rows_to_remove:
                    features_grouping[emotion] = features_grouping[emotion].drop(features_grouping[emotion].iloc[row].name) 
                    print(f'For subject {self.subject_mapping[row]}, {emotion} emotion prompt has been removed because it contains only one discrete time window')
            
            if affect_subject:
                features_grouping[emotion].rename(columns={'label': 'subject'}, inplace=True)

        return features_grouping

    def _clean_features(self, features_to_clean, replace=True):
        """
        Clean features by replacing all the nan values by the mean of non NaN values or removing them
        
        Parameters
        ----------
        features_to_clean : pd.DataFrame
            Dataframe with all the features to clean
        replace : bool, optional
            If True, replace the nan values by the mean of non NaN values, by default True
        """
        
        if replace:
            for feature in features_to_clean.index:
                if feature != 'label':
                    nan_index = [index for index, value in enumerate(features_to_clean[feature]) if math.isnan(value)]
                    if nan_index != []:
                        print(f'Cleaning feature {feature} by replacing NaN values by the mean across the values...')
                        mean_value = np.nanmean(features_to_clean[feature], axis=0)
                        features_to_clean[feature] = [mean_value if math.isnan(value) else value for value in features_to_clean[feature]]
        else:
            all_nan_indexes = []
            for feature in features_to_clean.index:
                if feature != 'label':
                    nan_index = [index for index, value in enumerate(features_to_clean[feature]) if math.isnan(value)]
                    if nan_index != []:
                        all_nan_indexes.append(nan_index)
                        print(f'Feature {feature} has {len(nan_index)} nan values')
            all_nan_indexes = np.unique([item for sublist in all_nan_indexes for item in sublist])

            if all_nan_indexes != []:
                for feature in features_to_clean.index:
                    if feature != 'label':
                        indexes = [index for index in all_nan_indexes if index < len(features_to_clean[feature])]
                        features_to_clean[feature] = np.delete(features_to_clean[feature], indexes)

                min_length = min([len(features_to_clean[feature]) for feature in features_to_clean.index if feature != 'label'])

                for feature in features_to_clean.index:
                    if feature != 'label':
                        if len(features_to_clean[feature]) > min_length:
                            features_to_clean[feature] = features_to_clean[feature][:min_length]
    
    def _filter_empatica(self, data, sample_rate, data_type):
        """
        Filter the Empatica data.

        Parameters
        ----------
        data : numpy.ndarray
            The data to filter.
        sample_rate : int
            The sample rate.
        data_type : str
            The data type.
        
        Returns
        -------
        numpy.ndarray
            The filtered data.
        """

        if data_type == 'acc':
            filtered_data = self._butter_lowpass_filter(data, 3, 15*int(sample_rate), order=5)
        elif data_type == 'bvp':
            filtered_data = median_filter(data, 15*int(sample_rate) + 1)
        elif data_type == 'eda':
            filtered_data = self._safe_savgol_filter(data, 15*int(sample_rate) + 1, 3)
        elif data_type == 'hr':
            filtered_data = median_filter(data, 15*int(sample_rate) + 1)
        elif data_type == 'temp':
            filtered_data = self._safe_savgol_filter(data, 15*int(sample_rate) + 1, 3)

        return filtered_data
    
    def _filter_audio(self, audio, sr):
        """
        Filter Audio Signal

        Parameters
        ----------
        audio : numpy.ndarray
            The audio signal.
        sr : int
            The sample rate.

        Returns
        -------
        numpy.ndarray
            The filtered audio signal.
        """

        audio_norm = librosa.util.normalize(np.array(audio))

        audio_nr = nr.reduce_noise(y=audio_norm, sr=sr)

        # Human voice range is typically between 300Hz and 3400Hz
        audio_filtered = self._bandpass_filter(audio_nr, lowcut=300, highcut=3400, sample_rate=sr)

        return audio_filtered
    
    def _bandpass_filter(self, data, lowcut, highcut, sample_rate, order=5):
        """
        Apply a bandpass filter to data.

        Parameters
        ----------
        data : np.ndarray
            The data to filter.
        lowcut : int
            The lowcut frequency.
        highcut : int
            The highcut frequency.
        sample_rate : int
            The sample rate.
        order : int
            The order of the filter.
        
        Returns
        -------
        np.ndarray
            The filtered data.
        """

        nyquist = 0.5 * sample_rate
        low = lowcut / nyquist
        high = highcut / nyquist
        b, a = butter(order, [low, high], btype='band')
        y = lfilter(b, a, data)

        return y
    
    def _butter_lowpass_filter(self, data, cutoff, sample_rate, order=5):
        """
        
        Parameters
        ----------
        data : np.ndarray
            The data to filter.
        cutoff : int
            The cutoff frequency.
        sample_rate : int
            The sample rate.
        order : int
            The order of the filter.
        
        Returns
        -------
        np.ndarray
            The filtered data.
        """

        nyquist = 0.5 * sample_rate
        normal_cutoff = cutoff / nyquist 
        b, a = butter(order, normal_cutoff, btype='low', analog=False)
        y = filtfilt(b, a, data)

        return y
    
    def _NormalizeAndSmooth_data(self, data, max_window_length=30, polyorder=3):
        """
        Normalize and smooth the data.

        Parameters
        ----------
        data : np.ndarray
            The data to normalize and smooth.
        max_window_length : int
            The maximum window length to use.
        polyorder : int
            The order of the polynomial to use.
        
        Returns
        -------
        np.ndarray
            The normalized and smoothed data.
        """

        data_smooth = self._safe_savgol_filter(data, max_window_length, polyorder)
        data_normalized = self.scaler.fit_transform(data_smooth.reshape(-1, 1)).flatten()

        return data_normalized

    def _safe_savgol_filter(self, data, max_window_length, polyorder):
        """
        Apply a Savitzky-Golay filter to data. This method ensures that the length of the data is not reduced by the filter.

        Parameters
        ----------
        data : np.ndarray
            The data to filter.
        max_window_length : int
            The maximum window length to use.
        polyorder : int
            The order of the polynomial to use.

        Returns
        -------
        np.ndarray
            The filtered data.
        """

        if len(data) < max_window_length:
            window_length = len(data) - 1 if len(data) % 2 == 0 else len(data)
        else:
            window_length = max_window_length

        polyorder = min(polyorder, window_length - 1)

        return savgol_filter(data, window_length, polyorder)

    def _get_timings(self, filepath, type='frontal'):
        """
        Get the timings for each emotion and for each subject.
        
        Parameters
        ----------
        filepath : str
            The path to the Excel file to load. 
        type : str
            The type of the recording (frontal or lateral).
        
        Returns
        -------
        emotion_timing : dict
            The timings for each emotion and for each subject.
        """

        df = pd.read_excel(filepath, skiprows=3, engine='openpyxl')

        frontal_rows = df[df['Record ID'].str.contains('frontal', case=False, na=False)].values.tolist()
        filtered_rows = df[df['Record ID'].str.contains(type, case=False, na=False)].values.tolist()

        emotion_timing = dict()
        clap_times = dict()

        for row, frontal_row in zip(filtered_rows, frontal_rows):
            id = int(row[0].split(" ", 1)[0])
            emotions = frontal_row[1].split(', ')
            emotions.append('Joy')
            
            for idx, item in enumerate(row[2:]):
                item = str(item)
                if 'f1' in item:
                    row[2+idx] = item.replace('f1 ', '')
                elif 'f2' in item or 'f3' in item or 'f4' in item or 'f5' in item:
                    time_match = re.search(r'(\d+):(\d+):(\d+)', item)

                    if time_match:
                        hours = int(time_match.group(1))
                        minutes = int(time_match.group(2))
                        seconds = int(time_match.group(3))
                        
                    if 'f2' in item:
                        updated_time = datetime(1, 1, 1, hours, minutes, seconds) + timedelta(minutes=12)
                        row[2+idx] = updated_time.strftime('%#H:%M:%S')
                    elif 'f3' in item:
                        updated_time = datetime(1, 1, 1, hours, minutes, seconds) + timedelta(minutes=24)
                        row[2+idx] = updated_time.strftime('%#H:%M:%S')
                    elif 'f4' in item:
                        updated_time = datetime(1, 1, 1, hours, minutes, seconds) + timedelta(minutes=36)
                        row[2+idx] = updated_time.strftime('%#H:%M:%S')
                    elif 'f5' in item:
                        updated_time = datetime(1, 1, 1, hours, minutes, seconds) + timedelta(minutes=48)
                        row[2+idx] = updated_time.strftime('%#H:%M:%S')
                        
            clap_time = str(row[2])
            start_cond_1 = str(row[3])
            end_cond_1 = str(row[4])
            start_cond_2 = str(row[5])
            end_cond_2 = str(row[6])
            start_cond_3 = str(row[7])
            end_cond_3 = str(row[8])
            start_cond_4 = str(row[9])
            end_cond_4 = str(row[10])
            
            emotion_timing[id] = {
                emotions[0]: (start_cond_1, end_cond_1),
                emotions[1]: (start_cond_2, end_cond_2),
                emotions[2]: (start_cond_3, end_cond_3),
                emotions[3]: (start_cond_4, end_cond_4)
            }

            clap_times[id] = clap_time
        
        return emotion_timing, clap_times
    
    def _construct_time(self, sample_rate, num_entries):
        """
        Construct a time array based on the sample rate and number of entries.
        
        Parameters
        ----------
        sample_rate : float
            The sample rate of the recording.
        num_entries : int
            The number of entries in the recording.
            
        Returns
        -------
        time : list
            The time array.
        """
        
        return [i / sample_rate for i in range(num_entries)]
    
    def _time_to_seconds(self,t):
        """
        Convert a time string to seconds.
        
        Parameters
        ----------
        t : str
            The time string to be converted.
            
        Returns
        -------
        seconds : int
            The time in seconds.
        """

        pt = datetime.strptime(t, '%H:%M:%S')

        return pt.hour * 3600 + pt.minute * 60 + pt.second
    
    def _seconds_to_time(self, seconds):
        """
        Convert seconds to a time string.
        
        Parameters
        ----------
        seconds : int
            The time in seconds to be converted.
            
        Returns
        -------
        t : str
            The time string in the format 'HH:MM:SS'.
        """

        td = timedelta(seconds=seconds)
        
        hours, remainder = divmod(td.seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        
        time_str = f'{hours:01}:{minutes:02}:{seconds:02}'
        
        return time_str

    def _timestamp_to_frame(self, timestamp, fps, shift=0):
        """
        Convert a timing to a frame number.

        Parameters
        ----------
        timestamp : str
            The timing to be converted.
        fps : int
            The FPS of the video.
        shift : int
            The shift to apply to the timestamp.

        Returns
        -------
        int
            The frame number.
        """

        total_seconds = self._time_to_seconds(timestamp)-shift

        return int(total_seconds * fps)

    def _convert_unix_timestamp_to_utc(self, unix_timestamp):
        
        """
        Convert a UNIX timestamp to UTC.
        
        Parameters
        ----------
        unix_timestamp : int
            The UNIX timestamp to be converted.
        
        Returns
        ------- 
        formatted_datetime : str
            The UTC time.
        """

        dt_object = datetime.utcfromtimestamp(unix_timestamp)
        formatted_datetime = dt_object.strftime('%Y-%m-%d %H:%M:%S')
        
        return formatted_datetime
    
    def _detect_clap_time(self, all_audio, sr, threshold = 0.85, webcam_video = False, path = None):
        """
        Detects the time of the first clap in the audio file.
        
        Parameters
        ----------
        all_audio : np.ndarray
            The audio data.
        sr : int
            The sampling rate.
        threshold : float
            The threshold for the onset strength envelope.
        webcam_video : bool
            Whether the file is a webcam video or not.
        path : str
            The path to the video file.
            
        Returns
        -------
        int
            The time of the first clap.
        """

        if webcam_video:
            audio_file_path = 'temp_extracted_webcam_audio.wav'

            video = VideoFileClip(path)

            end_time = 3*60 

            audio_clip = video.subclip(0, end_time).audio

            audio_clip.write_audiofile(audio_file_path, verbose=False, logger=None)

            audio, sr = librosa.load(audio_file_path, sr=11025)   
        else:
            audio = all_audio[:3 * 60 * sr]

        audio_env = np.abs(librosa.onset.onset_strength(y=audio, sr=sr))
        audio_env /= np.max(audio_env)

        clap_indices = np.nonzero(audio_env > threshold)[0]
        clap_times = librosa.frames_to_time(clap_indices, sr=sr)
        
        first_clap = int(clap_times[0])

        if webcam_video:
            os.remove(audio_file_path)

        return first_clap
    
    def _get_correct_timing(self, start, end):
        """
        Get the correct timing for the emotion with respect to the right video (f1, f2, ...).

        Parameters
        ----------
        start : str
            The start time of the emotion.
        end : str
            The end time of the emotion.
        
        Returns
        -------
        str
            The correct video file.
        str
            The start time according to the right video file.
        str
            The end time according to the right video file.
        """

        start_obj = datetime.strptime(start, "%H:%M:%S")
        end_obj = datetime.strptime(end, "%H:%M:%S")

        current_time_minutes = (start_obj - datetime.strptime("00:12:00", "%H:%M:%S")).total_seconds() / 60

        if current_time_minutes < 0:
            new_start = (start_obj - timedelta(minutes=0)).strftime("%H:%M:%S")
            new_end = (end_obj - timedelta(minutes=0)).strftime("%H:%M:%S")
            prefix = "f1"
        elif 0<current_time_minutes <12:
            new_start = (start_obj - timedelta(minutes=12)).strftime("%H:%M:%S")
            new_end = (end_obj - timedelta(minutes=12)).strftime("%H:%M:%S")
            prefix = "f2"
        elif 12<current_time_minutes < 24:
            new_start = (start_obj - timedelta(minutes=24)).strftime("%H:%M:%S")
            new_end = (end_obj - timedelta(minutes=24)).strftime("%H:%M:%S")
            prefix = "f3"
        elif 24<current_time_minutes < 36:
            new_start = (start_obj - timedelta(minutes=36)).strftime("%H:%M:%S")
            new_end = (end_obj - timedelta(minutes=36)).strftime("%H:%M:%S")
            prefix = "f4"
        elif 36<current_time_minutes < 48:
            new_start = (start_obj - timedelta(minutes=48)).strftime("%H:%M:%S")
            new_end = (end_obj - timedelta(minutes=48)).strftime("%H:%M:%S")
            prefix = "f5"

        new_end_obj = datetime.strptime(new_end, "%H:%M:%S")
        if (new_end_obj - datetime.strptime("00:12:00", "%H:%M:%S")).total_seconds() / 60 > 0:
            new_end = "00:12:00"

        return prefix, new_start, new_end
    
    def _equalize_emotion_segments(self, df):
        """
        Brings all the emotion segments to the same count

        Parameters
        ----------
        df : pd.DataFrame
            The dataframe to equalize.
        
        Returns
        ------- 
        pd.DataFrame
            The equalized dataframe.
        """

        emotion_counts = df['label'].value_counts()

        min_count = emotion_counts.min()

        resampled_df = df.groupby('label', group_keys=False).apply(lambda x: x.sample(n=min_count, random_state=42))

        resampled_df = resampled_df.sort_index()

        return resampled_df
    
    def get_info(self, df):
        """
        Get some information about the dataset

        Parameters
        ----------
        df : pandas.DataFrame
            The dataframe containing the dataset
        """

        print("Number of Shame segments: ", len(df[df['label'] == 'Shame']))
        print("Number of Frustration segments: ", len(df[df['label'] == 'Frustration']))
        print("Number of Joy segments: ", len(df[df['label'] == 'Joy']))
        print("Number of Pride segments: ", len(df[df['label'] == 'Pride']))
        print("Number of Neutral segments: ", len(df[df['label'] == 'Neutral']))
        print("")
        print("Features Set Shape: ", df.shape)
        print("")

    def get_feature_names(self, features, desired_measurement=['Empatica', 'Transcript', 'Audio', 'Webcam', 'GoPro', 'SRE', 'FaceReader']):
        """
        Extracts the feature names from the features dataframe and group them per datatype

        Parameters
        ----------
        features : pandas dataframe
            dataframe containing the features
        desired_measurement : list
            list containing the desired features
        
        Returns
        -------
        feature_names : dict
            dictionary containing the feature names for each feature group
        """

        if isinstance(desired_measurement, str):
            desired_measurement = [desired_measurement]
        
        feature_groups = {
            'Empatica': ['temp_mean', 'hr_range'],
            'Transcript': ['word-count', 'gunning-fog'],
            'Audio': ['mfccs_mean', 'crest-factor'],
            'Webcam': ['mouthCornerToMouthCorner_mean_dist', 'blink_rate'],
            'GoPro': ['shoulderDistance_mean', 'headInclination_var'],
            'SRE': ['sre-score', 'sre-score'],
            'FaceReader': ['neutral_mean', 'depth-position_var']
        }

        all_features_names = list(features.columns)
        feature_names = {}

        for group in feature_groups.keys():

            if group in desired_measurement:

                idx1 = all_features_names.index(feature_groups[group][0])
                idx2 = all_features_names.index(feature_groups[group][1])

                feature_names[group] = all_features_names[idx1:idx2+1]

        return feature_names
    
    def _extract_discrete_time_windows(self, data, nb_seconds):
        """
        Extract the discrete time windows for each emotion.

        Parameters
        ----------
        data : dict
            The dictionary containing the data.
        nb_seconds : int
            The number of seconds per window.

        Returns
        -------
        time_windows : dict
            The dictionary containing the time windows.
        """
        time_windows = {
            'Shame': {
                'Empatica': list(),
                'Audio': list(),
                'Webcam': list(),
                'GoPro': {
                    'Frontal': list(),
                    'Lateral': list()
                },
                'FaceReader': list()
            },
            'Frustration': {
                'Empatica': list(),
                'Audio': list(),
                'Webcam': list(),
                'GoPro': {
                    'Frontal': list(),
                    'Lateral': list()
                },
                'FaceReader': list()
            },
            'Pride': {
                'Empatica': list(),
                'Audio': list(),
                'Webcam': list(),
                'GoPro': {
                    'Frontal': list(),
                    'Lateral': list()
                },
                'FaceReader': list()
            },
            'Joy':{
                'Empatica': list(),
                'Audio': list(),
                'Webcam': list(),
                'GoPro': {
                    'Frontal': list(),
                    'Lateral': list()
                },
                'FaceReader': list()
            },
            'Neutral': {
                'Empatica': list(),
                'Audio': list(),
                'Webcam': list(),
                'GoPro': {
                    'Frontal': list(),
                    'Lateral': list()
                },
                'FaceReader': list()
            }
        }

        for rec in data.keys():
            if rec == 'Empatica':
                for emotion in data[rec]['eda']['data']['label'].unique():
                    d = data[rec]['eda']['data'][data[rec]['eda']['data']['label'] == emotion]
                    samples_per_window = int(nb_seconds * data[rec]['eda']['sample_rate'])
                    l = []
                    for start in range(0, len(d), samples_per_window):
                        end = start + samples_per_window

                        if end <= len(d):
                            l.append([start,end])
                    
                    time_windows[emotion][rec] = l
            elif rec == 'Audio':
                for emotion in data[rec]['data']['label'].unique():
                    d = data[rec]['data'][data[rec]['data']['label'] == emotion]
                    samples_per_window = int(nb_seconds * data[rec]['sample_rate'])
                    l = []
                    for start in range(0, len(d), samples_per_window):
                        end = start + samples_per_window

                        if end <= len(d):
                            l.append([start,end])   
                    
                    time_windows[emotion][rec] = l
            elif rec == 'Webcam':
                for emotion in data[rec]['label'].keys():
                    frames_per_window = int(nb_seconds * data[rec]['fps'])
                    start_frame, end_frame = data[rec]['label'][emotion]
                    l = []
                    for start in range(start_frame, end_frame, frames_per_window):
                        end = start + frames_per_window

                        if end <= end_frame:
                            l.append([start, end])
                    
                    time_windows[emotion][rec] = l
            elif rec == 'GoPro':
                for position in data[rec].keys():
                    for emotion in data[rec][position]['label'].keys():
                        frames_per_window = int(nb_seconds * data[rec][position]['fps'])
                        start_frame, end_frame = data[rec][position]['label'][emotion]
                        l = []
                        for start in range(start_frame, end_frame, frames_per_window):
                            end = start + frames_per_window

                            if end <= end_frame:
                                l.append([start, end])
                        
                        time_windows[emotion][rec][position] = l
            elif rec == 'FaceReader':
                for emotion in data[rec].keys():
                    sample_rate = 5
                    d = data[rec][emotion]['neutral']
                    samples_per_window = int(nb_seconds * sample_rate)
                    l = []
                    for start in range(0, len(d), samples_per_window):
                        end = start + samples_per_window

                        if end <= len(d):
                            l.append([start,end]) 
                    time_windows[emotion][rec] = l

        return time_windows

    def compare_nb_of_time_windows(self, data, desired_measurement, nb_seconds=10):
        """
        This function compares the number of time windows for each subject and each emotion.

        Parameters
        ----------
        data : dict
            Dictionary containing the data.
        desired_measurement : list
            List containing the desired features.
        nb_seconds : int
            The number of seconds per window.
        """

        for subject_id in data.keys():
            time_windows = self._extract_discrete_time_windows(data[subject_id], nb_seconds)      

            for emotion in ['Shame', 'Frustration', 'Pride', 'Joy']:
                basis = len(time_windows[emotion]['Empatica'])
                for rec in time_windows[emotion].keys():
                    if rec in desired_measurement:
                        if rec != 'Empatica':
                            if rec == 'GoPro':
                                for orientation in time_windows[emotion][rec].keys():
                                    if len(time_windows[emotion][rec][orientation]) != basis:
                                        print('ERROR: ', subject_id, emotion, rec, orientation)
                                        print('Length should be', basis, ' but is ', len(time_windows[emotion][rec][orientation]))
                                        print("--------------------")
                            else:
                                if len(time_windows[emotion][rec]) != basis:
                                    print('ERROR: ', subject_id, emotion, rec)
                                    print('Length should be', basis, ' but is ', len(time_windows[emotion][rec]))
                                    print("--------------------")
    
    def _log_msg(self, filename, msg=None):
        """
        Log a message in a file.

        Parameters
        ----------
        filename : str
            The name of the file.
        msg : str
            The message to log.
        """

        if filename is not None:
            with open(filename, "a") as file:
                if msg != None:
                    json.dump(msg, file)
                file.write("\n")

    def get_missing_labels(self, data, nb_emotions=5):
        """
        Check if there are missing labels in the data

        Parameters
        ----------
        data : dict
            Dictionary with the data
        """

        for subject_id in data.keys():
            if len(data[subject_id]['Empatica']['eda']['data']["label"].unique()) !=  nb_emotions:
                print('Empatica - Subject:', subject_id, data[subject_id]['Empatica']['eda']['data']['label'].unique())
            if len(data[subject_id]['Audio']['data']['label'].unique()) !=  nb_emotions:
                print('Audio - Subject:', subject_id, data[subject_id]['Audio']['data']['label'].unique())
            if len(data[subject_id]['Webcam']['label'].keys()) !=  nb_emotions:
                print('Webcam - Subject:', subject_id, data[subject_id]['Webcam']['label'].keys())
            if len(data[subject_id]['GoPro']['Frontal']['label'].keys()) !=  nb_emotions:
                print('GoPro Frontal - Subject:', subject_id, 'Frontal', len(data[subject_id]['GoPro']['Frontal']['label'].keys()), data[subject_id]['GoPro']['Frontal']['label'].keys())
            if len(data[subject_id]['GoPro']['Lateral']['label'].keys()) !=  nb_emotions:
                print(subject_id, 'Lateral', len(data[subject_id]['GoPro']['Lateral']['label'].keys()), data[subject_id]['GoPro']['Lateral']['label'].keys())

    def _convert_keys_to_string(self, dictionary):
        """
        Recursively converts dictionary keys to strings to enable the saving in JSON.

        Parameters
        ----------
        dictionary : dict
            The dictionary to convert.

        Returns
        -------
        dict
            The dictionary with the converted keys.
        """
        
        if isinstance(dictionary, dict):
            return {str(key): self._convert_keys_to_string(value) for key, value in dictionary.items()}
        elif isinstance(dictionary, list):
            return [self._convert_keys_to_string(element) for element in dictionary]
        else:
            return dictionary
        
    def _group_desired_measurement_for_correlation(self, desired_measurement):
        """
        Parameters
        ----------
        desired_measurement : list
            The desired features to study
        
        Returns
        -------
        correlation_to_study : dict
            The dictionary containing the features to study
        
        """
        correlation_to_study = dict()
        for feature1 in desired_measurement:
            for feature2 in desired_measurement:
                if [feature1, feature2] in correlation_to_study.values() or [feature2, feature1] in correlation_to_study.values():
                    continue
                else:
                    correlation_to_study[f"{feature1.lower()}_{feature2.lower()}"] = [feature1, feature2]

        return correlation_to_study
    

    def _get_subject_mapping(self):
        """
        Returns a dictionary mapping the subject index to the subject id.
        
        Returns
        -------
        subject_mapping : dict
            Dictionary mapping the subject index to the subject id.
        """

        all_subjects = [1001, 1003, 1007, 1008, 1009, 1011, 1013, 1015, 1016, 1017, 1020, 1021, 1022, 1023, 1024, 1025, 1026, 1029, 1031, 1032, 1033, 1034, 1036, 1037, 1039, 1040, 1041, 1042]
        subject_mapping = dict()
        idx = 0

        for subject in all_subjects:
            if subject in [1002, 1005, 1009, 1011, 1016, 1021, 1023, 1029, 1034, 1036]:
                continue
            subject_mapping[idx] = subject
            idx += 1

        return subject_mapping
    
    def get_population_statistics(self, file_path):
        """
        This function returns a dictionary with the population statistics

        Parameters:
        -----------
        file_path: string
            The path to the file containing the data
        
        Returns:
        --------
        statistics: dictionary
            A dictionary containing the population statistics
        """

        data = pd.read_csv(file_path)

        statistics = dict()

        subjects_id = data['Record ID'].tolist()
        subjects_gender = data['Which of the following best describes you? '].tolist()
        subjects_age = data['How old are you?'].tolist()
        subject_ethnicity = data['What is your primary ethnic identification? '].tolist()
        subject_orientation = data['Do you identify as: '].tolist()
        subject_date_of_diagnosis = data['In what year were you diagnosed with HIV?'].tolist()

        for idx, subject_id in enumerate(subjects_id):
            if not math.isnan(subjects_age[idx]):
                statistics[subject_id] = {
                    'Age': int(subjects_age[idx]),
                    'Gender': subjects_gender[idx],
                    'Ethnicity': subject_ethnicity[idx],
                    'Orientation': subject_orientation[idx],
                    'Date of Diagnosis': int(subject_date_of_diagnosis[idx],)
                }
            if subject_id == 1041:
                statistics[subject_id] = {
                    'Age': int(np.mean([int(age) for age in subjects_age if not math.isnan(age)])),
                    'Gender': subjects_gender[idx],
                    'Ethnicity': subject_ethnicity[idx],
                    'Orientation': subject_orientation[idx],
                    'Date of Diagnosis': int(subject_date_of_diagnosis[idx],)
                }

        return statistics
    
    def git_commit_push(self, repo_path, commit_message, branch_name="master"):
        """
        Commit and push the changes to the git repository.

        Parameters
        ----------
        repo_path : str
            The path to the repository.
        commit_message : str
            The commit message.
        branch_name : str
            The name of the branch.
        """
        try:
            repo = Repo(repo_path)
            
            if not repo.is_dirty(untracked_files=True):
                print("No changes to commit.")
                return

            origin = repo.remote(name='origin')
            origin.pull(branch_name)

            repo.git.add(all=True)

            repo.index.commit(commit_message)

            origin.push(refspec=f'{branch_name}:{branch_name}')
        
            print("Changes committed and pushed successfully.")
        except GitCommandError as e:
            print(f"Git command error: {e}")

    def extract_sre_stats(self, filename, plot=False):
        """
        Extracts and prints descriptive statistics for the SRE dataset and plots histograms for each emotion.

        Parameters:
        ----------
        filename : str
            The filename of the SRE dataset.
        plot : bool, optional
            Whether to plot histograms for each emotion. Default is False.
        """
        sre_features = pd.read_csv(filename)

        shame_scores = list()
        frustration_scores = list()
        joy_scores = list()
        pride_scores = list()

        for _, row in sre_features.iterrows():
            label = row['label']
            score = row['sre_score']
            if label == 'Shame':
                shame_scores.append(score)
            elif label == 'Frustration':
                frustration_scores.append(score)
            elif label == 'Joy':
                joy_scores.append(score)
            elif label == 'Pride':
                pride_scores.append(score)

        data = {
            'Subject': [1001, 1003, 1007, 1008, 1013, 1015, 1016, 1017, 1020, 1022, 1024, 1025, 1026, 1031, 1032, 1033, 1036, 1037, 1039, 1040, 1041, 1042],
            'Shame': shame_scores,
            'Frustration': frustration_scores,
            'Joy': joy_scores,
            'Pride': pride_scores
        }

        df = pd.DataFrame(data)

        desc_stats = df[['Shame', 'Frustration', 'Joy', 'Pride']].describe()
        print("Descriptive Statistics:\n", desc_stats)

        if plot:
            emotion_columns = df.columns[1:]
            fig, axs = plt.subplots(2, 2, figsize=(10, 8))

            for i, emotion in enumerate(emotion_columns):
                row = i // 2
                col = i % 2
                ax = axs[row, col]
                sns.histplot(df[emotion], kde=True, ax=ax, bins=10, color='red')
                ax.set_title(f'{emotion} Scores Distribution')
                ax.set_xlabel('Scores')
                ax.set_ylabel('Frequency')

            plt.tight_layout()
            plt.show()

            plt.figure(figsize=(10, 6))

            sns.boxplot(data=df.drop('Subject', axis=1), palette=['red', 'red', 'red', 'red'])

            plt.title('Emotion Scores Distribution')
            plt.ylabel('Scores')
            plt.show()

    def select_features_from_subjects(self, features, selected_subjects, include_neutral = False):
        """
        Filter the feature dataframe based on the selected subjects

        Parameters
        ----------
        features : pd.DataFrame
            The dataframe containing the features
        selected_subjects : list
            The list of selected subjects
        include_neutral : bool, optional
            Whether to include the neutral emotion. Default is False.

        Returns
        -------
        pd.DataFrame
            The filtered dataframe
        """
        
        rows_to_keep = list()
        for emotion in features['label'].unique():
            grouped_features = features[features['label']==emotion]
            rows = [idx for row, idx in enumerate(grouped_features.index) if self.subject_mapping[row] in selected_subjects]
            rows_to_keep.extend(rows)
        
        features = features.loc[rows_to_keep]

        if not include_neutral:
            features = features[features['label'] != 'Neutral']

        return features.sort_index()

    def default_converter(self, obj):
        """
        Convert an object to a JSON serializable format.

        Parameters
        ----------
        obj : object
            The object to convert.
        
        Returns
        -------
        object
            The JSON serializable object.
        """
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, pd.Series):
            return obj.to_dict() 
        else:
            raise TypeError(f"Object of type {obj.__class__.__name__} is not JSON serializable")
    
    def draw_ellipse(self, position, covariance, ax=None, **kwargs):
        """
        Draw an ellipse with a given position and covariance.
        
        Parameters
        ----------
        position : list
            The position of the ellipse.
        covariance : list
            The covariance of the ellipse.
        ax : matplotlib.axes._subplots.AxesSubplot, optional
            The axes to draw the ellipse on. Default is None.
        kwargs : dict
            Additional keyword arguments to pass to the ellipse.
        
        Returns
        -------
        matplotlib.patches.Ellipse
            The ellipse.
        """

        if ax is None:
            ax = plt.gca()
            
        eigenvals, eigenvecs = np.linalg.eigh(covariance)
        order = eigenvals.argsort()[::-1]
        eigenvals, eigenvecs = eigenvals[order], eigenvecs[:, order]

        angle = np.degrees(np.arctan2(*eigenvecs[:, 0][::-1]))
        width, height = 2 * np.sqrt(eigenvals)
        ellipse = Ellipse(xy=position, width=width, height=height, angle=angle, **kwargs)
        
        ax.add_artist(ellipse)

        return ellipse
    
    